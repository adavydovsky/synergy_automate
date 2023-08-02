import os

import pandas as pd
import numpy as np
import openpyxl as ox
import re

from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

from datetime import datetime, timedelta

import warnings
warnings.filterwarnings("ignore")

import string

import pyodbc

#------------------------------------------------------------
#now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Обновление отчётов по запланированным оплатам и прослушке...")

struktura = pd.read_excel(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Харитуня\Шаблоны\Сотрудники.xlsx', sheet_name='Структура')
status_lead = pd.read_excel(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Харитуня\Шаблоны\Сотрудники.xlsx', sheet_name='статусы')
all = pd.read_excel(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Харитуня\Таблица по сотрудникам КД1.xlsx', sheet_name='Все')

#------------------------------------------------------------

conn = pyodbc.connect(r'Driver={SQL Server};Server=MSK1-BIDB01;Database=DWH;Trusted_Connection=yes;')
cursor = conn.cursor()

tdata = f'''

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

declare @date date = getdate()
declare @month int = month(getdate())
declare @endate date = DATEFROMPARTS (2024, 01, 31);

SELECT DISTINCT
      COALESCE(D.DATE_CREATE, ASS_AR.Data, R.DATE_CREATE)  as "Дата создания", -- DATE_CREATE,
      D.CLOSE_DATE  as "Дата завершения",
      --cp.[NAME] "Продукт сделки",
      CAST(D.AMOUNT / COUNT(*) OVER (PARTITION BY D.ID_DEAL)  as decimal(33,3)) "Сумма",
      concat(cc.[FAMILY]+' ',cc.[NAME]+' ',cc.[SURNAME]) as "Контакт",
      D.[PROBABILITY] "Вероятность",
      sd.NAME as "Стадия сделки",
      R.FEATURES_1 "Источник",
      CAST(ASS_P.[QUANTITY] AS INT) AS [Количество продукта],
      D.AMOUNT "СуммаСделки",       
      concat(e.[LAST_NAME]+' ',e.[NAME]+' ',e.[SECOND_NAME]) as "Ответственный",    
      org.full_NAME "Структурное подразделение",
      D.CODE as "ID сделки из Битрикс",
      uts_D.[UF_CRM_AUDITION_ROP] "Прослушано РОП"
FROM [DWH].[dbo].DIC_DEAL D
          LEFT JOIN [DWH].[dbo].[DIC_EMPLOYEES] E ON D.ID_EMPLOYEES = E.ID_EMPLOYEES
          LEFT JOIN [DWH].[dbo].[ASS_EMPLOYEE_AND_ORGSTRUCTURE] ASS_OS ON E.ID_EMPLOYEES = ASS_OS.ID_EMPLOYEES
          LEFT JOIN [DWH].[dbo].[v_DIC_ORGSTRUCTURE] org on org.ID_ORGSTRUCTURE = ASS_OS.ID_ORGSTRUCTURE
          --подсоединение заявок
          LEFT JOIN [DWH].[dbo].ASS_REQUEST_DEAL ASS_RD ON ASS_RD.ID_DEAL = D.ID_DEAL
          LEFT JOIN [DWH].[dbo].[DIC_REQUEST] R ON ASS_RD.ID_REQUEST = R.ID_REQUEST
          --объявления
          LEFT JOIN [DWH].[dbo].ASS_AD_REQUEST ASS_AR ON ASS_AR.[ID_TYPE_ASS_AD_REQUEST] = 0   --Связь с учетом всех коректировок
                                         AND ASS_AR.ID_REQUEST = R.ID_REQUEST   
          LEFT JOIN [DWH].[dbo].DIC_AD AD ON AD.ID_AD = ASS_AR.ID_AD
          LEFT JOIN [DWH].[dbo].DIC_AD_GROUP AG ON AD.ID_AD_GROUP = AG.ID_AD_GROUP
          LEFT JOIN [DWH].[dbo].DIC_AD_CAMPAIGN AC ON AG.ID_AD_CAMPAIGN = AC.ID_AD_CAMPAIGN
  
          LEFT JOIN [DWH].[AdWrapper].[ASS_MARKETOLOG_CAMPAIGN_AdWrapper] ASS_MC ON ASS_MC.ID_AD_CAMPAIGN_DWH = AC.ID_AD_CAMPAIGN AND D.DATE_CREATE BETWEEN ASS_MC.DATE_START AND ASS_MC.DATE_FINISH
          LEFT JOIN [DWH].[AdWrapper].[ASS_AD_PROJECT_AdWrapper] ASS_PC ON ASS_PC.[ID_AD_CAMPAIGN_DWH] = AC.ID_AD_CAMPAIGN AND D.DATE_CREATE BETWEEN ASS_PC.DATE_START AND ASS_PC.DATE_FINISH

          --подсоединение инвойсов к сделке
          LEFT JOIN [DWH].[dbo].ASS_DEAL_INVOICE ASS_I_D ON ASS_I_D.ID_DEAL = D.ID_DEAL
          LEFT JOIN [DWH].[dbo].DIC_INVOICE I_D ON ASS_I_D.ID_INVOICE = I_D.ID_INVOICE
          --каталог продуктов сделки
          LEFT JOIN [DWH].[dbo].ASS_DEAL_CATALOG_PRODUCT ASS_P ON D.ID_DEAL = ASS_P.ID_DEAL
          LEFT JOIN [DWH].[dbo].[DIC_CATALOG_PRODUCT] cp on cp.ID_CATALOG_PRODUCT = ASS_P.ID_CATALOG_PRODUCT
          -- статус сделки
          left join [DWH].[dbo].[DIC_STATUS_DEAL] sd on sd.ID_STATUS_DEAL = d.ID_STATUS_DEAL
          --contact
          left join [DWH].[dbo].[DIC_CLIENT_CRM] cc on cc.[ID_CLIENT_CRM] = d.[ID_CLIENT_CRM]
          left join [DWH].[dbo].[DIC_CLIENT_CRM_PHONE] PhHome ON PhHome.[ID_CLIENT_CRM]=d.[ID_CLIENT_CRM] AND PhHome.[ID_TYPE_CLIENT_CRM_PHONE]=1   --  Домашний
          left join [DWH].[dbo].[DIC_CLIENT_CRM_PHONE] PhMob ON PhMob.[ID_CLIENT_CRM]=d.[ID_CLIENT_CRM] AND PhMob.[ID_TYPE_CLIENT_CRM_PHONE]=2      --Мобильный
          left join [DWH].[dbo].[DIC_CLIENT_CRM_PHONE] PhJob ON PhJob.[ID_CLIENT_CRM]=d.[ID_CLIENT_CRM] AND PhJob.[ID_TYPE_CLIENT_CRM_PHONE]=5      --Рабочий
          left join [DWH].[dbo].[DIC_CLIENT_CRM_PHONE] PhOst ON PhOst.[ID_CLIENT_CRM]=d.[ID_CLIENT_CRM] AND PhOst.[ID_TYPE_CLIENT_CRM_PHONE]=3      --Прочий
          -- group product deal
          left join [DWH].[dbo].[DIC_GROUP_PRODUCT_DEAL] dgp on dgp.[ID_GROUP_PRODUCT_DEAL] = d.[GROUP_PRODUCT_ID]
         
          -- расширение свойств сделок
          left join [DWH].[dbo].[DIC_DEAL_STAT] utsD ON D.[ID_DEAL]=utsD.[ID_DEAL]
          LEFT JOIN [stage].[CRM_b_uts_crm_deal] UTS_D ON UTS_D.VALUE_ID = D.code
  
WHERE
  D.ID_DEAL <> - 1
    --and  cast(COALESCE(D.DATE_CREATE, ASS_AR.Data, R.DATE_CREATE)  as date) between '2018-01-01' and '2018-12-31' -- фильтр по дате создания
    and  D.CLOSE_DATE  between @date and @endate -- фильтр по дате завершения 
    and org.full_NAME like '%Коммерческий департамент (КМВ)\%'
    --and  cc.BIRTHDATE  between '1990-01-01' and '1999-12-31' -- фильтр по  дате рождения
    --КАК ВАРИАНТ  and   YEAR(cc.BIRTHDATE)  between 1990 AND 1999 -- фильтр по  ГОДУ даты рождения  
'''


Employees = f'''

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

SELECT [EMPLOYEES] as "Менеджер"
      ,[LOG]
      ,[SP]

      ,CASE

       WHEN KC like '%4%' THEN 'КЦ 4'
       WHEN KC like '%3%' THEN 'КЦ 3'
       WHEN KC like '%центр 3%' THEN 'КЦ 3'
       WHEN KC like '%продаж 18%' THEN 'ОП 18'
             
       ELSE KC
  END
  as "КЦ"

         ,CASE

       WHEN [KC] like '%4%' THEN SUBSTRING([ОП],CHARINDEX(' ',[ОП])+1,LEN([ОП])) + ' ВР'
       WHEN ([ОП] in ('ОП 9')) and ([GP] in ('.1')) THEN [ОП] + [GP]
       WHEN ([ОП] in ('ОП 2')) and ([GP] in ('.1')) THEN 'ОП 8'
       WHEN ([ОП] in ('ОП 10')) and ([GP] in ('.2')) THEN [ОП] + [GP]
       WHEN ([ОП] in ('ОП 5')) and ([GP] in ('.1')) THEN [ОП] + [GP]
       WHEN ([ОП] in ('3 ЯР')) and ([GP] in ('.1')) THEN '3.1 ЯР'
       WHEN ([KC] in ('Отдел прямых продаж'))  THEN 'ОПП'
           WHEN [KC] like '%продаж 18%'  THEN 'ОП 18'  
       ELSE [ОП]
  END
  as "ОП"
      
               ,[PHONE]
      ,[STATUS]
      ,[ID_EMPL]
      ,[ID_ORG]

  FROM [DWH].[dbo].[KHTS_EMPL]
  where  
  [SP] in ('КД\')

'''

next = f'''

declare @startdate date = DATEFROMPARTS(2017, 01, 01)
declare @enddate date = getdate()-1
--declare @month int = month(getdate())


SELECT DISTINCT
      COALESCE(D.DATE_CREATE, ASS_AR.Data, R.DATE_CREATE)  as "Дата создания" -- DATE_CREATE,
      ,D.CLOSE_DATE  as "Дата завершения"
      --cp.[NAME] "Продукт сделки",
      ,concat(cc.[FAMILY]+' ',cc.[NAME]+' ',cc.[SURNAME]) as "Контакт"
      ,isnull(D.[PROBABILITY],0) "Вероятность"
      --,sd.NAME as "Стадия сделки"
      ,SUBSTRING(sd.[NAME],0,PATINDEX('% / %',sd.[NAME]))   as "Статус"
      --,CAST(ASS_P.[QUANTITY] AS INT) AS [Количество продукта]
     -- ,D.AMOUNT "СуммаСделки"
      ,isnull(IIF (ASS_P.[QUANTITY] = 1,CAST(D.AMOUNT as decimal(33)), CAST((D.AMOUNT - (D.AMOUNT / ASS_P.[QUANTITY]))*0.5 + (D.AMOUNT / ASS_P.[QUANTITY])  as decimal(33))),0) AS TOTAL_AMOUNT       
      ,concat(e.[LAST_NAME]+' ',e.[NAME]+' ',e.[SECOND_NAME]) as "Ответственный"
      ,iif(E2.[KC] like '%Ярославль%',iif(E2.[KC] like '%Ярославль%',SUBSTRING (E2.[ОП], 4,1),E2.[ОП])+' '+ iif(E2.[ОП] like '%Ярославль%',replace(E2.[ОП],SUBSTRING (E2.[ОП], 1,4),'ЯР'),E2.[ОП]),E2.[ОП]) as "ОП"
      ,replace(E2.[KC],' Ярославль','') as "КЦ" 
      --org.full_NAME "Структурное подразделение",
     , D.CODE as "ID сделки из Битрикс"
      ,isnull(utsD.[AUDITION_ROP],0) "Прослушано РОП"

FROM [DWH].[dbo].DIC_DEAL D

          LEFT JOIN [DWH].[dbo].[KHTS_EMPL] E2 on E2.[ID_EMPL] = D.ID_EMPLOYEES
          LEFT JOIN [DWH].[dbo].[DIC_EMPLOYEES] E ON D.ID_EMPLOYEES = E.ID_EMPLOYEES
          LEFT JOIN [DWH].[dbo].[ASS_EMPLOYEE_AND_ORGSTRUCTURE] ASS_OS ON E.ID_EMPLOYEES = ASS_OS.ID_EMPLOYEES
          LEFT JOIN [DWH].[dbo].[v_DIC_ORGSTRUCTURE] org on org.ID_ORGSTRUCTURE = ASS_OS.ID_ORGSTRUCTURE
          --подсоединение заявок
          LEFT JOIN [DWH].[dbo].ASS_REQUEST_DEAL ASS_RD ON ASS_RD.ID_DEAL = D.ID_DEAL
          LEFT JOIN [DWH].[dbo].[DIC_REQUEST] R ON ASS_RD.ID_REQUEST = R.ID_REQUEST
          --объявления
          LEFT JOIN [DWH].[dbo].ASS_AD_REQUEST ASS_AR ON ASS_AR.[ID_TYPE_ASS_AD_REQUEST] = 0   --Связь с учетом всех коректировок
                                         AND ASS_AR.ID_REQUEST = R.ID_REQUEST   
          LEFT JOIN [DWH].[dbo].DIC_AD AD ON AD.ID_AD = ASS_AR.ID_AD
          LEFT JOIN [DWH].[dbo].DIC_AD_GROUP AG ON AD.ID_AD_GROUP = AG.ID_AD_GROUP
          LEFT JOIN [DWH].[dbo].DIC_AD_CAMPAIGN AC ON AG.ID_AD_CAMPAIGN = AC.ID_AD_CAMPAIGN
  
          LEFT JOIN [DWH].[AdWrapper].[ASS_MARKETOLOG_CAMPAIGN_AdWrapper] ASS_MC ON ASS_MC.ID_AD_CAMPAIGN_DWH = AC.ID_AD_CAMPAIGN AND D.DATE_CREATE BETWEEN ASS_MC.DATE_START AND ASS_MC.DATE_FINISH
          LEFT JOIN [DWH].[AdWrapper].[ASS_AD_PROJECT_AdWrapper] ASS_PC ON ASS_PC.[ID_AD_CAMPAIGN_DWH] = AC.ID_AD_CAMPAIGN AND D.DATE_CREATE BETWEEN ASS_PC.DATE_START AND ASS_PC.DATE_FINISH

          --подсоединение инвойсов к сделке
          LEFT JOIN [DWH].[dbo].ASS_DEAL_INVOICE ASS_I_D ON ASS_I_D.ID_DEAL = D.ID_DEAL
          LEFT JOIN [DWH].[dbo].DIC_INVOICE I_D ON ASS_I_D.ID_INVOICE = I_D.ID_INVOICE
          --каталог продуктов сделки
          LEFT JOIN [DWH].[dbo].ASS_DEAL_CATALOG_PRODUCT ASS_P ON D.ID_DEAL = ASS_P.ID_DEAL
          LEFT JOIN [DWH].[dbo].[DIC_CATALOG_PRODUCT] cp on cp.ID_CATALOG_PRODUCT = ASS_P.ID_CATALOG_PRODUCT
          -- статус сделки
          left join [DWH].[dbo].[DIC_STATUS_DEAL] sd on sd.ID_STATUS_DEAL = d.ID_STATUS_DEAL
          --contact
          left join [DWH].[dbo].[DIC_CLIENT_CRM] cc on cc.[ID_CLIENT_CRM] = d.[ID_CLIENT_CRM]
          left join [DWH].[dbo].[DIC_CLIENT_CRM_PHONE] PhHome ON PhHome.[ID_CLIENT_CRM]=d.[ID_CLIENT_CRM] AND PhHome.[ID_TYPE_CLIENT_CRM_PHONE]=1   --  Домашний
          left join [DWH].[dbo].[DIC_CLIENT_CRM_PHONE] PhMob ON PhMob.[ID_CLIENT_CRM]=d.[ID_CLIENT_CRM] AND PhMob.[ID_TYPE_CLIENT_CRM_PHONE]=2      --Мобильный
          left join [DWH].[dbo].[DIC_CLIENT_CRM_PHONE] PhJob ON PhJob.[ID_CLIENT_CRM]=d.[ID_CLIENT_CRM] AND PhJob.[ID_TYPE_CLIENT_CRM_PHONE]=5      --Рабочий
          left join [DWH].[dbo].[DIC_CLIENT_CRM_PHONE] PhOst ON PhOst.[ID_CLIENT_CRM]=d.[ID_CLIENT_CRM] AND PhOst.[ID_TYPE_CLIENT_CRM_PHONE]=3      --Прочий
          -- group product deal
          left join [DWH].[dbo].[DIC_GROUP_PRODUCT_DEAL] dgp on dgp.[ID_GROUP_PRODUCT_DEAL] = d.[GROUP_PRODUCT_ID]
         
          -- расширение свойств сделок
          left join [DWH].[dbo].[DIC_DEAL_STAT] utsD ON D.[ID_DEAL]=utsD.[ID_DEAL]
  
WHERE
  D.ID_DEAL <> - 1
    --and  cast(COALESCE(D.DATE_CREATE, ASS_AR.Data, R.DATE_CREATE)  as date) between '2018-01-01' and '2018-12-31' -- фильтр по дате создания
    and  D.CLOSE_DATE  between @startdate and @enddate -- фильтр по дате завершения 
    and org.full_NAME like '%Коммерческий департамент (КМВ)\%'
    and ASS_P.[QUANTITY] <> 0
    --and D.[PROBABILITY] in ('1','2')
    --and  cc.BIRTHDATE  between '1990-01-01' and '1999-12-31' -- фильтр по  дате рождения
    --КАК ВАРИАНТ  and   YEAR(cc.BIRTHDATE)  between 1990 AND 1999 -- фильтр по  ГОДУ даты рождения  ия  
'''

tdata = pd.read_sql_query(tdata, conn)
Employees = pd.read_sql_query(Employees, conn)
next = pd.read_sql_query(next, conn)

#------------------------------------------------------------

def create_list(array, x, y):
    result = []
    found_x = False
    for index in array:
        if found_x:
            result.append(index)
            if index.value == y:
                break
        elif index.value == x:
            found_x = True
            result.append(index)
    return result

#------------------------------------------------------------

save = pd.read_excel(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\Отчёт по запланированным оплатам.xlsx', 
                     sheet_name='Отчёт'
                    )

save = save.drop(save.columns[0], axis=1)
save = save.loc[save['Отчет № 2.3'] == "Итого"]
save = save.reset_index(drop=True).reset_index()
save['index'] = save['index'].astype(str)
save['Итого'] = save['Отчет № 2.3'] + "_" + save['index']
save = save.drop(save.columns[0], axis=1)
column_to_move = save.pop("Итого")
save.insert(0, "Итого", column_to_move)

#------------------------------------------------------------

status_lead.columns = ['Статус сделки', 'один', 'два', 'три', 'четыре', 'пять', 'шесть', 'семь', 'восемь']
status_lead['два'] = status_lead['два'].fillna(0)
status_lead = status_lead.query('два != 0')
status_lead= status_lead.drop(['один', 'три', 'четыре', 'шесть'], axis=1)

#------------------------------------------------------------

struktura = struktura[['Сотрудник', 'ФИО', 'ДП', 'ОП', 'Логин', 'КЦ']]
struktura['ДП'] = struktura['ДП'].fillna(0)
struktura = struktura.query('ДП != 0 & Сотрудник != ""')
struktura = struktura.rename(columns={'Сотрудник': 'Менеджер'})

#------------------------------------------------------------
all['Пусто'] = "пусто"
all= all[['Менеджер', 'Пусто']]

#------------------------------------------------------------

tdata = tdata.query('Сумма != 0')
tdata['Ответственный'] = tdata['Ответственный'].replace('Ким Милена ', 'Ким Милена')
tdata['Ответственный'] = tdata['Ответственный'].replace('Мурзакулова Альбина ', 'Мурзакулова Альбина')
tdata['Ответственный'] = tdata['Ответственный'].replace('Скорогодская Валерия ', 'Скорогодская Валерия')
tdata['Ответственный'] = tdata['Ответственный'].replace('Юрченко Влада ', 'Юрченко Влада')

tdata = tdata.drop(['Сумма', 'Дата создания', 'Структурное подразделение'], axis=1)

tdata = tdata.rename(columns={'СуммаСделки': 'Сумма', 
                              'Стадия сделки': 'Статус сделки',
                              'Контакт': 'ФИО клиента',
                              'Дата завершения': 'Дата закрытия',
                              'Вероятность': 'Вероятность_оплаты',
                              'Ответственный': 'Менеджер',
                              'Количество продукта': 'Семестры'
                              })

#------------------------------------------------------------

tdata['Семестры'] = tdata['Семестры'].fillna(0)
tdata = tdata.loc[tdata['Семестры'] != 0]

tdata = tdata.merge(struktura, how='left', on='Менеджер')
tdata[['Статус сделки', 'del']]= tdata['Статус сделки'].str.split(' / ', expand=True)
tdata = tdata.drop('del', axis=1)

tdata = tdata.merge(status_lead, how='left', on='Статус сделки')
tdata = tdata.drop(['пять', 'семь', 'восемь'], axis=1)
tdata = tdata.rename(columns={'два': 'status'})

#------------------------------------------------------------

tdata = tdata.query('status == "В работе" | status == "Договор" | status == "Назначена ЛВ" | status == "Проведена ЛВ"')
tdata['1 сем'] = tdata['Сумма']/tdata['Семестры']
tdata['Чистая выручка'] = ((tdata['Сумма']-tdata['1 сем'])*0.5) + tdata['1 сем']
tdata['Прослушано РОП'] = tdata['Прослушано РОП'].fillna(2)
tdata['Прослушано РОП'] = tdata['Прослушано РОП'].astype(int)
tdata['Прослушано РОП'] = tdata['Прослушано РОП'].replace(2, 'нет')
tdata['Прослушано РОП'] = tdata['Прослушано РОП'].replace(0, 'нет')
tdata['Прослушано РОП'] = tdata['Прослушано РОП'].replace(1, 'да')

tdata['Дата закрытия'] = pd.to_datetime(tdata['Дата закрытия'])
tdata['месяц'] = tdata['Дата закрытия'].dt.month

tdata['месяц'] = tdata['месяц'].replace(1, 'Январь')
tdata['месяц'] = tdata['месяц'].replace(2, 'Февраль')
tdata['месяц'] = tdata['месяц'].replace(3, 'Март')
tdata['месяц'] = tdata['месяц'].replace(4, 'Апрель')
tdata['месяц'] = tdata['месяц'].replace(5, 'Май')
tdata['месяц'] = tdata['месяц'].replace(6, 'Июнь')
tdata['месяц'] = tdata['месяц'].replace(7, 'Июль')
tdata['месяц'] = tdata['месяц'].replace(8, 'Август')
tdata['месяц'] = tdata['месяц'].replace(9, 'Сентябрь')
tdata['месяц'] = tdata['месяц'].replace(10, 'Октябрь')
tdata['месяц'] = tdata['месяц'].replace(11, 'Ноябрь')
tdata['месяц'] = tdata['месяц'].replace(12, 'Декабрь')

#------------------------------------------------------------

tdata['Вероятность_оплаты'] = tdata['Вероятность_оплаты'].fillna(6)
tdata['Вероятность_оплаты'] = tdata['Вероятность_оплаты'].astype(int)
tdata = tdata.query('Вероятность_оплаты == 1 | Вероятность_оплаты == 2 | Вероятность_оплаты == 3 | Вероятность_оплаты == 4 | Вероятность_оплаты == 5 | Вероятность_оплаты == 0')
tdata[['Источник', 'ОП', 'КЦ']] = tdata[['Источник', 'ОП', 'КЦ']].fillna('нет')

tdata = tdata.merge(all, how='left', on='Менеджер')

tdata = tdata.rename(columns={'ОП': 'ОП_1', 'КЦ' : 'КЦ_1'})

#------------------------------------------------------------

def alert(row):
    if row['Вероятность_оплаты'] == 1:
        return row['Чистая выручка']
    elif row['Вероятность_оплаты'] == 2:
        return row['Чистая выручка']
    
tdata['Чистая'] = tdata.apply(alert, axis=1)  
tdata['Чистая'] = tdata['Чистая'].fillna(0)
tdata['Чистая'].value_counts()

tdata = tdata.merge(Employees, how='left', on='Менеджер')

tdata=tdata[['Статус сделки', 'ФИО клиента', 'Дата закрытия', 'Источник', 'Вероятность_оплаты',
       'Менеджер', 'Сумма', 'Семестры', 'ID сделки из Битрикс', 'status',
       'Прослушано РОП', 'ФИО', 'Логин', '1 сем', 'Чистая выручка', 'месяц', 'Пусто',
       'Чистая','КЦ', 'ОП', 'ОП_1']]

tdata = tdata.drop_duplicates()
tdata['ОП'] = tdata['ОП'].replace('ОП 3.1', 'ОП 3')

tdata[['Сумма', 'Семестры', '1 сем', 'Чистая выручка', 'Чистая']] = tdata[['Сумма', 
                                                                           'Семестры', 
                                                                           '1 сем', 
                                                                           'Чистая выручка', 
                                                                           'Чистая']].fillna(0).astype('int')

tdata = tdata.query('ОП != "ОП 14"')

g_1 = ['ОП 5', 'ОП 5.1', 'ОП 7', 'ОП 8', 'ОП 9', 'ОП 10', 'ОП 16']
g_2 = ['ОП 6', 'ОП 1', 'ОП 10.2', 'ОП 12', 'ОП 13', 'ОП 15']

tdata.loc[tdata['ОП'].isin(g_1), ['Группа']]='Пельванов'
tdata.loc[tdata['ОП'].isin(g_2), ['Группа']]='Группа №2'
tdata.loc[tdata['ОП'] == 'ОП 3', ['Группа']]='ОП 3'
tdata.loc[tdata['ОП'] == 'ОП 14.2', ['Группа']]= 'ОП 14'
tdata.loc[tdata['КЦ'] == 'КЦ 3', ['Группа']]= 'КЦ 3'
tdata.loc[tdata['КЦ'] == 'КЦ 4', ['Группа']]= 'Трофимов'

#------------------------------------------------------------

next = next.rename(columns={'Статус': 'Статус сделки'})
next = next.merge(status_lead, how='left', on='Статус сделки')
next = next.drop(['пять', 'семь', 'восемь'], axis=1)
next = next.rename(columns={'два': 'status', 'Статус сделки': 'Статус_сделки'})

next = next.query('status == "В работе" | status == "Договор" | status == "Назначена ЛВ" | status == "Проведена ЛВ"')
next = next.query('Статус_сделки != "Отправлено в АКАДА"')

next['Дата завершения'] = pd.to_datetime(next['Дата завершения'])
next['Год'] = next['Дата завершения'].dt.year
next['Месяц'] = next['Дата завершения'].dt.month

next = next.query('Статус_сделки != "Договор" & Статус_сделки != "Отправлены документы на ДО" & Статус_сделки != "Назначена ЛВ" & Контакт != ""')

next = next.query('ОП != "ОП 14"')

next.loc[next['ОП'].isin(g_1), ['Группа']]='Пельванов'
next.loc[next['ОП'].isin(g_2), ['Группа']]='Группа №2'
next.loc[next['ОП'] == 'ОП 3', ['Группа']]='ОП 3'
next.loc[next['ОП'] == 'ОП 14.2', ['Группа']]= 'ОП 14'
next.loc[next['КЦ'] == 'КЦ 3', ['Группа']]= 'КЦ 3'
next.loc[next['КЦ'] == 'КЦ 4 Воронеж', ['Группа']]= 'Трофимов'

#------------------------------------------------------------

wb = load_workbook(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\Отчёт по запланированным оплатам.xlsx')
#---------------------------
sheet = wb.get_sheet_by_name('Сводник_3')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(save, index=False, header=True):
    sheet.append(r)
#---------------------------
sheet = wb.get_sheet_by_name('Сводник_1')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(next, index=True, header=True):
    sheet.append(r)    
#---------------------------
sheet = wb.get_sheet_by_name('Сводник_2')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(tdata, index=True, header=True):
    sheet.append(r)

x_value = datetime.today().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
y_value = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

##---------------------------    
#sheet = wb.get_sheet_by_name('Отчёт')
#
#my_array = sheet['A']
#
#result_list = create_list(my_array, x_value, y_value)
#
#for index in result_list[:-1]:
#    start_row = re.findall(r'\d+', str(index))
#    start_row = int(start_row[0])
#    sheet.row_dimensions[start_row].hidden= True
##---------------------------

sheet = wb.get_sheet_by_name('Прослушка')

my_array = sheet['B']

result_list = create_list(my_array, x_value, y_value)

for index in result_list[:-1]:
    start_row = re.findall(r'\d+', str(index))
    start_row = int(start_row[0])
    sheet.row_dimensions[start_row].hidden= True
#---------------------------
wb.save(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\Отчёт по запланированным оплатам.xlsx')

#------------------------------------------------------------

print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Готово!")
print("")