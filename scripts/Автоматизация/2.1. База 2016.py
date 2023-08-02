import os

import pandas as pd
import numpy as np
import openpyxl as ox
import re
import shutil

from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

from datetime import datetime, timedelta

import warnings
warnings.filterwarnings("ignore")

import string

import pyodbc


print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Обновление БАЗЫ 2016")

БД = pd.read_excel(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\БАЗА КД ОБЩАЯ 2018.xlsx', sheet_name='2018')


БД.columns = ['ID сделки', 'ID встречи', 'месяц оплаты', 'ФИО АБИТУРИЕНТА', 'ТЕЛЕФОН', 'ОПЕРАТОР', 
              'ОТДЕЛ', 'ленд', 'источник', 'ID Лида', 'дата создания лида','первое обращение', 
              'ленд первого обращения', 'ХТС источник', 'ДАТА ПОДАЧИ', 'ОПЛАТА ПО ФАКТУ', 
              'Дата выписки', 'Дата счета', 'СТАВКА 1С', 'Сумма оплаты', 'Выручка РЕАЛ', 'Тип скидки', 
              'Размер скидки', 'Кол. Семестров', 'Номер Договора', 'Скидка/Полный', 'ВХОД', 
              'ВЫХОД', 'Коэфф', 'Примечание', 'Услуга', 'Услуга 2', 'форма', 'Специальность', 
              'Программа', 'ДАТА ЛВ', 'Принято', 'Визуал/ Офор.', 'оформление', 'СТАТУС ПЭО', 
              'Уровень 2', 'Уровень', 'Продукт ПМ', 'КЦ', 'Задача'
             ]


wb = load_workbook(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\БАЗА КД ОБЩАЯ 2016.xlsx')

sheet = wb.get_sheet_by_name('2018')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(БД, index=False, header=True):
    sheet.append(r)

wb.save(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\БАЗА КД ОБЩАЯ 2016.xlsx')

src_1 = r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\БАЗА КД ОБЩАЯ 2016.xlsx'
dest_1 = r'\\synergy.local\Documents\11.Коммерческий департамент\13.Отдел прямых продаж!\Зачисление\БАЗА КД ОБЩАЯ 2016.xlsx'

src_2 = r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\БАЗА КД ОБЩАЯ 2018.xlsx'
dest_2 = r'\\synergy.local\Documents\11.Коммерческий департамент\13.Отдел прямых продаж!\Зачисление\БАЗА КД ОБЩАЯ 2018.xlsx'

shutil.copyfile(src_1, dest_1)
shutil.copyfile(src_2, dest_2)

print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Готово!")
print("")