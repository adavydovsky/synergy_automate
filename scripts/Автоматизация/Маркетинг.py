import os

import pandas as pd
import numpy as np
import openpyxl as ox

from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

from datetime import datetime, timedelta

import warnings
warnings.filterwarnings("ignore")

import string

import pyodbc

print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Обновление маркетинга, пожалуйста, подождите...")

#day_m = datetime(2022, 12, 30)
day_m = datetime.today()
m = day_m.month
m_1 = m-1
y = day_m.year
y_1 = y-1

MonthDict={ 1 : "Январь",
       2 : "Февраль",
       3 : "Март",
       4 : "Апрель",
       5 : "Май",
       6 : "Июнь",
       7 : "Июль",
       8 : "Август",
       9 : "Сентябрь",
       10 : "Октябрь",
       11 : "Ноябрь",
       12 : "Декабрь"
          }

m = MonthDict[m]
m_1 = MonthDict[m_1]

monday = (day_m - timedelta(days=7)).strftime('%Y-%m-%d')
tuesday = (day_m - timedelta(days=6)).strftime('%Y-%m-%d')
wednesday = (day_m - timedelta(days=5)).strftime('%Y-%m-%d')
thursday = (day_m - timedelta(days=4)).strftime('%Y-%m-%d')
friday = (day_m - timedelta(days=3)).strftime('%Y-%m-%d')
saturday = (day_m - timedelta(days=2)).strftime('%Y-%m-%d')
sunday = (day_m - timedelta(days=1)).strftime('%Y-%m-%d')

month = day_m.strftime('%B')
#month = datetime(2022, 1, 30).strftime('%B')
                       
monday_22 = pd.to_datetime(monday) - timedelta(days=365)
tuesday_22 = pd.to_datetime(tuesday) - timedelta(days=365)
wednesday_22 = pd.to_datetime(wednesday) - timedelta(days=365)
thursday_22 = pd.to_datetime(thursday) - timedelta(days=365)
friday_22 = pd.to_datetime(friday) - timedelta(days=365)
saturday_22 = pd.to_datetime(saturday) - timedelta(days=365)
sunday_22 = pd.to_datetime(sunday) - timedelta(days=365)

monday_22 = monday_22.strftime('%Y-%m-%d')
tuesday_22 = tuesday_22.strftime('%Y-%m-%d')
wednesday_22 = wednesday_22.strftime('%Y-%m-%d')
thursday_22 = thursday_22.strftime('%Y-%m-%d')
friday_22 = friday_22.strftime('%Y-%m-%d')
saturday_22 = saturday_22.strftime('%Y-%m-%d')
sunday_22 = sunday_22.strftime('%Y-%m-%d')



conn = pyodbc.connect(r'Driver={SQL Server};Server=MSK1-BIDB01;Database=DWH;Trusted_Connection=yes;')
cursor = conn.cursor()



data = pd.read_excel(f'\\\\synergy.local\\Documents\\19.Группа мониторинга и сопровождения сделок\\01.Отчеты\Харитуня\\Шаблоны\\STAGE\\Расходы.xlsx', 
                     sheet_name='Лист3')
data_m = pd.read_excel(f'\\\\synergy.local\\Documents\\19.Группа мониторинга и сопровождения сделок\\01.Отчеты\\Харитуня\Шаблоны\\STAGE\\Расходы\\{y}\\{m_1}.xlsx', 
                     sheet_name='Лист3')
data_22 = pd.read_excel(f'\\\\synergy.local\\Documents\\19.Группа мониторинга и сопровождения сделок\\01.Отчеты\\Харитуня\\Шаблоны\\STAGE\\Расходы\\{y_1}\\{m}.xlsx', 
                        sheet_name='Лист3')
data_m_22 = pd.read_excel(f'\\\\synergy.local\\Documents\\19.Группа мониторинга и сопровождения сделок\\01.Отчеты\\Харитуня\\Шаблоны\\STAGE\\Расходы\\{y_1}\\{m_1}.xlsx', 
                          sheet_name='Лист3')

data = pd.concat([data, data_m])
data_22 = pd.concat([data_22, data_m_22])

status_lead = pd.read_excel(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Харитуня\Шаблоны\Сотрудники.xlsx', sheet_name='статусы')
sources = pd.read_excel(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Харитуня\Шаблоны\Сотрудники.xlsx', sheet_name='ленды')
ЧВ = pd.read_excel(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\!Расчет чистой выручки 0,25.xlsx', sheet_name='ЧВ')
dt = pd.read_excel(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Харитуня\КМВ\Реклама\Расход контекст.xlsx', sheet_name='КД 2021')



request_string_leads_2023 = '''

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

declare @enddate date = getdate()
declare @month int = month(getdate())
declare @startdate date = DATEFROMPARTS(2023, @month, 01); 

SELECT DISTINCT

       dr.[DATE_CREATE]    as "DATE"
      ,dr.[DATE_OP] as "Дата соединения на ОП"
      ,[FEATURES_6]  as "Старый/новый"       
      ,dr.[FEATURES_5] as "Ленд"
      ,s.tag1 as tag
      ,dr.[FEATURES_1] as "source"
      ,fc.[SPENT_BY_REQ] as "Расход"
      ,iif(dsr.[NAME] not like '%/%',dsr.[NAME], SUBSTRING(dsr.[NAME],0,PATINDEX('% / %',dsr.[NAME])))   as "Статус"              
      ,concat(e.[LAST_NAME]+' ',e.[NAME]+' ',e.[SECOND_NAME]) as "Ответственный"
      ,RU.[Marketer] as Marketer
      ,ul.[UF_LIDFORMA]
      ,PB.[name] as PB
      ,dr.[CODE] as id  
             

FROM [DWH].[dbo].[DIC_REQUEST] dr 

            LEFT JOIN [DWH].[dbo].[DIC_REQUEST_UTM] RU ON RU.[ID_REQUEST] = DR.[ID_REQUEST]

-- статусы

            LEFT JOIN  [DWH].[dbo].[DIC_STATUS_REQUEST] dsr ON dsr.ID_STATUS_REQUEST = dr.ID_STATUS_REQUEST

-- источник

            LEFT JOIN [DWH].[stage].[CRM_b_uts_crm_lead] ul ON ul.value_id = dr.CODE
            LEFT JOIN [DWH].[stage].[CRM_b_crm_lead] l ON l.ID = dr.code
            LEFT JOIN (SELECT [NAME] AS name_source, [STATUS_ID]
            FROM [DWH].[stage].[CRM_b_crm_status]
            WHERE [ENTITY_ID] = 'SOURCE') src ON src.[STATUS_ID] = l.[SOURCE_ID]

            LEFT JOIN ASS_REQUEST_SOURCE ARS WITH(NOLOCK) ON dR.ID_REQUEST = ARS.ID_REQUEST
            LEFT JOIN DIC_SOURCE S WITH(NOLOCK) ON ARS.ID_SOURCE = S.ID_SOURCE

-- исполнители

            LEFT JOIN [DWH].[dbo].DIC_EMPLOYEES E ON dR.Id_EMPLOYEES = E.ID_EMPLOYEES
            LEFT JOIN [DWH].[dbo].[ASS_EMPLOYEE_AND_ORGSTRUCTURE] ASS_OS ON E.ID_EMPLOYEES = ASS_OS.ID_EMPLOYEES

-- оргструктура

            LEFT JOIN [DWH].[dbo].[v_DIC_ORGSTRUCTURE] org on org.ID_ORGSTRUCTURE = ISNULL(ASS_OS.ID_ORGSTRUCTURE,-1)

-- продукт бюджета

            LEFT JOIN [DWH].[dbo].ASS_REQUEST_PRODUCT_BUDGET RPB With(nolock) ON dR.ID_REQUEST=RPB.ID_REQUEST and dR.DATE_CREATE= RPB.R_DATE
            LEFT JOIN [DWH].[dbo].DIC_PRODUCT_BUDGET PB With(nolock) ON RPB.ID_PRODUCT_BUDGET=PB.ID_PRODUCT_BUDGET

-- инвойсы

            LEFT JOIN [DWH].[dbo].ASS_REQUEST_DEAL ASS_RD ON ASS_RD.ID_REQUEST = dR.ID_REQUEST
            LEFT JOIN [DWH].[dbo].ASS_DEAL_INVOICE ASS_DI ON ASS_DI.ID_DEAL = ASS_RD.ID_DEAL
            LEFT JOIN [DWH].[dbo].DIC_INVOICE D_I ON D_I.ID_INVOICE = ASS_DI.ID_INVOICE
            LEFT JOIN [DWH].[dbo].DIC_STATUS_INVOICE S_I ON S_I.ID_STATUS_INVOICE = D_I.ID_STATUS_INVOICE

--расход

            LEFT JOIN [DWH].[dbo].[FCT_REQUEST] fc With(nolock) on fc.[ID_REQUEST] = dr.ID_REQUEST

WHERE 1=1 

            AND isnull(dr.SIGN_DELETED,0) =0
            AND dr.[DATE_CREATE] BETWEEN @startdate AND @enddate
            AND org.full_NAME LIKE '%Коммерческий департамент (КМВ)\%'
            AND org.name_2 NOT LIKE '%МАП%'
            AND src.name_source LIKE '%Веб%'
            AND dr.[FEATURES_5] NOT LIKE '%itshool2022%'
            AND dr.[FEATURES_5] NOT LIKE '%startprof%'
            AND dr.[FEATURES_5] NOT LIKE '%itshool2022_kd%'
            AND dr.[FEATURES_5] NOT LIKE '%mc-psychology%'
            
'''



request_string_vhod_2023 = '''

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

declare @enddate date = getdate()
declare @month int = month(getdate())
declare @startdate date = DATEFROMPARTS(2023, @month, 01); 

SELECT DISTINCT

     dr.[DATE_CREATE]    as "DATE"
    ,dr.[DATE_OP] as "Дата соединения на ОП"
    ,[FEATURES_6]  as "Старый/новый"       
    ,dr.[FEATURES_5] as "Ленд"
    ,s.tag1 as tag
    ,dr.[FEATURES_1] as "source"
    ,fc.[SPENT_BY_REQ] as "Расход"
    ,iif(dsr.[NAME] not like '%/%',dsr.[NAME], SUBSTRING(dsr.[NAME],0,PATINDEX('% / %',dsr.[NAME])))   as "Статус"              
    ,concat(e.[LAST_NAME]+' ',e.[NAME]+' ',e.[SECOND_NAME]) as "Ответственный"
    ,RU.[Marketer] as Marketer
    ,ul.[UF_LIDFORMA]
    ,PB.[name] as PB
    ,dr.[CODE] as id  
             
FROM [DWH].[dbo].[DIC_REQUEST] dr 

            LEFT JOIN [DWH].[dbo].[DIC_REQUEST_UTM] RU on RU.[ID_REQUEST] = DR.[ID_REQUEST]
            
 -- статусы
 
            LEFT JOIN  [DWH].[dbo].[DIC_STATUS_REQUEST] dsr on dsr.ID_STATUS_REQUEST = dr.ID_STATUS_REQUEST

-- источник

            LEFT JOIN [DWH].[stage].[CRM_b_uts_crm_lead] ul on ul.value_id = dr.CODE
            LEFT JOIN [DWH].[stage].[CRM_b_crm_lead] l on l.ID = dr.code
            LEFT JOIN (SELECT [NAME] as name_source, [STATUS_ID]
            FROM [DWH].[stage].[CRM_b_crm_status]
            WHERE [ENTITY_ID] = 'SOURCE') src on src.[STATUS_ID] = l.[SOURCE_ID]

            LEFT JOIN ASS_REQUEST_SOURCE ARS WITH(NOLOCK) ON dR.ID_REQUEST = ARS.ID_REQUEST
            LEFT JOIN DIC_SOURCE S WITH(NOLOCK) ON ARS.ID_SOURCE = S.ID_SOURCE
            
-- исполнители

            LEFT JOIN [DWH].[dbo].DIC_EMPLOYEES E ON dR.Id_EMPLOYEES = E.ID_EMPLOYEES            
            LEFT JOIN [DWH].[dbo].[ASS_EMPLOYEE_AND_ORGSTRUCTURE] ASS_OS ON E.ID_EMPLOYEES = ASS_OS.ID_EMPLOYEES

-- оргструктура

            LEFT JOIN [DWH].[dbo].[v_DIC_ORGSTRUCTURE] org on org.ID_ORGSTRUCTURE = ISNULL(ASS_OS.ID_ORGSTRUCTURE,-1)

-- продукт бюджета

            LEFT JOIN [DWH].[dbo].ASS_REQUEST_PRODUCT_BUDGET RPB With(nolock) ON dR.ID_REQUEST=RPB.ID_REQUEST and dR.DATE_CREATE= RPB.R_DATE
            LEFT JOIN [DWH].[dbo].DIC_PRODUCT_BUDGET PB With(nolock) ON RPB.ID_PRODUCT_BUDGET=PB.ID_PRODUCT_BUDGET

-- инвойсы

            LEFT JOIN [DWH].[dbo].ASS_REQUEST_DEAL ASS_RD ON ASS_RD.ID_REQUEST = dR.ID_REQUEST
            LEFT JOIN [DWH].[dbo].ASS_DEAL_INVOICE ASS_DI ON ASS_DI.ID_DEAL = ASS_RD.ID_DEAL
            LEFT JOIN [DWH].[dbo].DIC_INVOICE D_I ON D_I.ID_INVOICE = ASS_DI.ID_INVOICE
            LEFT JOIN [DWH].[dbo].DIC_STATUS_INVOICE S_I ON S_I.ID_STATUS_INVOICE = D_I.ID_STATUS_INVOICE

--расход

            LEFT JOIN [DWH].[dbo].[FCT_REQUEST] fc With(nolock) on fc.[ID_REQUEST] = dr.ID_REQUEST

WHERE 1=1 

            AND isnull(dr.SIGN_DELETED,0) =0
            AND dr.[DATE_CREATE] between @startdate and @enddate
            AND org.full_NAME like '%Коммерческий департамент (КМВ)\%'
            AND org.name_2 not like '%МАП%'
            AND src.name_source like '%Вход%'


'''


request_string_leads_2022 = '''

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

declare @month int = month(getdate())
declare @day int = day(getdate())
declare @startdate date = DATEFROMPARTS(2022, @month, 01)
declare @enddate date = DATEFROMPARTS(2022, @month, @day)

SELECT DISTINCT

              dr.[DATE_CREATE]    as "DATE"
            ,dr.[DATE_OP] as "Дата соединения на ОП"
            ,[FEATURES_6]  as "Старый/новый"       
            ,dr.[FEATURES_5] as "Ленд"
            ,s.tag1 as tag
            ,dr.[FEATURES_1] as "source"
            ,fc.[SPENT_BY_REQ] as "Расход"
            ,iif(dsr.[NAME] not like '%/%',dsr.[NAME], SUBSTRING(dsr.[NAME],0,PATINDEX('% / %',dsr.[NAME])))   as "Статус"              
            ,concat(e.[LAST_NAME]+' ',e.[NAME]+' ',e.[SECOND_NAME]) as "Ответственный"
            ,RU.[Marketer] as Marketer
            ,ul.[UF_LIDFORMA]
            ,PB.[name] as PB
            ,dr.[CODE] as id  
             

FROM [DWH].[dbo].[DIC_REQUEST] dr 

            LEFT JOIN [DWH].[dbo].[DIC_REQUEST_UTM] RU on RU.[ID_REQUEST] = DR.[ID_REQUEST]
            
 -- статусы
 
            LEFT JOIN  [DWH].[dbo].[DIC_STATUS_REQUEST] dsr on dsr.ID_STATUS_REQUEST = dr.ID_STATUS_REQUEST

-- источник

            LEFT JOIN [DWH].[stage].[CRM_b_uts_crm_lead] ul on ul.value_id = dr.CODE
            LEFT JOIN [DWH].[stage].[CRM_b_crm_lead] l on l.ID = dr.code
            LEFT JOIN (SELECT [NAME] as name_source, [STATUS_ID]
            FROM [DWH].[stage].[CRM_b_crm_status]
            WHERE [ENTITY_ID] = 'SOURCE') src on src.[STATUS_ID] = l.[SOURCE_ID]

            LEFT JOIN ASS_REQUEST_SOURCE ARS WITH(NOLOCK) ON dR.ID_REQUEST = ARS.ID_REQUEST
            LEFT JOIN DIC_SOURCE S WITH(NOLOCK) ON ARS.ID_SOURCE = S.ID_SOURCE
            
-- исполнители

            LEFT JOIN [DWH].[dbo].DIC_EMPLOYEES E ON dR.Id_EMPLOYEES = E.ID_EMPLOYEES
            LEFT JOIN [DWH].[dbo].[ASS_EMPLOYEE_AND_ORGSTRUCTURE] ASS_OS ON E.ID_EMPLOYEES = ASS_OS.ID_EMPLOYEES

-- оргструктура

            LEFT JOIN [DWH].[dbo].[v_DIC_ORGSTRUCTURE] org on org.ID_ORGSTRUCTURE = ISNULL(ASS_OS.ID_ORGSTRUCTURE,-1)

-- продукт бюджета

            LEFT JOIN [DWH].[dbo].ASS_REQUEST_PRODUCT_BUDGET RPB With(nolock) ON dR.ID_REQUEST=RPB.ID_REQUEST and dR.DATE_CREATE= RPB.R_DATE
            LEFT JOIN [DWH].[dbo].DIC_PRODUCT_BUDGET PB With(nolock) ON RPB.ID_PRODUCT_BUDGET=PB.ID_PRODUCT_BUDGET

-- инвойсы

            LEFT JOIN [DWH].[dbo].ASS_REQUEST_DEAL ASS_RD ON ASS_RD.ID_REQUEST = dR.ID_REQUEST
            LEFT JOIN [DWH].[dbo].ASS_DEAL_INVOICE ASS_DI ON ASS_DI.ID_DEAL = ASS_RD.ID_DEAL
            LEFT JOIN [DWH].[dbo].DIC_INVOICE D_I ON D_I.ID_INVOICE = ASS_DI.ID_INVOICE
            LEFT JOIN [DWH].[dbo].DIC_STATUS_INVOICE S_I ON S_I.ID_STATUS_INVOICE = D_I.ID_STATUS_INVOICE

--расход

            LEFT JOIN [DWH].[dbo].[FCT_REQUEST] fc With(nolock) on fc.[ID_REQUEST] = dr.ID_REQUEST

WHERE 1=1 

            AND isnull(dr.SIGN_DELETED,0) =0
            AND dr.[DATE_CREATE] between @startdate and @enddate
            AND org.full_NAME like '%Коммерческий департамент (КМВ)\%'
            AND org.name_2 not like '%МАП%'
            AND src.name_source like '%Веб%'
            AND dr.[FEATURES_5] not like '%itshool2022%'
            AND dr.[FEATURES_5] not like '%startprof%'
            AND dr.[FEATURES_5] not like '%itshool2022_kd%'
            AND dr.[FEATURES_5] not like '%mc-psychology%'

 

'''



request_string_vhod_2022 = '''

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

declare @month int = month(getdate())
declare @day int = day(getdate())
declare @startdate date = DATEFROMPARTS(2022, @month, 01)
declare @enddate date = DATEFROMPARTS(2022, @month, @day)

SELECT DISTINCT

     dr.[DATE_CREATE]    as "DATE"
    ,dr.[DATE_OP] as "Дата соединения на ОП"
    ,[FEATURES_6]  as "Старый/новый"       
    ,dr.[FEATURES_5] as "Ленд"
    ,s.tag1 as tag
    ,dr.[FEATURES_1] as "source"
    ,fc.[SPENT_BY_REQ] as "Расход"
    ,iif(dsr.[NAME] not like '%/%',dsr.[NAME], SUBSTRING(dsr.[NAME],0,PATINDEX('% / %',dsr.[NAME])))   as "Статус"              
    ,concat(e.[LAST_NAME]+' ',e.[NAME]+' ',e.[SECOND_NAME]) as "Ответственный"
    ,RU.[Marketer] as Marketer
    ,ul.[UF_LIDFORMA]
    ,PB.[name] as PB
    ,dr.[CODE] as id  
             
FROM [DWH].[dbo].[DIC_REQUEST] dr 

            LEFT JOIN [DWH].[dbo].[DIC_REQUEST_UTM] RU on RU.[ID_REQUEST] = DR.[ID_REQUEST]
            
 -- статусы
 
            LEFT JOIN  [DWH].[dbo].[DIC_STATUS_REQUEST] dsr on dsr.ID_STATUS_REQUEST = dr.ID_STATUS_REQUEST

-- источник

            LEFT JOIN [DWH].[stage].[CRM_b_uts_crm_lead] ul on ul.value_id = dr.CODE
            LEFT JOIN [DWH].[stage].[CRM_b_crm_lead] l on l.ID = dr.code
            LEFT JOIN (SELECT [NAME] as name_source, [STATUS_ID]
            FROM [DWH].[stage].[CRM_b_crm_status]
            WHERE [ENTITY_ID] = 'SOURCE') src on src.[STATUS_ID] = l.[SOURCE_ID]

            LEFT JOIN ASS_REQUEST_SOURCE ARS WITH(NOLOCK) ON dR.ID_REQUEST = ARS.ID_REQUEST
            LEFT JOIN DIC_SOURCE S WITH(NOLOCK) ON ARS.ID_SOURCE = S.ID_SOURCE
            
-- исполнители

            LEFT JOIN [DWH].[dbo].DIC_EMPLOYEES E ON dR.Id_EMPLOYEES = E.ID_EMPLOYEES            
            LEFT JOIN [DWH].[dbo].[ASS_EMPLOYEE_AND_ORGSTRUCTURE] ASS_OS ON E.ID_EMPLOYEES = ASS_OS.ID_EMPLOYEES

-- оргструктура

            LEFT JOIN [DWH].[dbo].[v_DIC_ORGSTRUCTURE] org on org.ID_ORGSTRUCTURE = ISNULL(ASS_OS.ID_ORGSTRUCTURE,-1)

-- продукт бюджета

            LEFT JOIN [DWH].[dbo].ASS_REQUEST_PRODUCT_BUDGET RPB With(nolock) ON dR.ID_REQUEST=RPB.ID_REQUEST and dR.DATE_CREATE= RPB.R_DATE
            LEFT JOIN [DWH].[dbo].DIC_PRODUCT_BUDGET PB With(nolock) ON RPB.ID_PRODUCT_BUDGET=PB.ID_PRODUCT_BUDGET

-- инвойсы

            LEFT JOIN [DWH].[dbo].ASS_REQUEST_DEAL ASS_RD ON ASS_RD.ID_REQUEST = dR.ID_REQUEST
            LEFT JOIN [DWH].[dbo].ASS_DEAL_INVOICE ASS_DI ON ASS_DI.ID_DEAL = ASS_RD.ID_DEAL
            LEFT JOIN [DWH].[dbo].DIC_INVOICE D_I ON D_I.ID_INVOICE = ASS_DI.ID_INVOICE
            LEFT JOIN [DWH].[dbo].DIC_STATUS_INVOICE S_I ON S_I.ID_STATUS_INVOICE = D_I.ID_STATUS_INVOICE

--расход

            LEFT JOIN [DWH].[dbo].[FCT_REQUEST] fc With(nolock) on fc.[ID_REQUEST] = dr.ID_REQUEST

WHERE 1=1 

            AND isnull(dr.SIGN_DELETED,0) =0
            AND dr.[DATE_CREATE] between @startdate and @enddate
            AND org.full_NAME like '%Коммерческий департамент (КМВ)\%'
            AND org.name_2 not like '%МАП%'
            AND src.name_source like '%Вход%'

'''


#cursor.execute(requestString)
leads = pd.read_sql_query(request_string_leads_2023, conn)
vhod = pd.read_sql_query(request_string_vhod_2023, conn)
leads_2022 = pd.read_sql_query(request_string_leads_2022, conn)
vhod_2022 = pd.read_sql_query(request_string_vhod_2022, conn)


sources = sources.drop(["Ленд", "Распределение", "Стоимость", "РК", "ВС/ВХ", "Принадлежность РК ДП", "Продукт", "Unnamed: 7", "Unnamed: 8"], axis=1)
sources = sources.fillna('Пусто')
sources= sources.query('Источник != "Пусто"')
sources['ИСТ']= sources['ИСТ'].replace("vk","В контакте")


sources = sources.rename(columns={'Источник': 'source'})


ЧВ = ЧВ.iloc[2:]
ЧВ.columns = ['ID сделки', 'КЦ', 'ОТДЕЛ', 'ОПЕРАТОР', 
              'ОПЛАТА_ПО_ФАКТУ', 'ДАТА ПОДАЧИ', 'Выручка РЕАЛ', 
              'Кол сем', '1 сем', 'Чистая', 'Продукт', 'ID лида', 
              'дата_создания', 'Ленд', 'source', 'Источник ХТС', 'Срок', 
              'Продукт_2', 'Договор', 'Встреча', 'Клиент', 'Услуга_2', 
              'null', 'null', 'null', 'null']


ЧВ['ОПЛАТА_ПО_ФАКТУ'] = pd.to_datetime(ЧВ.ОПЛАТА_ПО_ФАКТУ)
ЧВ['дата_создания'] = ЧВ['дата_создания'].astype(str)


ЧВ = ЧВ.query('дата_создания != "00:00:00"')
ЧВ['дата_создания'] = pd.to_datetime(ЧВ.дата_создания)


ЧВ = ЧВ.query('ОПЛАТА_ПО_ФАКТУ >= "01.01.2020"')
ЧВ.reset_index(drop=True, inplace=True)

ЧВ = ЧВ.drop(["Ленд", "Источник ХТС", "Срок", 
              "Продукт_2", "Договор", "Встреча", 
              "Продукт", "null", "null", "null", "null"], axis=1)


ЧВ = ЧВ.merge(sources, how='left', on='source')
ЧВ = ЧВ.drop(['source', 'Метка'], axis=1)
ЧВ = ЧВ.rename(columns={'ИСТ': 'source'})


ЧВ = ЧВ.fillna('Пусто')
ЧВ= ЧВ.query('source != ["Пусто", "органика", "прочие", "рассылка"]')


status_lead.columns = ['status', 'один', 'два', 'три', 'четыре', 'пять', 'шесть', 'семь', 'восемь']
status_lead= status_lead.drop(['два', 'три', 'четыре', 'пять'], axis=1)

status_lead = status_lead.query('один != "deal"')


data= data.fillna(0)
data_22= data_22.fillna(0)

data.columns = ['unnamed_0', 'id_lead', 'id_employee','dt_create', 'expenditure', 
                'id_product', 'product_name', 'source', 'org', 'status', 
                'source_group', 'url_convert', 'lend', 'account', 'campaign', 'medium']

data_22.columns = ['unnamed_0', 'id_lead', 'id_employee','dt_create', 'expenditure', 
                'id_product', 'product_name', 'source', 'org', 'status', 
                'source_group', 'url_convert', 'lend', 'account', 'campaign', 'medium']

data['id_product'] = data['id_product'].astype(int)
data_22['id_product'] = data_22['id_product'].astype(int)

data[['status', 'del']]= data['status'].str.split(' / ', expand=True)
data_22[['status', 'del']]= data_22['status'].str.split(' / ', expand=True)


data = data.drop(['unnamed_0', 'id_employee', 'id_product', 'org', 
                  'del', 'source_group', 'url_convert', 'lend', 'account', 
                  'campaign', 'medium'], axis=1)

data_22 = data_22.drop(['unnamed_0', 'id_employee', 'id_product', 'org', 
                  'del', 'source_group', 'url_convert', 'lend', 'account', 
                  'campaign', 'medium'], axis=1)


data = data.merge(status_lead, how='left', on='status')
data = data.drop(['status', 'шесть', 'семь', 'восемь'], axis=1)
data = data.rename(columns={'один': 'status'})

data_22 = data_22.merge(status_lead, how='left', on='status')
data_22 = data_22.drop(['status', 'шесть', 'семь', 'восемь'], axis=1)
data_22 = data_22.rename(columns={'один': 'status'})

data = data.merge(sources, how='left', on='source')
data = data.drop(['source', 'Метка'], axis=1)
data = data.rename(columns={'ИСТ': 'source'})

data_22 = data_22.merge(sources, how='left', on='source')
data_22 = data_22.drop(['source', 'Метка'], axis=1)
data_22 = data_22.rename(columns={'ИСТ': 'source'})


pivot_table_1 = data.pivot_table(index='source', values='id_lead', aggfunc='count')
pivot_table_2 = data.pivot_table(index='source', values='expenditure', aggfunc='sum')

pivot_table_lead= pivot_table_1.merge(pivot_table_2, how='left', on='source')
pivot_table_lead['expenditure'] = pivot_table_lead['expenditure'].astype(int)

data_eff = data.query('status == "эфф"')

data_pivot = data.query('dt_create >= @monday & dt_create <= @sunday')
data_pivot = data_pivot.pivot_table(index='source', values='expenditure', aggfunc='sum')

data_pivot_2 = data_22.query('dt_create >= @monday_22 & dt_create <= @sunday_22')
data_pivot_2 = data_pivot_2.pivot_table(index='source', values='expenditure', aggfunc='sum')

data_pivot_all = data_pivot.merge(data_pivot_2, how='outer', on='source').reset_index()
data_pivot_all.columns = ['source', '2023', '2022']

pivot_table_3 = data_eff.pivot_table(index='source', values='id_lead', aggfunc='count')
pivot_table_4 = data_eff.pivot_table(index='source', values='expenditure', aggfunc='sum')

pivot_table_eff_lead= pivot_table_3.merge(pivot_table_4, how='left', on='source')
pivot_table_eff_lead['expenditure'] = pivot_table_eff_lead['expenditure'].astype(int)

ЧВ_pivot= ЧВ.query('ОПЛАТА_ПО_ФАКТУ >= "01.01.2023"')
ЧВ_pivot= ЧВ.query('дата_создания >= "01.01.2023"')


pivot_table_5 = ЧВ_pivot.pivot_table(index='source', values='ID лида', aggfunc='count')
pivot_table_6 = ЧВ_pivot.pivot_table(index='source', values='Чистая', aggfunc='sum')

pivot_table_чв = pivot_table_5.merge(pivot_table_6, how='left', on='source')

l = ["ВПО Заодист", "ВПО/General", "Колледж", "Магистратура", "Аспирантура", "Бренд",
     "НТШ all", "НТШ дети", "НТШ Актерский МК", "Анимация"]

dt = dt[['Дата', 'Продукт', 'Источник', 'Триграмма', 'Расход', 'Лиды', 'Cpl']]
dt['Месяц'] = dt['Дата'].dt.strftime('%B')

#dt = dt.query('Продукт != "Synergy IT Academy" & Продукт != "Бизнес лагерь"') #del 'Месяц == @month'
dt = dt.query('Дата >= "2023-01-01"')


dt[['Клики', 'Лиды', 'Cpl']] = dt[['Клики', 'Лиды', 'Cpl']].fillna(0)
dt[['Клики', 'Лиды', 'Cpl']] = dt[['Клики', 'Лиды', 'Cpl']].astype('int')


pivot_1 = dt.pivot_table(index='Дата', columns='Источник', values='Расход', aggfunc='sum')
pivot_2 = dt.pivot_table(index='Дата', columns='Источник', values='Лиды', aggfunc='sum')
pivot = pivot_1.merge(pivot_2, on='Дата', how='outer')
pivot.columns = ['google_расход', 'yandex_расход', 'google_лиды', 'yandex_лиды']


pivot_table = data.pivot_table(index='dt_create', values='expenditure', aggfunc='sum')


filter_status = ['Дубль', 'Ошибка номера', 'Повторные заявки', 'Спам']


leads.columns = ['date', 'date_op', 'old_new', 'lend', 
                 'tag', 'source', 'expenses', 'status', 
                 'resp', 'marketer', 'uf_lidforma', 'pb', 'id']


leads[['tag', 'source']] = leads[['tag', 'source']].fillna('unknown')


leads['tag'] = np.where((leads.tag == 'unknown'), leads.source, leads.tag)


leads.loc[leads['tag'] == 'unknown','tag'] = np.nan
leads.loc[leads['source'] == 'unknown','source'] = np.nan


leads['tag'] = leads['tag'].replace('tik_tok', 'Tiktok')
leads['tag'] = leads['tag'].fillna('organic')


leads['date'] = pd.to_datetime(leads.date)
leads['date_op'] = leads['date_op'].dt.strftime('%Y-%m-%d')


vhod.columns = ['date', 'date_op', 'old_new', 'lend', 
                 'tag', 'source', 'expenses', 'status', 
                 'resp', 'marketer', 'uf_lidforma', 'pb', 'id']


vhod[['tag', 'source']] = vhod[['tag', 'source']].fillna('unknown')


vhod['tag'] = np.where((vhod.tag == 'unknown'), vhod.source, vhod.tag)


vhod.loc[vhod['tag'] == 'unknown','tag'] = np.nan
vhod.loc[vhod['source'] == 'unknown','source'] = np.nan


vhod['tag'] = vhod['tag'].replace('tik_tok', 'Tiktok')
vhod['tag'] = vhod['tag'].fillna('organic')


vhod['date'] = pd.to_datetime(leads.date)
vhod['date_op'] = vhod['date_op'].dt.strftime('%Y-%m-%d')


leads_2022.columns = ['date', 'date_op', 'old_new', 'lend', 
                 'tag', 'source', 'expenses', 'status', 
                 'resp', 'marketer', 'uf_lidforma', 'pb', 'id']


leads_2022[['tag', 'source']] = leads_2022[['tag', 'source']].fillna('unknown')


leads_2022['tag'] = np.where((leads_2022.tag == 'unknown'), leads_2022.source, leads_2022.tag)


leads_2022.loc[leads_2022['tag'] == 'unknown','tag'] = np.nan
leads_2022.loc[leads_2022['source'] == 'unknown','source'] = np.nan


leads_2022['tag'] = leads_2022['tag'].replace('tik_tok', 'Tiktok')
leads_2022['tag'] = leads_2022['tag'].fillna('organic')


leads_2022['date'] = pd.to_datetime(leads_2022.date)
leads_2022['date_op'] = leads_2022['date_op'].dt.strftime('%Y-%m-%d')


vhod_2022.columns = ['date', 'date_op', 'old_new', 'lend', 
                 'tag', 'source', 'expenses', 'status', 
                 'resp', 'marketer', 'uf_lidforma', 'pb', 'id']


vhod_2022[['tag', 'source']] = vhod_2022[['tag', 'source']].fillna('unknown')


vhod_2022['tag'] = np.where((vhod_2022.tag == 'unknown'), vhod_2022.source, vhod_2022.tag)


vhod_2022.loc[vhod_2022['tag'] == 'unknown','tag'] = np.nan
vhod_2022.loc[vhod_2022['source'] == 'unknown','source'] = np.nan


vhod_2022['tag'] = vhod_2022['tag'].replace('tik_tok', 'Tiktok')
vhod_2022['tag'] = vhod_2022['tag'].fillna('organic')


vhod_2022['date'] = pd.to_datetime(leads.date)
vhod_2022['date_op'] = vhod_2022['date_op'].dt.strftime('%Y-%m-%d')


date = pd.to_datetime('today')

date = date.strftime('%Y-%m-%d')


def day_report(x, y):
    
    gross = leads.query('date == @x')
    eff = leads.query('date == @x & status not in @filter_status')
    soed = leads.query('date == @x & date_op == @x & status not in @filter_status')
    sm = leads.query('date == @x & lend == "synergymobileapp" & status not in @filter_status')
    sm_soed = leads.query('date == @x & lend == "synergymobileapp" & date_op == @x & status not in @filter_status')
    
    gross_22 = leads_2022.query('date == @y')
    eff_22 = leads_2022.query('date == @y & status not in @filter_status')
    soed_22 = leads_2022.query('date == @y & date_op == @y & status not in @filter_status')
    sm_22 = leads_2022.query('date == @y & lend == "synergymobileapp" & status not in @filter_status')
    vh_22 = vhod_2022.query('date == @y & status not in @filter_status')
    
    vh = vhod.query('date == @x & status not in @filter_status')
    vh_soed = vhod.query('date == @x & date_op == @x & status not in @filter_status')
    
    pivot_table_1 = gross.pivot_table(index='tag', values='id', aggfunc='count')
    pivot_table_2 = eff.pivot_table(index='tag', values='id', aggfunc='count')
    pivot_table_3 = soed.pivot_table(index='tag', values='id', aggfunc='count')
    pivot_table_4 = sm.pivot_table(index='lend', values='id', aggfunc='count')
    pivot_table_5 = sm_soed.pivot_table(index='lend', values='id', aggfunc='count')
    pivot_table_6 = vh.pivot_table(index='tag', values='id', aggfunc='count')
    pivot_table_7 = vh_soed.pivot_table(index='tag', values='id', aggfunc='count')
    
    pivot_table_8 = gross_22.pivot_table(index='tag', values='id', aggfunc='count')
    pivot_table_9 = eff_22.pivot_table(index='tag', values='id', aggfunc='count')
    pivot_table_10 = sm_22.pivot_table(index='lend', values='id', aggfunc='count')
    pivot_table_11 = vh_22.pivot_table(index='tag', values='id', aggfunc='count')
    pivot_table_12 = soed_22.pivot_table(index='tag', values='id', aggfunc='count')
    
    #Основной сводник
    
    if len(pivot_table_1) > 0 and len(pivot_table_2) > 0:
        pivot_table_day = pivot_table_1.merge(pivot_table_2, on='tag', how='outer').reset_index()
        pivot_table_day.columns = ['tag', 'gross', 'eff']
        if len(pivot_table_3) > 0:
            pivot_table_day = pivot_table_day.merge(pivot_table_3, on='tag', how='outer')
            pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed']
            if len(pivot_table_6) > 0:
                pivot_table_day = pivot_table_day.merge(pivot_table_6, on='tag', how='outer')
                pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh']
                if len(pivot_table_7) > 0:
                    pivot_table_day = pivot_table_day.merge(pivot_table_7, on='tag', how='outer')
                    pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed']
                else:
                    pivot_table_day = pivot_table_day.merge(pivot_table_7, on='tag', how='outer')
                    pivot_table_day['vh_soed'] = np.nan
                    pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed']
            else:        
                pivot_table_day = pivot_table_day.merge(pivot_table_6, on='tag', how='outer')
                pivot_table_day['vh'] = np.nan
                pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh']
                pivot_table_day = pivot_table_day.merge(pivot_table_7, on='tag', how='outer')
                pivot_table_day['vh_soed'] = np.nan
                pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed']
        else:
            pivot_table_day = pivot_table_day.merge(pivot_table_3, on='tag', how='outer')
            pivot_table_day['soed'] = np.nan
            pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed']
            pivot_table_day = pivot_table_day.merge(pivot_table_6, on='tag', how='outer')
            pivot_table_day['vh'] = np.nan
            pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh']
            pivot_table_day = pivot_table_day.merge(pivot_table_7, on='tag', how='outer')
            pivot_table_day['vh_soed'] = np.nan
            pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed']
    else:
        pivot_table_day = pivot_table_1.merge(pivot_table_2, on='tag', how='outer').reset_index()
        pivot_table_day[['gross', 'eff']] = np.nan
        pivot_table_day.columns = ['tag', 'gross', 'eff']
        pivot_table_day = pivot_table_day.merge(pivot_table_3, on='tag', how='outer')
        pivot_table_day['soed'] = np.nan
        pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed']
        pivot_table_day = pivot_table_day.merge(pivot_table_6, on='tag', how='outer')
        pivot_table_day['vh'] = np.nan
        pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh']
        pivot_table_day = pivot_table_day.merge(pivot_table_7, on='tag', how='outer')
        pivot_table_day['vh_soed'] = np.nan
        pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed']
        
    #Подсоединение 2022 к основному своднику
        
    if len(pivot_table_8) > 0:
        pivot_table_day = pivot_table_day.merge(pivot_table_8, on='tag', how='outer')
        pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22']
        if len(pivot_table_9) > 0:
            pivot_table_day = pivot_table_day.merge(pivot_table_9, on='tag', how='outer')
            pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22']
            if len(pivot_table_11) > 0:
                pivot_table_day = pivot_table_day.merge(pivot_table_11, on='tag', how='outer')
                pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22', 'vh_22']
                if len(pivot_table_12) > 0:
                    pivot_table_day = pivot_table_day.merge(pivot_table_12, on='tag', how='outer')
                    pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22', 'vh_22', 'soed_22']
                else:
                    pivot_table_day = pivot_table_day.merge(pivot_table_12, on='tag', how='outer')
                    pivot_table_day['soed_22'] = np.nan
                    pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22', 'vh_22', 'soed_22']
            else:
                pivot_table_day = pivot_table_day.merge(pivot_table_11, on='tag', how='outer')
                pivot_table_day['vh_22'] = np.nan
                pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22', 'vh_22']
                pivot_table_day = pivot_table_day.merge(pivot_table_12, on='tag', how='outer')
                pivot_table_day['soed_22'] = np.nan
                pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22', 'vh_22', 'soed_22', '0']
        else:
            pivot_table_day = pivot_table_day.merge(pivot_table_9, on='tag', how='outer')
            pivot_table_day['eff_22'] = np.nan
            pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22']
            pivot_table_day = pivot_table_day.merge(pivot_table_11, on='tag', how='outer')
            pivot_table_day['vh_22'] = np.nan
            pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22', 'vh_22']
            pivot_table_day = pivot_table_day.merge(pivot_table_12, on='tag', how='outer')
            pivot_table_day['soed_22'] = np.nan
            pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22', 'vh_22', 'soed_22']
    else:
        pivot_table_day = pivot_table_day.merge(pivot_table_8, on='tag', how='outer')
        pivot_table_day['gross_22'] = np.nan
        pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22']
        pivot_table_day = pivot_table_day.merge(pivot_table_9, on='tag', how='outer')
        pivot_table_day['eff_22'] = np.nan
        pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22']  
        pivot_table_day = pivot_table_day.merge(pivot_table_11, on='tag', how='outer')
        pivot_table_day['vh_22'] = np.nan
        pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22', 'vh_22']    
        pivot_table_day = pivot_table_day.merge(pivot_table_12, on='tag', how='outer')
        pivot_table_day['soed_22'] = np.nan
        pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22', 'vh_22', 'soed_22']    
    #сводная таблица по СМ
        
    if len(pivot_table_4) > 0 and len(pivot_table_5) > 0:
        pivot_table_sm_day = pivot_table_4.merge(pivot_table_5, on='lend', how='outer').reset_index()
        pivot_table_sm_day.columns = ['lend', 'eff', 'soed']
    else:
        pivot_table_sm_day = pivot_table_4.merge(pivot_table_5, on='lend', how='outer').reset_index()
        pivot_table_sm_day['soed'] = np.nan
        #pivot_table_sm_day.columns = ['lend', 'eff', 'soed']
        
    #Подсоединение 2022 года к своднику по СМ
        
    if len(pivot_table_10) > 0:
        pivot_table_sm_day = pivot_table_sm_day.merge(pivot_table_10, on='lend', how='outer')
        #pivot_table_sm_day.columns = ['tag', 'eff', 'soed', 'eff_21']
    else:
        pivot_table_sm_day = pivot_table_sm_day.merge(pivot_table_10, on='lend', how='outer')
        pivot_table_sm_day['eff_22'] = np.nan
        #pivot_table_sm_day.columns = ['tag', 'eff', 'soed', 'eff_21']
        
    frames = [pivot_table_day, pivot_table_sm_day]
    result = pd.concat(frames).reset_index(drop=True)
    
    return result



def week_report(x_1, y_1, x_2, y_2):
    
    gross = leads[leads['date'].between(x_1, x_2, inclusive=True)]
    
    eff = leads.query('status not in @filter_status')
    eff = eff[eff['date'].between(x_1, x_2, inclusive=True)]
    
    soed = leads.query('status not in @filter_status')
    soed = soed[soed['date'].between(x_1, x_2, inclusive=True)]
    soed = soed[soed['date_op'].between(x_1, x_2, inclusive=True)]
    
    sm = leads.query('lend == "synergymobileapp" & status not in @filter_status')
    sm = sm[sm['date'].between(x_1, x_2, inclusive=True)]
    
    sm_soed = leads.query('lend == "synergymobileapp" & status not in @filter_status')
    sm_soed = sm_soed[sm_soed['date'].between(x_1, x_2, inclusive=True)]
    sm_soed = sm_soed[sm_soed['date_op'].between(x_1, x_2, inclusive=True)]
    
    gross_22 = leads_2022[leads_2022['date'].between(y_1, y_2, inclusive=True)]
    
    eff_22 = leads_2022.query('status not in @filter_status')
    eff_22 = eff_22[eff_22['date'].between(y_1, y_2, inclusive=True)]
    
    soed_22 = leads_2022.query('status not in @filter_status')
    soed_22 = soed_22[soed_22['date'].between(y_1, y_2, inclusive=True)]
    soed_22 = soed_22[soed_22['date_op'].between(y_1, y_2, inclusive=True)] 
    
    sm_22 = leads_2022.query('lend == "synergymobileapp" & status not in @filter_status')
    sm_22 = sm_22[sm_22['date'].between(y_1, y_2, inclusive=True)]
    
    vh_22 = vhod_2022.query('status not in @filter_status')
    vh_22 = vh_22[vh_22['date'].between(y_1, y_2, inclusive=True)]
    
    vh = vhod.query('status not in @filter_status')
    vh = vh[vh['date'].between(x_1, x_2, inclusive=True)]
    
    vh_soed = vhod.query('status not in @filter_status')
    vh_soed = vh_soed[vh_soed['date'].between(x_1, x_2, inclusive=True)]
    vh_soed = vh_soed[vh_soed['date_op'].between(x_1, x_2, inclusive=True)]
    
    pivot_table_1 = gross.pivot_table(index='tag', values='id', aggfunc='count')
    pivot_table_2 = eff.pivot_table(index='tag', values='id', aggfunc='count')
    pivot_table_3 = soed.pivot_table(index='tag', values='id', aggfunc='count')
    pivot_table_4 = sm.pivot_table(index='lend', values='id', aggfunc='count')
    pivot_table_5 = sm_soed.pivot_table(index='lend', values='id', aggfunc='count')
    pivot_table_6 = vh.pivot_table(index='tag', values='id', aggfunc='count')
    pivot_table_7 = vh_soed.pivot_table(index='tag', values='id', aggfunc='count')
    
    pivot_table_8 = gross_22.pivot_table(index='tag', values='id', aggfunc='count')
    pivot_table_9 = eff_22.pivot_table(index='tag', values='id', aggfunc='count')
    pivot_table_10 = sm_22.pivot_table(index='lend', values='id', aggfunc='count')
    pivot_table_11 = vh_22.pivot_table(index='tag', values='id', aggfunc='count')
    pivot_table_12 = soed_22.pivot_table(index='tag', values='id', aggfunc='count')
    
    #Основной сводник
    
    if len(pivot_table_1) > 0 and len(pivot_table_2) > 0:
        pivot_table_day = pivot_table_1.merge(pivot_table_2, on='tag', how='outer').reset_index()
        pivot_table_day.columns = ['tag', 'gross', 'eff']
        if len(pivot_table_3) > 0:
            pivot_table_day = pivot_table_day.merge(pivot_table_3, on='tag', how='outer')
            pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed']
            if len(pivot_table_6) > 0:
                pivot_table_day = pivot_table_day.merge(pivot_table_6, on='tag', how='outer')
                pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh']
                if len(pivot_table_7) > 0:
                    pivot_table_day = pivot_table_day.merge(pivot_table_7, on='tag', how='outer')
                    pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed']
                else:
                    pivot_table_day = pivot_table_day.merge(pivot_table_7, on='tag', how='outer')
                    pivot_table_day['vh_soed'] = np.nan
                    pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed']
            else:        
                pivot_table_day = pivot_table_day.merge(pivot_table_6, on='tag', how='outer')
                pivot_table_day['vh'] = np.nan
                pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh']
                pivot_table_day = pivot_table_day.merge(pivot_table_7, on='tag', how='outer')
                pivot_table_day['vh_soed'] = np.nan
                pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed']
        else:
            pivot_table_day = pivot_table_day.merge(pivot_table_3, on='tag', how='outer')
            pivot_table_day['soed'] = np.nan
            pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed']
            pivot_table_day = pivot_table_day.merge(pivot_table_6, on='tag', how='outer')
            pivot_table_day['vh'] = np.nan
            pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh']
            pivot_table_day = pivot_table_day.merge(pivot_table_7, on='tag', how='outer')
            pivot_table_day['vh_soed'] = np.nan
            pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed']
    else:
        pivot_table_day = pivot_table_1.merge(pivot_table_2, on='tag', how='outer').reset_index()
        pivot_table_day[['gross', 'eff']] = np.nan
        pivot_table_day.columns = ['tag', 'gross', 'eff']
        pivot_table_day = pivot_table_day.merge(pivot_table_3, on='tag', how='outer')
        pivot_table_day['soed'] = np.nan
        pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed']
        pivot_table_day = pivot_table_day.merge(pivot_table_6, on='tag', how='outer')
        pivot_table_day['vh'] = np.nan
        pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh']
        pivot_table_day = pivot_table_day.merge(pivot_table_7, on='tag', how='outer')
        pivot_table_day['vh_soed'] = np.nan
        pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed']
        
    #Подсоединение 2022 к основному своднику
        
    if len(pivot_table_8) > 0:
        pivot_table_day = pivot_table_day.merge(pivot_table_8, on='tag', how='outer')
        pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22']
        if len(pivot_table_9) > 0:
            pivot_table_day = pivot_table_day.merge(pivot_table_9, on='tag', how='outer')
            pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22']
            if len(pivot_table_11) > 0:
                pivot_table_day = pivot_table_day.merge(pivot_table_11, on='tag', how='outer')
                pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22', 'vh_22']
                if len(pivot_table_12) > 0:
                    pivot_table_day = pivot_table_day.merge(pivot_table_12, on='tag', how='outer')
                    pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22', 'vh_22', 'soed_22']
                else:
                    pivot_table_day = pivot_table_day.merge(pivot_table_12, on='tag', how='outer')
                    pivot_table_day['soed_22'] = np.nan
                    pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22', 'vh_22', 'soed_22']
            else:
                pivot_table_day = pivot_table_day.merge(pivot_table_11, on='tag', how='outer')
                pivot_table_day['vh_22'] = np.nan
                pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22', 'vh_22']
                pivot_table_day = pivot_table_day.merge(pivot_table_12, on='tag', how='outer')
                pivot_table_day['soed_22'] = np.nan
                pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22', 'vh_22', 'soed_22', '0']
        else:
            pivot_table_day = pivot_table_day.merge(pivot_table_9, on='tag', how='outer')
            pivot_table_day['eff_22'] = np.nan
            pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22']
            pivot_table_day = pivot_table_day.merge(pivot_table_11, on='tag', how='outer')
            pivot_table_day['vh_22'] = np.nan
            pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22', 'vh_22']
            pivot_table_day = pivot_table_day.merge(pivot_table_12, on='tag', how='outer')
            pivot_table_day['soed_22'] = np.nan
            pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22', 'vh_22', 'soed_22']
    else:
        pivot_table_day = pivot_table_day.merge(pivot_table_8, on='tag', how='outer')
        pivot_table_day['gross_22'] = np.nan
        pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22']
        pivot_table_day = pivot_table_day.merge(pivot_table_9, on='tag', how='outer')
        pivot_table_day['eff_22'] = np.nan
        pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22']  
        pivot_table_day = pivot_table_day.merge(pivot_table_11, on='tag', how='outer')
        pivot_table_day['vh_22'] = np.nan
        pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22', 'vh_22']    
        pivot_table_day = pivot_table_day.merge(pivot_table_12, on='tag', how='outer')
        pivot_table_day['soed_22'] = np.nan
        pivot_table_day.columns = ['tag', 'gross', 'eff', 'soed', 'vh', 'vh_soed', 'gross_22', 'eff_22', 'vh_22', 'soed_22']    
    #сводная таблица по СМ
        
    if len(pivot_table_4) > 0 and len(pivot_table_5) > 0:
        pivot_table_sm_day = pivot_table_4.merge(pivot_table_5, on='lend', how='outer').reset_index()
        pivot_table_sm_day.columns = ['lend', 'eff', 'soed']
    else:
        pivot_table_sm_day = pivot_table_4.merge(pivot_table_5, on='lend', how='outer').reset_index()
        pivot_table_sm_day['soed'] = np.nan
        pivot_table_sm_day.columns = ['lend', 'eff', 'soed']
        
    #Подсоединение 2022 года к своднику по СМ
        
    if len(pivot_table_10) > 0:
        pivot_table_sm_day = pivot_table_sm_day.merge(pivot_table_10, on='lend', how='outer')
        pivot_table_sm_day.columns = ['tag', 'eff', 'soed', 'eff_22']
    else:
        pivot_table_sm_day = pivot_table_sm_day.merge(pivot_table_10, on='lend', how='outer')
        pivot_table_sm_day['eff_22'] = np.nan
        pivot_table_sm_day.columns = ['tag', 'eff', 'soed', 'eff_22']
        
    frames = [pivot_table_day, pivot_table_sm_day]
    result = pd.concat(frames).reset_index(drop=True)
    
    return result


result_monday = day_report(monday, monday_22)
result_tuesday = day_report(tuesday, tuesday_22)
result_wednesday = day_report(wednesday, wednesday_22)
result_thursday = day_report(thursday, thursday_22)
result_friday = day_report(friday, friday_22)
result_saturday = day_report(saturday, saturday_22)
result_sunday = day_report(sunday, sunday_22)
result_week = week_report(monday, monday_22, sunday, sunday_22)


# Коневерсия источников

wb = load_workbook(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Харитуня\КМВ\Реклама\Артём автоматизация\Источники.xlsx')

sheet = wb.get_sheet_by_name('Сводник_1')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(pivot_table_lead, index=True, header=True):
    sheet.append(r)

sheet = wb.get_sheet_by_name('Сводник_2')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(pivot_table_eff_lead, index=True, header=True):
    sheet.append(r)

sheet = wb.get_sheet_by_name('Сводник_3')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(pivot_table_чв, index=True, header=True):
    sheet.append(r)


wb.save(r"\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Харитуня\КМВ\Реклама\Артём автоматизация\Источники.xlsx")


# Отчёт по бюджету

wb = load_workbook(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Харитуня\КМВ\Реклама\Артём автоматизация\FileName\Отчет по бюджету РК.xlsx')

sheet = wb.get_sheet_by_name('Сводник_1')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(pivot, index=True, header=True):
    sheet.append(r)
    
sheet = wb.get_sheet_by_name('Сводник_2')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(pivot_table, index=True, header=True):
    sheet.append(r)


wb.save(f'\\\\synergy.local\\Documents\\19.Группа мониторинга и сопровождения сделок\\01.Отчеты\\Харитуня\\КМВ\\Реклама\\Артём автоматизация\\Отчет по бюджету РК {month}.xlsx')


# РК КД Неделя

wb = load_workbook(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Харитуня\КМВ\Реклама\Артём автоматизация\FileName\РК КД.xlsx')

sheet = wb.get_sheet_by_name('Сводник_1')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(result_week, index=True, header=True):
    sheet.append(r)

sheet['N1'] = pd.to_datetime(monday) #.strftime('%d.%m.%Y')
sheet['N2'] = pd.to_datetime(tuesday) #.strftime('%d.%m.%Y')
sheet['N3'] = pd.to_datetime(wednesday) #.strftime('%d.%m.%Y')
sheet['N4'] = pd.to_datetime(thursday) #.strftime('%d.%m.%Y')
sheet['N5'] = pd.to_datetime(friday) #.strftime('%d.%m.%Y')
sheet['N6'] = pd.to_datetime(saturday) #.strftime('%d.%m.%Y')
sheet['N7'] = pd.to_datetime(sunday) #.strftime('%d.%m.%Y')

sheet = wb.get_sheet_by_name('Сводник_2')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(pivot, index=True, header=True):
    sheet.append(r)

sheet = wb.get_sheet_by_name('Сводник_3')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(data_pivot_all, index=False, header=True):
    sheet.append(r)

wb.save(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Харитуня\КМВ\Реклама\Артём автоматизация\РК КД Неделя.xlsx')

print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Готово!")
print("")