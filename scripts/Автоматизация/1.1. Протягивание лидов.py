import os
import re
import win32com.client

import pandas as pd
import numpy as np
import openpyxl as ox

from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

from datetime import datetime, timedelta

import warnings
warnings.filterwarnings("ignore")

import string

import pyodbc


print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Лиды протягиваются, пожалуйста, подождите...")

data = pd.read_excel(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\БАЗА КД ОБЩАЯ 2018.xlsx', sheet_name='2018')

conn = pyodbc.connect(r'Driver={SQL Server};Server=MSK1-BIDB01;Database=DWH;Trusted_Connection=yes;')
cursor = conn.cursor()

data = data[['ID сделки', 'ленд']]

data = data.tail(500)

data = data.fillna(0)

data = data.loc[data['ID сделки'] != 0]

dt = data['ID сделки'].astype('str')

dt = [f", '{el}'" for el in data['ID сделки']]

dt = list(dt)

s = ''.join(dt)


request_string_leads = f'''

SELECT DISTINCT
	 --даты
	  CONVERT(varchar(10), COALESCE(D.DATE_CREATE, ASS_AR.Data, R.DATE_CREATE),104)  as "Дата создания" -- DATE_CREATE,
	  --,CONVERT(varchar(10), D.DATE_MODIFY, 104) as "Дата изменения"
	 -- ,CONVERT(varchar(10), D.CLOSE_DATE, 104)  as "Дата завершения"
	  --,R.[DATE_OP]
	  --маркетинговые метки
	 -- ,d.[UTM_CAMPAIGN]  as "Кампания"
	 ,src.NAME as "Источник 1"
	  ,R.FEATURES_1 "Источник"
      ,R.FEATURES_5  "Ленд"
	 -- ,R.[DATE_CREATE]
	 -- --клиент
	 -- ,D.name  as "Название сделки" --NAME_DEAL
	 -- ,concat(cc.[FAMILY]+' ',cc.[NAME]+' ',cc.[SURNAME]) as "Контакт"
	 -- ,cc.BIRTHDATE "Дата рождения"
	 -- ,isnull(PhHome.PHONE,'') "Телефон дом."
	 -- ,isnull(PhMob.PHONE,'') "Телефон моб."
	 -- ,isnull(PhJob.PHONE,'') "Телефон раб."
	 -- ,isnull(PhOst.PHONE,'') "Телефон проч."
	 -- ,coalesce(cc.EMAIL_HOME,cc.EMAIL_WORK) as "e-mail"
	 -- --продукт
	 -- ,cp.[NAME] "Продукт сделки"
	 -- ,dgp.NAME as "Группа продуктов"
	 -- ,PB.[Name] as "Продукт бюджета"
	 -- ,CAST(ASS_P.[QUANTITY] AS INT) AS "Количество продукта"
 -- ,CAST(D.AMOUNT / COUNT(*) OVER (PARTITION BY D.ID_DEAL)  as decimal(33,0)) "Сумма"
	 -- ,CAST(D.AMOUNT as decimal(33,0)) as "СуммаСделки"	      
	 -- ,IIF(isnull(D.SIGN_MAT_CAP,0)=1, 'Да', '') "Материнский капитал в сделке"
	 -- --вероятность+прослушка
	-- ,D.[PROBABILITY] "Вероятность"
	 -- ,replace(replace(isnull(utsD.[AUDITION_ROP],'0'),'0','нет'),'1','да') as "Прослушано РОП"
	 -- --статус
	--,sd.NAME as "Стадия сделки"
	 -- --ответственный	 
	 -- ,concat(e.[LAST_NAME]+' ',e.[NAME]+' ',e.[SECOND_NAME]) as "Ответственный"
	--  ,org.full_NAME "Структурное подразделение"
	 -- ,D.ID_CFO_VISUAL  as "Визуал"
	 -- --инвойсы
	 --  --,CONVERT(varchar(10), D_I.[DATE_CREATE],104) as "Дата создания инвойса"
	 --  ,CONVERT(varchar(10), D_I.[DATE_PAY],104) as "Дата оплаты"
  --     ,CAST(D_I.[AMOUNT] as decimal(33,0)) as "Сумма инвойса"
	 ----  ,S_I.[NAME] as "Статус инвойса"
	 --  --дела
		--,CONVERT(varchar(10), CAST(CA.[CREATED] as date),104) as "Дата создания дела"
		--,CONVERT(varchar(10), CAST(CA.[START_TIME] as date),104) as "Дата дела"
	 --   ,CAST(CA.[START_TIME] as time(0)) as "Время дела"
		--,CA.[COMPLETED] as "Статус дела"
		--,utsD.[CODE_LABEL_BASE_KD]
	 --  --id
	  ,D.CODE as "id deal crm"
	  ,R.[CODE] as "id lead crm"
	  --,R.[ID_REQUEST] as "id lead base"
	  --,RPB.[ID_PRODUCT_BUDGET] as "id продукт бюджета"	  
	  --,cc.ID_CLIENT_CRM as "id контакта"
	  --,ASS_D_I.[ID_INVOICE] as "id инвойса base"
	  --,D_I.[CODE] as "id инвойса crm"
	  


FROM [DWH].[dbo].DIC_DEAL D
		  LEFT JOIN [DWH].[dbo].[DIC_EMPLOYEES] E ON D.ID_EMPLOYEES = E.ID_EMPLOYEES
		  LEFT JOIN [DWH].[dbo].[ASS_EMPLOYEE_AND_ORGSTRUCTURE] ASS_OS ON E.ID_EMPLOYEES = ASS_OS.ID_EMPLOYEES
		  LEFT JOIN [DWH].[dbo].[v_DIC_ORGSTRUCTURE] org on org.ID_ORGSTRUCTURE = ASS_OS.ID_ORGSTRUCTURE
		 
		  --подсоединение заявок
		  LEFT JOIN [DWH].[dbo].ASS_REQUEST_DEAL ASS_RD ON ASS_RD.ID_DEAL = D.ID_DEAL
		  LEFT JOIN [DWH].[dbo].[DIC_REQUEST] R ON ASS_RD.ID_REQUEST = R.ID_REQUEST

		   -- источник
                
            LEFT JOIN [DWH].[STAGE].[CRM_B_UTS_CRM_LEAD] UL WITH(NOLOCK) ON UL.VALUE_ID = R.CODE
            LEFT JOIN [DWH].[STAGE].[CRM_B_CRM_LEAD] L WITH(NOLOCK) ON L.ID = R.CODE
            INNER JOIN (SELECT [NAME], [STATUS_ID]
		                FROM [DWH].[STAGE].[CRM_B_CRM_STATUS] WITH(NOLOCK)
		                WHERE [ENTITY_ID] = 'SOURCE') SRC ON SRC.[STATUS_ID] = L.[SOURCE_ID]		
       
                    
                    
		    --объявления
		  LEFT JOIN [DWH].[dbo].ASS_AD_REQUEST ASS_AR ON ASS_AR.[ID_TYPE_ASS_AD_REQUEST] = 0   --Связь с учетом всех коректировок
										 AND ASS_AR.ID_REQUEST = R.ID_REQUEST   
		  LEFT JOIN [DWH].[dbo].DIC_AD AD ON AD.ID_AD = ASS_AR.ID_AD
		  LEFT JOIN [DWH].[dbo].DIC_AD_GROUP AG ON AD.ID_AD_GROUP = AG.ID_AD_GROUP
		  LEFT JOIN [DWH].[dbo].DIC_AD_CAMPAIGN AC ON AG.ID_AD_CAMPAIGN = AC.ID_AD_CAMPAIGN
   		  --подсоединение инвойсов к сделке
		  LEFT JOIN [DWH].[dbo].ASS_DEAL_INVOICE ASS_D_I ON ASS_D_I.ID_DEAL = D.ID_DEAL
		  LEFT JOIN [DWH].[dbo].DIC_INVOICE D_I ON ASS_D_I.ID_INVOICE = D_I.ID_INVOICE
		  LEFT JOIN [DWH].[dbo].DIC_STATUS_INVOICE S_I ON S_I.ID_STATUS_INVOICE = D_I.ID_STATUS_INVOICE
		  --каталог продуктов сделки
		  LEFT JOIN [DWH].[dbo].ASS_DEAL_CATALOG_PRODUCT ASS_P ON D.ID_DEAL = ASS_P.ID_DEAL
		  LEFT JOIN [DWH].[dbo].[DIC_CATALOG_PRODUCT] cp on cp.ID_CATALOG_PRODUCT = ASS_P.ID_CATALOG_PRODUCT
		  -- статус сделки
		  left join [DWH].[dbo].[DIC_STATUS_DEAL] sd on sd.ID_STATUS_DEAL = d.ID_STATUS_DEAL
		  --contact
		  left join [DWH].[dbo].[DIC_CLIENT_CRM] cc on cc.[ID_CLIENT_CRM] = d.[ID_CLIENT_CRM]
		  left join [DWH].[dbo].[DIC_CLIENT_CRM_PHONE] PhHome ON PhHome.[ID_CLIENT_CRM]=d.[ID_CLIENT_CRM] AND PhHome.[ID_TYPE_CLIENT_CRM_PHONE]=1	--	Домашний
		  left join [DWH].[dbo].[DIC_CLIENT_CRM_PHONE] PhMob ON PhMob.[ID_CLIENT_CRM]=d.[ID_CLIENT_CRM] AND PhMob.[ID_TYPE_CLIENT_CRM_PHONE]=2		--Мобильный
		  left join [DWH].[dbo].[DIC_CLIENT_CRM_PHONE] PhJob ON PhJob.[ID_CLIENT_CRM]=d.[ID_CLIENT_CRM] AND PhJob.[ID_TYPE_CLIENT_CRM_PHONE]=5		--Рабочий
		  left join [DWH].[dbo].[DIC_CLIENT_CRM_PHONE] PhOst ON PhOst.[ID_CLIENT_CRM]=d.[ID_CLIENT_CRM] AND PhOst.[ID_TYPE_CLIENT_CRM_PHONE]=3		--Прочий
		  -- group product deal
		  left join [DWH].[dbo].[DIC_GROUP_PRODUCT_DEAL] dgp on dgp.[ID_GROUP_PRODUCT_DEAL] = d.[GROUP_PRODUCT_ID]
		  -- расширение свойств сделок
		  left join [DWH].[dbo].[DIC_DEAL_STAT] utsD ON D.[ID_DEAL]=utsD.[ID_DEAL]
		  -- продукт бюджета
			LEFT JOIN [DWH].[dbo].ASS_REQUEST_PRODUCT_BUDGET RPB With(nolock) ON R.ID_REQUEST=RPB.ID_REQUEST and R.DATE_CREATE= RPB.R_DATE
			LEFT JOIN [DWH].[dbo].DIC_PRODUCT_BUDGET PB With(nolock) ON RPB.ID_PRODUCT_BUDGET=PB.ID_PRODUCT_BUDGET
			--дела
			LEFT JOIN [DWH].[stage].[CRM_b_crm_act] CA ON CA.[OWNER_ID] = D.[CODE]
		  
  
WHERE
  D.ID_DEAL <> - 1
  and D.CODE in ('5599690' {s})

'''


leads = pd.read_sql_query(request_string_leads, conn)

leads = leads[['id deal crm', 'Дата создания', 'Источник 1', 'Источник', 'Ленд', 'id lead crm']]


leads['Источник'] = leads['Источник'].fillna('0')
leads_2 = leads.query('Источник == "0"')
leads = leads.query('Источник != "0"')


leads_2['Источник'] = leads_2['Источник 1']
leads_2['Источник'] = leads_2['Источник'].map({'Аккаунтинг/Accounting': 'свой контакт', 
                                               'Входящий звонок / Incoming call': 'входящий звонок', 
                                               'Веб-сайт / Website':'органик', 
                                               'Импорт / Import':'импорт'})

leads = leads.append(leads_2, ignore_index=True)


leads['Ленд'] = leads['Ленд'].fillna('0')
leads_2 = leads.query('Ленд == "0"')
leads = leads.query('Ленд != "0"')


leads_2['Ленд'] = leads_2['Источник 1']
leads_2['Ленд'] = leads_2['Ленд'].map({'Аккаунтинг/Accounting': 'свой контакт', 
                                        'Входящий звонок / Incoming call': 'входящий звонок',  
                                        'Импорт / Import':'импорт'})

leads = leads.append(leads_2, ignore_index=True)

leads = leads.rename(columns={'id deal crm': 'ID сделки'})
leads = leads.merge(data, how="right", on ="ID сделки")

leads[['ID сделки', 'id lead crm']] = leads[['ID сделки', 'id lead crm']].fillna(0)
leads[['ID сделки', 'id lead crm']] = leads[['ID сделки', 'id lead crm']].astype(int)
leads = leads[['ID сделки', 'Дата создания', 'Источник 1', 'Источник', 'Ленд', 'id lead crm']]

wb = load_workbook(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\leads.xlsx')

sheet = wb.get_sheet_by_name('Sheet_1')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(data, index=False, header=True):
    sheet.append(r)

sheet = wb.get_sheet_by_name('Sheet_2')
sheet.delete_rows(1, sheet.max_row)    

for r in dataframe_to_rows(leads, index=False, header=True):
    sheet.append(r)


wb.save(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\leads.xlsx')

#client = win32com.client.Dispatch("Excel.Application")
#client.Visible = False
#wb = client.Workbooks.Open(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\leads.xlsx')
#ws = wb.Worksheets('Отчёт')
#ws.Range("C2:F501").Copy()
#wb.Close()
#client.Quit()
#
#
#client = win32com.client.Dispatch("Excel.Application")
#client.Visible = False
#wb = client.Workbooks.Open(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\leads.xlsx')
#ws = wb.Worksheets('Отчёт')
#ws.Range("A2").Copy()
#wb.Close()
#client.Quit()

print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Готово!")
print("")