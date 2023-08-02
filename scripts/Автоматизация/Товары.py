import os
from tqdm import tqdm

import pandas as pd
import numpy as np
import openpyxl as ox
import re

from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

from datetime import datetime, timedelta

import warnings
warnings.filterwarnings("ignore")

import string

import pyodbc

print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Расшифровываю товары, пожалуйста, подождите...")

data = pd.read_excel(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Внесение оплат\Песочница.xlsx', sheet_name='2')

data["Товар"] = data["Товар"].fillna("0")




for i in data['Товар']:
    
#------------------------------ТОВАР 1--------------------------------\ 
    
    if "БАК" in i and "аочн" in i:
        data.loc[data['Товар'] == i, ['t']] = "впо-дист"
    elif "БАК" in i and "вечер" in i:
        data.loc[data['Товар'] == i, ['t']] = "впо"
    elif "БАК" in i and "вых" in i:
        data.loc[data['Товар'] == i, ['t']] = "впо"        
    elif "БАК" in i and ("классическ") in i:
        data.loc[data['Товар'] == i, ['t']] = "впо"
    elif "ДОК" in i and "аочн" in i:
        data.loc[data['Товар'] == i, ['t']] = "впо-дист"
    elif "ДОК"  in i and ("классическ") in i:
        data.loc[data['Товар'] == i, ['t']] = "впо"
    elif "ДОК" in i and "вечер" in i:
        data.loc[data['Товар'] == i, ['t']] = "впо"
    elif "ДОК" in i and "вых" in i:
        data.loc[data['Товар'] == i, ['t']] = "впо"
        
#-----------------------------------------------------------------\   БАК 

    elif "МАГ" in i and "аочн" in i:
        data.loc[data['Товар'] == i, ['t']] = "маг-дист"
    elif "МАГ" in i and "вых" in i:
        data.loc[data['Товар'] == i, ['t']] = "маг"        
    elif "МАГ"  in i and ("классическ") in i:
        data.loc[data['Товар'] == i, ['t']] = "маг" 
    elif "МАГ" in i and "вечер" in i:
        data.loc[data['Товар'] == i, ['t']] = "маг"
        
#-----------------------------------------------------------------\   МАГ

    elif "СПО" in i and "аочн" in i:
        data.loc[data['Товар'] == i, ['t']] = "спо-дист"
    elif "СПО" in i and "вых" in i:
        data.loc[data['Товар'] == i, ['t']] = "спо"
    elif "СПО"  in i and ("классическ") in i:
        data.loc[data['Товар'] == i, ['t']] = "спо" 
    elif "СПО" in i and "вечер" in i:
        data.loc[data['Товар'] == i, ['t']] = "спо"
    elif "СПЦ" in i and "аочн" in i:
        data.loc[data['Товар'] == i, ['t']] = "спо-дист"
    elif "СПЦ"  in i and ("классическ") in i:
        data.loc[data['Товар'] == i, ['t']] = "спо"
    elif "СПЦ" in i and "вечер" in i:
        data.loc[data['Товар'] == i, ['t']] = "спо"
        
#-----------------------------------------------------------------\  СПО

    elif "АСП" in i and "аочн" in i:
        data.loc[data['Товар'] == i, ['t']] = "асп"
    elif "АСП"  in i:
        data.loc[data['Товар'] == i, ['t']] = "асп" 
        
#-----------------------------------------------------------------\   АСП

    elif "СОО" in i and "кола" in i:
        data.loc[data['Товар'] == i, ['t']] = "школа"
        
#-----------------------------------------------------------------\   школа

    elif "SYNERGY" in i:
        data.loc[data['Товар'] == i, ['t']] = "курсы"     
        
#-----------------------------------------------------------------\   SYNERGY

    elif "Годовая" in i or "Школа" in i:
        data.loc[data['Товар'] == i, ['t']] = "курсы"     
        
#-----------------------------------------------------------------\   Годовая подписка

    elif "ДО" in i and "ДОК" not in i:
        data.loc[data['Товар'] == i, ['t']] = "курсы"
        
#-----------------------------------------------------------------\   Прочие курсы
        
    elif "Курс" in i:
        data.loc[data['Товар'] == i, ['t']] = "курсы" 
    
    else:
        data.loc[data['Товар'] == i, ['t']] = "0"
        
#-----------------------------------------------------------------\   Курсы

for i in data['Товар']:

#------------------------------ТОВАР 2--------------------------------\ 

    if "ВВС" in i or "ввс" in i:
        data.loc[data['Товар'] == i, ['t2']] = "ввс"
        
#-----------------------------------------------------------------\   Курсы
        
    elif "ВЕБ" in i or "веб" in i and "класс" not in i:
        data.loc[data['Товар'] == i, ['t2']] = "веб"

#-----------------------------------------------------------------\   Курсы

    elif "МАП" in i or "МОСАП" in i:
        data.loc[data['Товар'] == i, ['t2']] = "МОСАП"

#-----------------------------------------------------------------\   Курсы

    elif "едицин" in i:
        data.loc[data['Товар'] == i, ['t2']] = "мед"
        
#-----------------------------------------------------------------\   Курсы        

    elif "МОИ" in i or "МТИ" in i:
        data.loc[data['Товар'] == i, ['t2']] = "МОИ"
        
#-----------------------------------------------------------------\   Курсы        
        
    elif "театр" in i or "актер" in i:
        data.loc[data['Товар'] == i, ['t2']] = "НТШ"
        
#-----------------------------------------------------------------\   Курсы        
        
    elif "2 в 1" in i:
        data.loc[data['Товар'] == i, ['t2']] = "2в1"
        
#-----------------------------------------------------------------\   Курсы        
        
    elif "чтени" in i:
        data.loc[data['Товар'] == i, ['t2']] = "СЧ"
        
#-----------------------------------------------------------------\   Курсы        
        
    elif "Сверхпамять" in i:
        data.loc[data['Товар'] == i, ['t2']] = "РП"
        
#-----------------------------------------------------------------\   Курсы        
        
    elif "сзн" in i:
        data.loc[data['Товар'] == i, ['t2']] = "сзн"
        
#-----------------------------------------------------------------\   Курсы        

    elif "цзн" in i:
        data.loc[data['Товар'] == i, ['t2']] = "цзн"
        
#-----------------------------------------------------------------\   Курсы

    elif "Проф" in i:
        data.loc[data['Товар'] == i, ['t2']] = "курсы"
        
#-----------------------------------------------------------------\   Курсы 
        
    else:
        data.loc[data['Товар'] == i, ['t2']] = ""
        
#------------------------------МЕНЕДЖЕР--------------------------------\ 

#def search_partial_text(src, dst):
#    dst_buf = dst
#    result = 0
#    for char in src:
#        if char in dst_buf:
#            dst_buf = dst_buf.replace(char, '', 1)
#            result += 1
#    r1 = int(result / len(src) * 100)
#    r2 = int(result / len(dst) * 100)
#    return r1 if r1 < r2 else r2
#
#замены = []

#-----------------------------------------------------------------\

#conn = pyodbc.connect(r'Driver={SQL Server};Server=MSK1-BIDB01;Database=DWH;Trusted_Connection=yes;')
#cursor = conn.cursor()
#
#employees = f'''
#
#SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED
#
#SELECT [EMPLOYEES]
#FROM [DWH].[dbo].[KHTS_EMPL]
#WHERE 
#   [SP] in ('КД')
#   AND [STATUS] = 'работает'
#
#'''
#
#empl = pd.read_sql_query(employees, conn)
#td = pd.DataFrame()
#td['МП'] = empl['EMPLOYEES']

#-----------------------------------------------------------------\
        
#names = {0: 'name_1', 1: 'name_2', 2: 'name_3', 3: 'name_4', 4: 'name_5'}
#
#data_names = data['МЕНЕДЖЕР'].str.split(' ', expand=True).rename(columns=names).fillna('')
#data_names = data_names[['name_1', 'name_2', 'name_3', 'name_4']]
#
#data[['name_1', 'name_2', 'name_3', 'name_4']] = data_names[['name_1', 'name_2', 'name_3', 'name_4']]
#
#data['МП'] = data['name_1'] + " " + data['name_2']   + " " + data['name_3'] #+ " " + data['name_4']

#-----------------------------------------------------------------\ 

#for i in tqdm(data['МП']):
#
#    s1 = i
#    t9_2 = []
#    t9_k = []
#    one = data.loc[data['МП'] == i, ['МП']].iloc[0, 0]
#
#    for index in td['МП']:
#        s2 = index
#        two = td.loc[td['МП'] == index, ['МП']].iloc[0, 0]
#        search_partial = search_partial_text(s1, s2)
#
#        if search_partial >= 90:
#            t9_2.append(two)
#            t9_k.append(search_partial)
#
#    t9 = pd.DataFrame({'2': t9_2,
#                       'k': t9_k})
#
#    if len(t9) > 1:
#
#        tr = t9.sort_values(by='k', ascending=False).iloc[0, 1]
#        fl = t9.sort_values(by='k', ascending=False).iloc[1, 1]
#        mp = t9.sort_values(by='k', ascending=False).iloc[0, 0]
#
#        if tr != fl:
#            data.loc[data['МП'] == s1, ['МП']] = mp
#            if s1 != mp:
#                замены.append({s1: mp})
#
#    elif len(t9) == 1:
#
#        mp = t9.iloc[0, 0]
#        data.loc[data['МП'] == s1, ['МП']] = mp
#        if s1 != mp:
#            замены.append({s1: mp})
#pass
#
#print('')
#print("Произведённые замены фамилий:")
#print('')
#print(замены)

#-----------------------------------------------------------------\ 
        
data = data[['t', 't2']]

#-----------------------------------------------------------------\
        
wb = load_workbook(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Внесение оплат\Товары.xlsx')

sheet = wb.get_sheet_by_name('Sheet_2')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(data, index=False, header=True):
    sheet.append(r)
    
wb.save(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Внесение оплат\Товары.xlsx')

print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Готово!")
print("")