import os

import pandas as pd
import numpy as np
import openpyxl as ox
import re
from win32com import client

from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

from datetime import datetime, timedelta

import warnings
warnings.filterwarnings("ignore")

import string

import pyodbc

#now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Обновление отчёта по распределению лидов на МП, пожалуйста, подождите...")

day_m = datetime.today()

d = day_m.day
m = day_m.month

if m == 1:
    month = 'Январь'
elif m == 2:
    month = 'Февраль'
elif m == 3:
    month = 'Март'
elif m == 4:
    month = 'Апрель'
elif m == 5:
    month = 'Май'
elif m == 6:
    month = 'Июнь'
elif m == 7:
    month = 'Июль'
elif m == 8:
    month = 'Август'
elif m == 9:
    month = 'Сентябрь'
elif m == 10:
    month = 'Октябрь'
elif m == 11:
    month = 'Ноябрь'
else:
    month = 'Декабрь'
    

conn = pyodbc.connect(r'Driver={SQL Server};Server=MSK1-BIDB01;Database=DWH;Trusted_Connection=yes;')
cursor = conn.cursor()

leads_eff = f'''

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

declare @date date = getdate()-1
declare @month int = month(getdate()-1)
declare @startdate date = DATEFROMPARTS(2022, @month, 01)
declare @endate date = @date;

with t as(
select distinct
dr.[DATE_CREATE]    as "DATE"
,dr.[DATE_OP] as "Дата соединения на ОП"
,year(dr.[DATE_CREATE]) as "year"

,[FEATURES_6]  as "Старый/новый"       
,dr.[FEATURES_5] as "Ленд"
,CASE

    WHEN s.tag1 like '%органика%' THEN 'органика'
    WHEN s.tag1 like '%yandex%' THEN 'Яндекс'
    WHEN s.tag1 like '%google%' THEN 'Google'
    WHEN s.tag1 like '%facebook%' THEN 'facebook'
    WHEN s.tag1 like '%edunetwork%' THEN 'edunetwork'
    WHEN s.tag1 like '%(пусто)%' THEN 'органика'
    WHEN s.tag1 like '%TikTok%' THEN 'TikTok'
    WHEN s.tag1 like '%рассылка%' THEN 'рассылка'
    WHEN dr.[FEATURES_1] like '%studika%' THEN 'studika'
    
    ELSE 'прочее'
END as "tag"


,dr.[FEATURES_1] as "source"
,fc.[SPENT_BY_REQ] as "Расход"
,iif(dsr.[NAME] not like '%/%',dsr.[NAME], SUBSTRING(dsr.[NAME],0,PATINDEX('% / %',dsr.[NAME])))   as "Статус"              
,concat(e.[LAST_NAME]+' ',e.[NAME]+' ',e.[SECOND_NAME]) as "Ответственный"
,CASE

    WHEN org.[NAME] like '%центр 1%' THEN 'КЦ 1'
    WHEN org.[NAME] like '%центр 2%' THEN 'КЦ 2'
    WHEN org.[NAME] like '%центр 3%' THEN 'КЦ 3'
    WHEN org.[NAME] like '%центр 4%' THEN 'КЦ 4'
    
    ELSE 'др'
  END as "СП"

,RU.[Marketer] as Marketer
,dr.[CODE] as id  
  

from [DWH].[dbo].[DIC_REQUEST] dr 

            left join [DWH].[dbo].[DIC_REQUEST_UTM] RU on RU.[ID_REQUEST] = DR.[ID_REQUEST]
-- статусы
            left join  [DWH].[dbo].[DIC_STATUS_REQUEST] dsr on dsr.ID_STATUS_REQUEST = dr.ID_STATUS_REQUEST
-- источник

LEFT JOIN [DWH].[STAGE].[CRM_B_UTS_CRM_LEAD] UL WITH(NOLOCK) ON UL.VALUE_ID = dr.CODE
LEFT JOIN [DWH].[STAGE].[CRM_B_CRM_LEAD] L WITH(NOLOCK) ON L.ID = dr.CODE
INNER JOIN (SELECT [NAME], [STATUS_ID]
           FROM [DWH].[STAGE].[CRM_B_CRM_STATUS] WITH(NOLOCK)
           WHERE [ENTITY_ID] = 'SOURCE' AND (STATUS_ID='WEB' or STATUS_ID= '7')
           ) SRC ON SRC.[STATUS_ID] = L.[SOURCE_ID]     

              LEFT JOIN ASS_REQUEST_SOURCE ARS WITH(NOLOCK) ON dR.ID_REQUEST = ARS.ID_REQUEST
              LEFT JOIN DIC_SOURCE S WITH(NOLOCK) ON ARS.ID_SOURCE = S.ID_SOURCE
-- исполнители
LEFT JOIN [DWH].[dbo].DIC_EMPLOYEES E With(nolock) ON dr.Id_EMPLOYEES = E.ID_EMPLOYEES
JOIN (SELECT * FROM  [DWH].[dbo].[v_DIC_ORG_EMPL] where [lv1_NAME] like '%Коммерческий департамент (КМВ)%') org on org.[ID_EMPLOYEES] = E.ID_EMPLOYEES

-- продукт бюджета
            LEFT JOIN [DWH].[dbo].ASS_REQUEST_PRODUCT_BUDGET RPB With(nolock) ON dR.ID_REQUEST=RPB.ID_REQUEST and dR.DATE_CREATE= RPB.R_DATE
            LEFT JOIN [DWH].[dbo].DIC_PRODUCT_BUDGET PB With(nolock) ON RPB.ID_PRODUCT_BUDGET=PB.ID_PRODUCT_BUDGET
--расход
            LEFT JOIN [DWH].[dbo].[FCT_REQUEST] fc With(nolock) on fc.[ID_REQUEST] = dR.ID_REQUEST

where 1=1 

and isnull(dr.SIGN_DELETED,0) =0
and year(dr.[DATE_CREATE]) in ('2021', '2022','2023')
and month(dr.[DATE_CREATE]) = @month
and day(dr.[DATE_CREATE]) <= day(getdate()-1)
--and org.full_NAME like '%Коммерческий департамент (КМВ)\\%'
and org.[NAME] not like '%МАП%'
--and src.name_source like '%Веб%'
aND dsr.[NAME] NOT IN ('Дубль / Double','Спам / Spam','Ошибка номера / Error number', 'Спам / Spam','Повторные заявки / Reapplication')
and dr.[FEATURES_5] not in ('estr2022' , 'ege_rf', 'leadform_CZ_747828', 'leadform_CZN_SYN_PP_713311','proftest')
)

select
t.[year]
, count (t.id) as "Кол"

from t
group by 
t.[year]

'''

leads = f'''

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

declare @date date = getdate()-1
declare @month int = month(getdate())
declare @startdate date = DATEFROMPARTS(2022, @month, 01)
declare @endate date = @date;

select distinct
              
dr.[DATE_CREATE]    as "DATE"
,dr.[DATE_OP] as "Дата соединения на ОП"
,year(dr.[DATE_OP]) as "Год"

,[FEATURES_6]  as "Старый/новый"       
,dr.[FEATURES_5] as "Ленд"

,CASE

    WHEN s.tag1 like '%органика%' THEN 'органика'
    WHEN s.tag1 like '%yandex%' THEN 'Яндекс'
    WHEN s.tag1 like '%google%' THEN 'Google'
    WHEN s.tag1 like '%facebook%' THEN 'facebook'
    WHEN s.tag1 like '%edunetwork%' THEN 'edunetwork'
    WHEN s.tag1 like '%(пусто)%' THEN 'органика'
    WHEN s.tag1 like '%TikTok%' THEN 'TikTok'
    WHEN s.tag1 like '%рассылка%' THEN 'рассылка'
    WHEN dr.[FEATURES_1] like '%studika%' THEN 'studika'
    
    ELSE 'прочее'
  END as "tag"


,dr.[FEATURES_1] as "source"
,fc.[SPENT_BY_REQ] as "Расход"
,iif(dsr.[NAME] not like '%/%',dsr.[NAME], SUBSTRING(dsr.[NAME],0,PATINDEX('% / %',dsr.[NAME])))   as "Статус"              
,concat(e.[LAST_NAME]+' ',e.[NAME]+' ',e.[SECOND_NAME]) as "Ответственный"
--,CASE
--
--    WHEN org.[NAME] like '%центр 1%' THEN 'КЦ 1'
--    WHEN org.[NAME] like '%центр 2%' THEN 'КЦ 2'
--    WHEN org.[NAME] like '%центр 3%' THEN 'КЦ 3'
--    WHEN org.[NAME] like '%центр 4%' THEN 'КЦ 4'
--    ELSE 'др'
--  END as "СП"

,RU.[Marketer] as Marketer
,dr.[CODE] as id  
             

         from [DWH].[dbo].[DIC_REQUEST] dr 

            left join [DWH].[dbo].[DIC_REQUEST_UTM] RU on RU.[ID_REQUEST] = DR.[ID_REQUEST]
-- статусы
            left join  [DWH].[dbo].[DIC_STATUS_REQUEST] dsr on dsr.ID_STATUS_REQUEST = dr.ID_STATUS_REQUEST
-- источник

LEFT JOIN [DWH].[STAGE].[CRM_B_UTS_CRM_LEAD] UL WITH(NOLOCK) ON UL.VALUE_ID = dr.CODE
LEFT JOIN [DWH].[STAGE].[CRM_B_CRM_LEAD] L WITH(NOLOCK) ON L.ID = dr.CODE
INNER JOIN (SELECT [NAME], [STATUS_ID]
           FROM [DWH].[STAGE].[CRM_B_CRM_STATUS] WITH(NOLOCK)
           WHERE [ENTITY_ID] = 'SOURCE' AND (STATUS_ID='WEB' or STATUS_ID= '7')
           ) SRC ON SRC.[STATUS_ID] = L.[SOURCE_ID]

                  LEFT JOIN ASS_REQUEST_SOURCE ARS WITH(NOLOCK) ON dR.ID_REQUEST = ARS.ID_REQUEST
                  LEFT JOIN DIC_SOURCE S WITH(NOLOCK) ON ARS.ID_SOURCE = S.ID_SOURCE
-- исполнители
LEFT JOIN [DWH].[dbo].DIC_EMPLOYEES E With(nolock) ON dr.Id_EMPLOYEES = E.ID_EMPLOYEES
JOIN (SELECT * FROM  [DWH].[dbo].[v_DIC_ORG_EMPL] where [lv1_NAME] like '%Коммерческий департамент (КМВ)%') org on org.[ID_EMPLOYEES] = E.ID_EMPLOYEES

-- продукт бюджета
            LEFT JOIN [DWH].[dbo].ASS_REQUEST_PRODUCT_BUDGET RPB With(nolock) ON dR.ID_REQUEST=RPB.ID_REQUEST and dR.DATE_CREATE= RPB.R_DATE
            LEFT JOIN [DWH].[dbo].DIC_PRODUCT_BUDGET PB With(nolock) ON RPB.ID_PRODUCT_BUDGET=PB.ID_PRODUCT_BUDGET
--расход
            LEFT JOIN [DWH].[dbo].[FCT_REQUEST] fc With(nolock) on fc.[ID_REQUEST] = dR.ID_REQUEST

where 1=1 

and isnull(dr.SIGN_DELETED,0) =0
and year(dr.[DATE_OP]) in ('2021', '2022','2023')
and month(dr.[DATE_OP]) = @month
--and org.full_NAME like '%Коммерческий департамент (КМВ)\%'
and org.[NAME] not like '%МАП%'
--and src.name_source like '%Веб%'
and dr.[FEATURES_5] not in ('estr2022' , 'ege_rf', 'leadform_CZ_747828', 'leadform_CZN_SYN_PP_713311','proftest')
 

'''

empl = f'''

SELECT [EMPLOYEES] as "Ответственный"
      ,CASE

       WHEN KC like '%4%' THEN 'КЦ 4'
       WHEN KC like '%3%' THEN 'КЦ 3'
       WHEN KC like '%центр 3%' THEN 'КЦ 3'
       WHEN KC like '%продаж 18%' THEN 'ОП 18'
             
       ELSE KC
  END
  as "KC"

         ,CASE

       WHEN [KC] like '%4%' THEN SUBSTRING([ОП],CHARINDEX(' ',[ОП])+1,LEN([ОП])) + ' ВР'
       WHEN ([ОП] in ('ОП 9')) and ([GP] in ('.1')) THEN [ОП] + [GP]
       WHEN ([ОП] in ('ОП 2')) and ([GP] in ('.1')) THEN 'ОП 8'
       WHEN ([ОП] in ('ОП 10')) and ([GP] in ('.2')) THEN 'ОП 10.2'
       WHEN ([ОП] in ('ОП 5')) and ([GP] in ('.1')) THEN [ОП] + [GP]
       WHEN ([ОП] in ('3 ЯР')) and ([GP] in ('.1')) THEN '3.1 ЯР'
       WHEN ([KC] in ('Отдел прямых продаж'))  THEN 'ОПП'
           WHEN [KC] like '%продаж 18%'  THEN 'ОП 18'  
       ELSE [ОП]
  END
  as "ОП"


        ,CASE

       WHEN [KC] = 'КЦ 3' THEN 'КЦ 3'
       WHEN [KC] like '%4%' THEN 'Трофимов'
       WHEN [ОП] = 'ОП 5' THEN 'Пельванов'
       WHEN [ОП] = 'ОП 5.1' THEN 'Пельванов'
       WHEN [ОП] = 'ОП 7' THEN 'Пельванов'
       WHEN [ОП] = 'ОП 8' THEN 'Пельванов'
       WHEN [ОП] = 'ОП 9' THEN 'Пельванов'
       WHEN [ОП] = 'ОП 10' THEN 'Пельванов'
       WHEN [ОП] = 'ОП 16' THEN 'Пельванов'
       WHEN [ОП] = 'ОП 6' THEN 'Группа №2'
       WHEN [ОП] = 'ОП 1' THEN 'Группа №2'
       WHEN [ОП] = 'ОП 10.1' THEN 'Группа №2'
       WHEN [ОП] = 'ОП 12' THEN 'Группа №2'
       WHEN [ОП] = 'ОП 13' THEN 'Группа №2'
       WHEN [ОП] = 'ОП 15' THEN 'Группа №2'
       WHEN [ОП] = 'ОП 3' THEN 'ОП 3'
       WHEN [ОП] like '%14%' THEN 'ОП 14'
       ELSE 'др'
  END
  as "Группа"






  FROM [DWH].[dbo].[KHTS_EMPL]
  WHERE 
  [SP] in ('КД')
  AND [STATUS] = 'работает'
 

'''

leads_eff = pd.read_sql_query(leads_eff, conn)
leads = pd.read_sql_query(leads, conn)
empl = pd.read_sql_query(empl, conn)
leads = leads.merge(empl, on='Ответственный', how='left')

leads['День'] = leads['Дата соединения на ОП'].dt.day
leads = leads.query('Статус != "Дубль" & Статус != "Ошибка номера" & Статус != "Повторные заявки" & Статус != "Спам" & Статус != "Черный список"')

g_1 = ['ОП 5', 'ОП 5.1', 'ОП 7', 'ОП 8', 'ОП 9', 'ОП 10', 'ОП 16']
g_2 = ['ОП 6', 'ОП 1', 'ОП 10.2', 'ОП 12', 'ОП 13', 'ОП 15']

baza_planov = pd.read_excel(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Харитуня\Планы\Выручка\База планов 2020.xlsx', 
                            sheet_name='Планы_2020')
baza_planov = baza_planov.query('Менеджер != "Вакансия"')
baza_planov.loc[baza_planov['ОП'].isin(g_1), ['Группа']]='Пельванов'
baza_planov.loc[baza_planov['ОП'].isin(g_2), ['Группа']]='Группа №2'
baza_planov.loc[baza_planov['ОП'] == 'ОП 3', ['Группа']]='ОП 3'
baza_planov.loc[baza_planov['ОП'] == 'ОП 14.2', ['Группа']]= 'ОП 14'
baza_planov.loc[baza_planov['КЦ'] == 'КЦ 3', ['Группа']]= 'КЦ 3'
baza_planov.loc[baza_planov['КЦ'] == 'КЦ 4', ['Группа']]= 'Трофимов'

leads = leads.query('День <= (@d-1)')
leads = leads.query('День <= (@d-1)')
leads_pivot = leads.pivot_table(index='Группа', columns='Год', values='id', aggfunc='count')

leads_pivot = leads_pivot.T.fillna(0)
leads_pivot['Общий итог'] = ( leads_pivot['Пельванов'] + 
                              leads_pivot['Группа №2'] + 
                              leads_pivot['КЦ 3'] + 
                              leads_pivot['Трофимов'] +
                              leads_pivot['ОП 3'] +
                              leads_pivot['ОП 14'] +
                              leads_pivot['др']
                             )
leads_pivot = leads_pivot.T

baza_planov_m = baza_planov.query('Месяц == @month')
baza_planov_m_pivot = baza_planov_m.pivot_table(index='Группа', columns='Год', values='Менеджер', aggfunc='count')

baza_planov_pivot = baza_planov.pivot_table(index='Месяц.1', columns='Год', values='Менеджер', aggfunc='count')



wb = load_workbook(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\Распределение лидов на МП.xlsx')

sheet = wb.get_sheet_by_name('Сводник_1')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(leads_pivot, index=True, header=True):
    sheet.append(r)
    

sheet = wb.get_sheet_by_name('Сводник_2')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(baza_planov_m_pivot, index=True, header=True):
    sheet.append(r)
    
    
sheet = wb.get_sheet_by_name('Сводник_3')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(baza_planov_pivot, index=True, header=True):
    sheet.append(r)
    
    
sheet = wb.get_sheet_by_name('Сводник_4')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(leads_eff, index=True, header=True):
    sheet.append(r)

wb.save(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\Распределение лидов на МП.xlsx')

xlApp = client.Dispatch("Excel.Application")
books = xlApp.Workbooks.Open(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\Распределение лидов на МП.xlsx')
ws = books.Worksheets[0]
ws.Visible = 1
ws.Range("B1:J46").ExportAsFixedFormat(0, r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\PDF\Распределение лидов на МП.pdf')
books.Close(True)
#os.startfile(r'C:\Users\ADavydovskiy\Desktop\Аналитика\Распределение лидов на МП.pdf', 'print')

print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Готово!")
print("")