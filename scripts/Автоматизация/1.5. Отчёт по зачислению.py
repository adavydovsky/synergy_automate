import os
import shutil
import pandas as pd
import numpy as np
import openpyxl as ox
from win32com import client

from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

from datetime import datetime, timedelta

import warnings
warnings.filterwarnings("ignore")

import string

import pyodbc

print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Обновление отчёта по зачислению, пожалуйста, подождите...")

day_m = datetime.today()
y = day_m.year
m_1 = day_m.month-1
y_1 = y-1
y_2 = y-2

dt = 0
if day_m.day <= 5:
    dt = day_m.replace(day=1, month=m_1).strftime('%m.%d.%Y')
else:
    dt = day_m.replace(day=1).strftime('%m.%d.%Y')

wb = ox.load_workbook(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\Отчёт по зачислению.xlsx'
                      ,data_only=True)
sheet = wb.get_sheet_by_name('Отчёт')
z_22 = sheet['E12'].value
z_23 = sheet['E31'].value
nz_22 = sheet['D58'].value
nz_23 = sheet['D77'].value

data = pd.read_excel(r'\\synergy.local\Documents\11.Коммерческий департамент\13.Отдел прямых продаж!\Зачисление\Для Альбины.xlsx', 
                     sheet_name='Для заполнения')

data = data.loc[data['Дата оплаты'] != 'Дубль']
data = data.drop(columns=['№', 'Договор', "Абитуриент", "Встреча",
                          "Менеджер", "Визуалы", "Дата звонка", "Дата доплаты", 
                          "ДАТА ПОДАЧИ", "Сумма доплаты", "STAGE_DO.Статус", "STAGE_OPP.Статус",
                          "STAGE_MOI.Статус", "Комментарий", "Дата зачисления план", 
                          "Дата зачисления факт", "Типовой статус РОП", "Комментарий РОП",
                          "Дата зачисления план РОП", "Дата зачисления факт РОП", "архив",
                          "Ответсвенный ГО", "Дата первого звонка", "Дата следующего звонка",
                          "Успешный старт"])

data['Год'] = data['Дата оплаты'].dt.year
data['Месяц'] = data['Дата оплаты'].dt.month
data[['Год', 'Месяц']]= data[['Год', 'Месяц']].fillna(0)
data[['Год', 'Месяц']] =data[['Год', 'Месяц']].astype(int)
data = data.loc[data['Дата оплаты'] >= "01.01.2021"]
data = data.loc[data['Дата оплаты'] < dt]
data['Категория клиента'] = data['Категория клиента'].replace(' 5 Д. НД', '')
data = data.loc[data['Категория клиента'] != 'Абитуриент оформил заявление на возврат ДС']
data = data.loc[data['Категория клиента'] != 'абитуриент сообщил о желании оформить возврат ДС']
data = data.loc[data['Категория клиента'] != 'Абитуриент сообщил о желании оформить возврат ДС']
data['Категория клиента'] = data['Категория клиента'].replace('неполный', 'Неполный')

data.loc[data['Категория клиента'] == "Зачислен", ['Статус']] = 1
data.loc[data['Категория клиента'] == "Зачислен без ПД", ['Статус']] = 1
data.loc[data['Категория клиента'] == "Полный ПД", ['Статус']] = 1
data.loc[data['Статус'] != 1, ['Статус']] = 0
data['Статус'] = data['Статус'].astype(int)

def svod(year):
    x_1 = 0
    x_2 = 0
    x_3 = 0
    x_4 = 0
    
    x_1 = data.loc[((data['Год'] == year))] 
    x_1 = x_1.pivot_table(index='Месяц',
                              columns='Категория клиента',
                              values='ID сделки',
                              aggfunc='count')
    x_1.loc[:,'Всего оплат'] = x_1.sum(axis=1)
    
    x_2 = data.loc[((data['Год'] == year))] 
    x_2 = x_2.loc[((x_2['Категория клиента'] == 'Зачислен без ПД') | (x_2['Категория клиента'] == np.nan))]
    x_2 = x_2.pivot_table(index='Месяц',
                                  columns='Статус зачисления',
                                  values='ID сделки',
                                  aggfunc='count')
    
    x_3 = data.loc[((data['Год'] == year))] 
    x_3 = x_3.loc[((x_3['Статус зачисления'] == 'Оригиналы документов собраны') | (x_3['Статус зачисления'] == 'Полный пакет документов'))]
    x_3 = x_3.pivot_table(index='Месяц',
                                  values='ID сделки',
                                  aggfunc='count')
    
    x_4 = data.loc[((data['Год'] == year))] 
    x_4 = x_4.loc[((x_4['Категория клиента'] == 'Неполный пакет документов') | (x_4['Категория клиента'] == 'Нет личного дела') | (x_4['Категория клиента'] == 'Недоплата'))]
    x_4 = x_4.pivot_table(index='Месяц',
                                  values='ID сделки',
                                  aggfunc='count')
    x_4.loc[:,'Не зачислен'] = x_4.sum(axis=1)
    
    x = x_1.merge(x_2, on='Месяц', how='outer')
    x = x.merge(x_3, on='Месяц', how='outer')
    x = x.merge(x_4, on='Месяц', how='outer')
    
    x.loc['Общий итог',:] = x.sum(axis=0)
    
    return x


year = svod(y)
year_1 = svod(y_1)
year_2 = svod(y_2)

wb = load_workbook(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\Отчёт по зачислению.xlsx')

sheet = wb.get_sheet_by_name('filename')
sheet.delete_rows(1, sheet.max_row)

sheet['A1'] = z_22
sheet['A2'] = z_23
sheet['A3'] = nz_22
sheet['A4'] = nz_23

sheet = wb.get_sheet_by_name('Год')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(year, index=True, header=True):
    sheet.append(r)

sheet = wb.get_sheet_by_name('Год-1')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(year_1, index=True, header=True):
    sheet.append(r)
    
sheet = wb.get_sheet_by_name('Год-2')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(year_2, index=True, header=True):
    sheet.append(r)
    
wb.save(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\Отчёт по зачислению.xlsx')

xlApp = client.Dispatch("Excel.Application")
books = xlApp.Workbooks.Open(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\Отчёт по зачислению.xlsx')
ws = books.Worksheets[0]
ws.Visible = 1
ws.Range("B1:I99").ExportAsFixedFormat(0, r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\PDF\Отчёт по зачислению.pdf')
books.Close(True)
#os.startfile(r'C:\Users\ADavydovskiy\Desktop\Аналитика\Отчёт по зачислению.pdf', 'print')

src_1 = r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\Отчёт по зачислению.xlsx'
dest_1 = r'\\synergy.local\Documents\11.Коммерческий департамент\13.Отдел прямых продаж!\Зачисление\Отчёт по зачислению.xlsx'

shutil.copyfile(src_1, dest_1)

print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Готово!")
print("")