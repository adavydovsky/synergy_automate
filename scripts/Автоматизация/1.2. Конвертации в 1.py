import os

import pandas as pd
import numpy as np
import openpyxl as ox
from win32com import client
import re

from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

from datetime import datetime, timedelta

import warnings
warnings.filterwarnings("ignore")

import string

import pyodbc

#now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Обновление отчёта по конвертации в 1, пожалуйста, подождите...")

conn = pyodbc.connect(r'Driver={SQL Server};Server=MSK1-BIDB01;Database=DWH;Trusted_Connection=yes;')
cursor = conn.cursor()

leads_yes = f'''

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

 declare @edate date = getdate()-1
 declare @month int = month(getdate()-1)
 declare @sdate date = getdate()-1;


select 

--даты
        R.[CODE] as "id_lead_crm" --id лида crm
       ,R.[DATE_CREATE] as "Дата создания"
       ,cast(R.[DATE_OP] as date) as "Дата соединения на ОП"
       ,R.[FEATURES_5] as "Ленд"
       ,D.[PROBABILITY] "Вероятность"
       ,Rdop.CODE_LABEL_BASE_KD as "Метка базы"
--статус
       ,iif(dsr.[NAME] not like '%/%',dsr.[NAME], SUBSTRING(dsr.[NAME],0,PATINDEX('% / %',dsr.[NAME])))   as "Статус"
--Ответственный
       ,e.LAST_NAME+' '+e.NAME+' '+ISNULL(e.SECOND_NAME,'') as "Ответственный"
       ,CASE

            WHEN org.[NAME] like '%\Колл-центр 1%' THEN 'КЦ 1'
            WHEN org.[NAME] like '%\Колл-центр 2%' THEN 'КЦ 2'
            WHEN org.[NAME] like '%\Колл-центр 3%' THEN 'КЦ 3'
            WHEN org.[NAME] like '%\Колл-центр 4%' THEN 'КЦ 4'
            
            ELSE '-'
            
        END as "КЦ"

       ,CASE

            WHEN org.[NAME] like '%1\Отдел продаж 2%' THEN 'ОП 2'
            WHEN org.[NAME] like '%1\Отдел продаж 3%' THEN 'ОП 3'
            WHEN org.[NAME] like '%1\Отдел продаж 4%' THEN 'ОП 4'
            WHEN org.[NAME] like '%1\Отдел продаж 5\%' THEN 'ОП 5.1'
            WHEN org.[NAME] like '%1\Отдел продаж 5' THEN 'ОП 5'
            WHEN org.[NAME] like '%1\Отдел продаж 6%' THEN 'ОП 6'
            WHEN org.[NAME] like '%1\Отдел продаж 7%' THEN 'ОП 7'
            WHEN org.[NAME] like '%1\Отдел продаж 8%' THEN 'ОП 8'
            WHEN org.[NAME] like '%1\Отдел продаж 9\%' THEN 'ОП 9.1'
            WHEN org.[NAME] like '%1\Отдел продаж 9' THEN 'ОП 9'
            WHEN org.[NAME] like '%1\Отдел продаж 10\%' THEN 'ОП 10.2'
            WHEN org.[NAME] like '%1\Отдел продаж 10' THEN 'ОП 10'
            WHEN org.[NAME] like '%1\Отдел продаж 1' THEN 'ОП 1'
            WHEN org.[NAME] like '%2\Отдел продаж 12%' THEN 'ОП 12'
            WHEN org.[NAME] like '%2\Отдел продаж 13%' THEN 'ОП 13'
            WHEN org.[NAME] like '%2\Отдел продаж 14\%' THEN 'ОП 14.1'
            WHEN org.[NAME] like '%2\Отдел продаж 14' THEN 'ОП 14'
            WHEN org.[NAME] like '%2\Отдел продаж 15%' THEN 'ОП 15'
            WHEN org.[NAME] like '%2\Отдел продаж 16%' THEN 'ОП 16'
            WHEN org.[NAME] like '%2\Отдел продаж 17%' THEN 'ОП 17'
            WHEN org.[NAME] like '%2\Отдел продаж 18%' THEN 'ОП 18'
            WHEN org.[NAME] like '%Ярославль\Отдел продаж 1%' THEN 'ЯР 1'
            WHEN org.[NAME] like '%Ярославль\Отдел продаж 2%' THEN 'ЯР 2'
            WHEN org.[NAME] like '%Ярославль\Отдел продаж 3%' THEN 'ЯР 3'
            WHEN org.[NAME] like '%Ярославль\Отдел продаж 4%' THEN 'ЯР 4'
            WHEN org.[NAME] like '%Ярославль\Отдел продаж 5%' THEN 'ЯР 5'
            WHEN org.[NAME] like '%Ярославль\Отдел продаж 6%' THEN 'ЯР 6'
            WHEN org.[NAME] like '%Воронеж\Отдел продаж 1%' THEN 'ВР 1'
            WHEN org.[NAME] like '%Воронеж\Отдел продаж 2%' THEN 'ВР 2'
            WHEN org.[NAME] like '%Воронеж\Отдел продаж 3%' THEN 'ВР 3'
            WHEN org.[NAME] like '%Воронеж\Отдел продаж 4%' THEN 'ВР 4'
            WHEN org.[NAME] like '%Воронеж\Отдел продаж 5%' THEN 'ВР 5'
            WHEN org.[NAME] like '%Воронеж\Отдел продаж 6%' THEN 'ВР 6'
            
            ELSE '-'
        END as "ОП"



from [DWH].[dbo].[DIC_REQUEST] R
left join [DWH].[dbo].DIC_DEAL D on D.[ID_REQUEST] = R.[ID_REQUEST]
left join [dbo].[DIC_REQUEST_UTM] RU With(nolock) on RU.[ID_REQUEST] = R.[ID_REQUEST]

-- статусы

left join  [DWH].[dbo].[DIC_STATUS_REQUEST] dsr With(nolock) on dsr.ID_STATUS_REQUEST = R.ID_STATUS_REQUEST

-- источник

LEFT JOIN [DWH].[STAGE].[CRM_B_UTS_CRM_LEAD] UL WITH(NOLOCK) ON UL.VALUE_ID = R.CODE
LEFT JOIN [DWH].[STAGE].[CRM_B_CRM_LEAD] L WITH(NOLOCK) ON L.ID = R.CODE
INNER JOIN (SELECT [NAME], [STATUS_ID]
           FROM [DWH].[STAGE].[CRM_B_CRM_STATUS] WITH(NOLOCK)
           WHERE [ENTITY_ID] = 'SOURCE' AND (STATUS_ID='WEB' or STATUS_ID= '7')
           ) SRC ON SRC.[STATUS_ID] = L.[SOURCE_ID]     

LEFT JOIN ASS_REQUEST_SOURCE ARS WITH(NOLOCK) ON R.ID_REQUEST = ARS.ID_REQUEST
LEFT JOIN DIC_SOURCE S WITH(NOLOCK) ON ARS.ID_SOURCE = S.ID_SOURCE

-- исполнители
LEFT JOIN [DWH].[dbo].DIC_EMPLOYEES E With(nolock) ON R.Id_EMPLOYEES = E.ID_EMPLOYEES
JOIN (SELECT * FROM  [DWH].[dbo].[v_DIC_ORG_EMPL] where [lv1_NAME] like '%Коммерческий департамент (КМВ)%') org on org.[ID_EMPLOYEES] = E.ID_EMPLOYEES

-- доп. коды лидов

LEFT JOIN [DWH].[dbo].[DIC_REQUEST_STAT] Rdop With(nolock) on R.ID_REQUEST = Rdop.ID_REQUEST

-- сделки

LEFT JOIN [DWH].[dbo].ASS_REQUEST_DEAL ASS_RD With(nolock) ON ASS_RD.ID_REQUEST = R.ID_REQUEST
LEFT JOIN [DWH].[dbo].[DIC_DEAL] DD With(nolock) ON DD.ID_DEAL = ASS_RD.ID_DEAL




where 1=1 

and cast(R.[DATE_OP] as date) between @sdate and @edate
and org.[NAME] not like '%Омский%'
and org.[NAME] not like '%МАП%'
and org.[NAME] not like '%МОИ%'
--and org.full_NAME like '%Коммерческий департамент (КМВ)\%'
--and src.name_source like '%Веб%'

'''

leads_konv_1 = f'''

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

 declare @edate date = getdate()-1
 declare @month int = month(getdate()-1)
 declare @sdate date = DATEFROMPARTS(2023, @month, 01);


select 

--даты
       R.[CODE] as "id_lead_crm" --id лида crm
      ,R.[DATE_CREATE] as "Дата создания"
      ,cast(R.[DATE_OP] as date) as "Дата соединения на ОП"
      ,R.[FEATURES_5] as "Ленд"
      ,D.[PROBABILITY] "Вероятность"
      ,Rdop.CODE_LABEL_BASE_KD as "Метка базы"
--статус
      ,iif(dsr.[NAME] not like '%/%',dsr.[NAME], SUBSTRING(dsr.[NAME],0,PATINDEX('% / %',dsr.[NAME])))   as "Статус"
--Ответственный
      ,e.LAST_NAME+' '+e.NAME+' '+ISNULL(e.SECOND_NAME,'') as "Ответственный"
      ,CASE

           WHEN org.[NAME] like '%\Колл-центр 1%' THEN 'КЦ 1'
           WHEN org.[NAME] like '%\Колл-центр 2%' THEN 'КЦ 2'
           WHEN org.[NAME] like '%\Колл-центр 3%' THEN 'КЦ 3'
           WHEN org.[NAME] like '%\Колл-центр 4%' THEN 'КЦ 4'
           
           ELSE '-'
       END as "КЦ"
      ,CASE

           WHEN org.[NAME] like '%1\Отдел продаж 2%' THEN 'ОП 2'
           WHEN org.[NAME] like '%1\Отдел продаж 3%' THEN 'ОП 3'
           WHEN org.[NAME] like '%1\Отдел продаж 4%' THEN 'ОП 4'
           WHEN org.[NAME] like '%1\Отдел продаж 5\%' THEN 'ОП 5.1'
           WHEN org.[NAME] like '%1\Отдел продаж 5' THEN 'ОП 5'
           WHEN org.[NAME] like '%1\Отдел продаж 6%' THEN 'ОП 6'
           WHEN org.[NAME] like '%1\Отдел продаж 7%' THEN 'ОП 7'
           WHEN org.[NAME] like '%1\Отдел продаж 8%' THEN 'ОП 8'
           WHEN org.[NAME] like '%1\Отдел продаж 9\%' THEN 'ОП 9.1'
           WHEN org.[NAME] like '%1\Отдел продаж 9' THEN 'ОП 9'
           WHEN org.[NAME] like '%1\Отдел продаж 10\%' THEN 'ОП 10.2'
           WHEN org.[NAME] like '%1\Отдел продаж 10' THEN 'ОП 10'
           WHEN org.[NAME] like '%1\Отдел продаж 1' THEN 'ОП 1'
           WHEN org.[NAME] like '%2\Отдел продаж 12%' THEN 'ОП 12'
           WHEN org.[NAME] like '%2\Отдел продаж 13%' THEN 'ОП 13'
           WHEN org.[NAME] like '%2\Отдел продаж 14\%' THEN 'ОП 14.1'
           WHEN org.[NAME] like '%2\Отдел продаж 14' THEN 'ОП 14'
           WHEN org.[NAME] like '%2\Отдел продаж 15%' THEN 'ОП 15'
           WHEN org.[NAME] like '%2\Отдел продаж 16%' THEN 'ОП 16'
           WHEN org.[NAME] like '%2\Отдел продаж 17%' THEN 'ОП 17'
           WHEN org.[NAME] like '%2\Отдел продаж 18%' THEN 'ОП 18'
           WHEN org.[NAME] like '%Ярославль\Отдел продаж 1%' THEN 'ЯР 1'
           WHEN org.[NAME] like '%Ярославль\Отдел продаж 2%' THEN 'ЯР 2'
           WHEN org.[NAME] like '%Ярославль\Отдел продаж 3%' THEN 'ЯР 3'
           WHEN org.[NAME] like '%Ярославль\Отдел продаж 4%' THEN 'ЯР 4'
           WHEN org.[NAME] like '%Ярославль\Отдел продаж 5%' THEN 'ЯР 5'
           WHEN org.[NAME] like '%Ярославль\Отдел продаж 6%' THEN 'ЯР 6'
           WHEN org.[NAME] like '%Воронеж\Отдел продаж 1%' THEN 'ВР 1'
           WHEN org.[NAME] like '%Воронеж\Отдел продаж 2%' THEN 'ВР 2'
           WHEN org.[NAME] like '%Воронеж\Отдел продаж 3%' THEN 'ВР 3'
           WHEN org.[NAME] like '%Воронеж\Отдел продаж 4%' THEN 'ВР 4'
           WHEN org.[NAME] like '%Воронеж\Отдел продаж 5%' THEN 'ВР 5'
           WHEN org.[NAME] like '%Воронеж\Отдел продаж 6%' THEN 'ВР 6'
           
           ELSE '-'
        END as "ОП"


from [DWH].[dbo].[DIC_REQUEST] R
left join [DWH].[dbo].DIC_DEAL D on D.[ID_REQUEST] = R.[ID_REQUEST]
left join [dbo].[DIC_REQUEST_UTM] RU With(nolock) on RU.[ID_REQUEST] = R.[ID_REQUEST]
 -- статусы
left join  [DWH].[dbo].[DIC_STATUS_REQUEST] dsr With(nolock) on dsr.ID_STATUS_REQUEST = R.ID_STATUS_REQUEST
-- источник

LEFT JOIN [DWH].[STAGE].[CRM_B_UTS_CRM_LEAD] UL WITH(NOLOCK) ON UL.VALUE_ID = R.CODE
LEFT JOIN [DWH].[STAGE].[CRM_B_CRM_LEAD] L WITH(NOLOCK) ON L.ID = R.CODE
INNER JOIN (SELECT [NAME], [STATUS_ID]
           FROM [DWH].[STAGE].[CRM_B_CRM_STATUS] WITH(NOLOCK)
           WHERE [ENTITY_ID] = 'SOURCE' AND (STATUS_ID='WEB' or STATUS_ID= '7')
           ) SRC ON SRC.[STATUS_ID] = L.[SOURCE_ID]

LEFT JOIN ASS_REQUEST_SOURCE ARS WITH(NOLOCK) ON R.ID_REQUEST = ARS.ID_REQUEST
LEFT JOIN DIC_SOURCE S WITH(NOLOCK) ON ARS.ID_SOURCE = S.ID_SOURCE
-- исполнители
LEFT JOIN [DWH].[dbo].DIC_EMPLOYEES E With(nolock) ON R.Id_EMPLOYEES = E.ID_EMPLOYEES
JOIN (SELECT * FROM  [DWH].[dbo].[v_DIC_ORG_EMPL] where [lv1_NAME] like '%Коммерческий департамент (КМВ)%') org on org.[ID_EMPLOYEES] = E.ID_EMPLOYEES

-- доп. коды лидов
LEFT JOIN [DWH].[dbo].[DIC_REQUEST_STAT] Rdop With(nolock) on R.ID_REQUEST = Rdop.ID_REQUEST

-- сделки
LEFT JOIN [DWH].[dbo].ASS_REQUEST_DEAL ASS_RD With(nolock) ON ASS_RD.ID_REQUEST = R.ID_REQUEST
LEFT JOIN [DWH].[dbo].[DIC_DEAL] DD With(nolock) ON DD.ID_DEAL = ASS_RD.ID_DEAL




where 1=1 

 and cast(R.[DATE_OP] as date) between @sdate and @edate
 and org.[NAME] not like '%Омский%'
 and org.[NAME] not like '%МАП%'
 and org.[NAME] not like '%МОИ%'
 --and org.full_NAME like '%Коммерческий департамент (КМВ)\%'
 --and src.name_source like '%Веб%'
 
'''

leads_yes = pd.read_sql_query(leads_yes, conn)
leads_konv_1 = pd.read_sql_query(leads_konv_1, conn)

leads_konv_1['id_lead_crm'] = leads_konv_1['id_lead_crm'].astype(int)
leads_konv_1['Дата создания'] = pd.to_datetime(leads_konv_1['Дата создания'])
leads_konv_1['Дата соединения на ОП'] = pd.to_datetime(leads_konv_1['Дата соединения на ОП'])

leads_konv_1 = leads_konv_1.loc[leads_konv_1['Метка базы'] != "ЧБ"]
leads_konv_1 = leads_konv_1.loc[leads_konv_1['Метка базы'] != "ЧБ "]

leads_konv_1['День недели'] = leads_konv_1['Дата создания'].dt.dayofweek
leads_konv_1['День недели'] = leads_konv_1['День недели'].astype(str)

leads_konv_1['День недели'] = leads_konv_1['День недели'].replace('0', 'Понедельник')
leads_konv_1['День недели'] = leads_konv_1['День недели'].replace('1', 'Вторник')
leads_konv_1['День недели'] = leads_konv_1['День недели'].replace('2', 'Среда')
leads_konv_1['День недели'] = leads_konv_1['День недели'].replace('3', 'Четверг')
leads_konv_1['День недели'] = leads_konv_1['День недели'].replace('4', 'Пятница')
leads_konv_1['День недели'] = leads_konv_1['День недели'].replace('5', 'Суббота')
leads_konv_1['День недели'] = leads_konv_1['День недели'].replace('6', 'Воскрсенье')

leads_konv_1['Ответственный'] = leads_konv_1['Ответственный'].fillna(0)
leads_konv_1 = leads_konv_1.query('Ответственный != 0 & Ответственный != "Гришунова Ольга Львовна" & Ответственный != "Кузнецова Дарья Андреевна"')
leads_konv_1 = leads_konv_1.drop_duplicates()

leads_yes['id_lead_crm'] = leads_yes['id_lead_crm'].astype(int)
leads_yes['Дата создания'] = pd.to_datetime(leads_yes['Дата создания'])
leads_yes['Дата соединения на ОП'] = pd.to_datetime(leads_yes['Дата соединения на ОП'])

leads_yes = leads_yes.loc[leads_yes['Метка базы'] != "ЧБ"]
leads_yes = leads_yes.loc[leads_yes['Метка базы'] != "ЧБ "]

leads_yes['День недели'] = leads_yes['Дата создания'].dt.dayofweek
leads_yes['День недели'] = leads_yes['День недели'].astype(str)

leads_yes['День недели'] = leads_yes['День недели'].replace('0', 'Понедельник')
leads_yes['День недели'] = leads_yes['День недели'].replace('1', 'Вторник')
leads_yes['День недели'] = leads_yes['День недели'].replace('2', 'Среда')
leads_yes['День недели'] = leads_yes['День недели'].replace('3', 'Четверг')
leads_yes['День недели'] = leads_yes['День недели'].replace('4', 'Пятница')
leads_yes['День недели'] = leads_yes['День недели'].replace('5', 'Суббота')
leads_yes['День недели'] = leads_yes['День недели'].replace('6', 'Воскрсенье')

leads_yes['Ответственный'] = leads_yes['Ответственный'].fillna(0)
leads_yes = leads_yes.query('Ответственный != 0 & Ответственный != "Гришунова Ольга Львовна" & Ответственный != "Кузнецова Дарья Андреевна"')
leads_yes = leads_yes.drop_duplicates()

leads_konv_1_pivot = leads_konv_1.pivot_table(index='ОП', values='id_lead_crm', aggfunc='count')

leads_konv_2 = leads_konv_1.query('Вероятность == 1')


leads_konv_1_pivot_2 = leads_konv_2.pivot_table(index='ОП', values='id_lead_crm', aggfunc='count')
leads_konv_1_pivot = leads_konv_1_pivot.merge(leads_konv_1_pivot_2, how='left', on='ОП')

leads_konv_1_pivot = leads_konv_1_pivot.fillna(0)
leads_konv_1_pivot = leads_konv_1_pivot.astype(int)
leads_konv_1_pivot.columns = ['Все соединённые', 'Сконвертированные в 1']

leads_yes_pivot = leads_yes.pivot_table(index='ОП', values='id_lead_crm', aggfunc='count')

leads_yes_2 = leads_yes.query('Вероятность == 1')
leads_yes_pivot_2 = leads_yes_2.pivot_table(index='ОП', values='id_lead_crm', aggfunc='count')
leads_yes_pivot = leads_yes_pivot.merge(leads_yes_pivot_2, how='left', on='ОП')

leads_yes_pivot = leads_yes_pivot.fillna(0)
leads_yes_pivot = leads_yes_pivot.astype(int)

try:
    leads_yes_pivot.columns = ['Все соединённые', 'Сконвертированные в 1']
except:
    print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Увы, не удалось обнаружить данных за вчера..")

wb = load_workbook(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\По конвертации в 1.xlsx')

sheet = wb.get_sheet_by_name('Сводник_1')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(leads_konv_1_pivot, index=True, header=True):
    sheet.append(r)
    

sheet = wb.get_sheet_by_name('Сводник_2')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(leads_yes_pivot, index=True, header=True):
    sheet.append(r)

wb.save(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\По конвертации в 1.xlsx')



xlApp = client.Dispatch("Excel.Application")
books = xlApp.Workbooks.Open(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\По конвертации в 1.xlsx')
ws = books.Worksheets[0]
ws.Visible = 1
ws.Range("B1:L42").ExportAsFixedFormat(0, r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\PDF\По конвертации в 1.pdf')
books.Close(True)
#os.startfile(r'C:\Users\ADavydovskiy\Desktop\Аналитика\По конвертации в 1.pdf', 'print')

print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Готово!")
print("")