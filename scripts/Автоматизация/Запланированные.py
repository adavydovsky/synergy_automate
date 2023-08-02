import pandas as pd
import numpy as np
import time
from datetime import datetime, timedelta
from tqdm import tqdm
import openpyxl as ox
import re
import os

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains

import warnings
warnings.filterwarnings("ignore")

log = os.getlogin()

print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Выгрузка запланированных, пожалуйста, подождите...")


driver = webdriver.Chrome(executable_path = f'C:\\Users\\{log}\\scripts\\YandexDriver\\yandexdriver.exe')
x = driver.get('https://corp.synergy.ru/crm/reports.synergy/plan_oplat/')
elem = Select(driver.find_element(By.ID, 'period'))
elem.select_by_value('cday')

rec = driver.find_element(By.ID, 'rec').click()

prob = driver.find_element(By.XPATH, 
                           '/html/body/table/tbody/tr[2]/td/table/tbody/tr[1]/td[2]/table/tbody/tr[2]/td/div/div[2]/div/div[2]/form/div[6]/input[2]'
                          ).click()


stage = Select(driver.find_element(By.ID, 'stage'))
stage.select_by_value('В работе')

elem = driver.find_element(By.XPATH, '//*[@id="dept_chosen"]/a/span').click()

element = driver.find_element(By.XPATH, '//*[@id="dept_chosen"]/div/div/input')
ActionChains(driver) \
    .send_keys_to_element(element, "(КМВ)", Keys.ENTER)\
    .perform()


    
find = driver.find_element(By.XPATH, 
                           '//*[@id="workarea-content"]/div/div[2]/form/div[14]/input'
                          ).click()
time.sleep(20)

tab = driver.find_element(By.XPATH, '//*[@id="mytable"]')    
tab_html=tab.get_attribute('outerHTML')
tab_dfs=pd.read_html(tab_html)
df = tab_dfs[0]
y = df['Сумма за день'][1]
df['Сумма за день'] = df['Сумма за день'].replace(y, np.nan)
df['Сумма за день'] = df['Сумма за день'][-2:].fillna(y)

driver.close()
    
#-----------------------------------------------------------------------

plan = df.drop(['№', 'Дата создания', 'Дата оплаты', 'Визуал'], axis=1)
plan = plan.drop(plan.index[-1:])
plan = plan.rename(columns={'Дата завершения': 'Дата закрытия'})

zaplan = pd.read_excel(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Запланированные.xlsx', sheet_name='1')
index = zaplan['№'].head(1)
index = index.to_string(index=False)
zaplan = zaplan.drop('№', axis=1)

index = re.sub("[^0-9]", "", index)
index = int(index)
index = index-1

frames = [zaplan, plan]
zaplan = pd.concat(frames).reset_index(drop=True)
zaplan['№'] = 0

for i in range(len(zaplan)):
    index= index+1
    zaplan['№'][i] = "G" + str(index)

zaplan = zaplan[['№','Сделка', 'Статус сделки', 'ФИО клиента', 'Город', 'Дата закрытия',
       'Источник (UTM SOURCE)', 'Источник', 'Продукт', 'Вероятность оплаты',
       'Сумма', 'Сумма за день', 'Менеджер', 'Дата создания лида',
       'Дата соединения на ОП', 'Группа продуктов', 'ID лида',
       'Прослушано РОП', 'Количество', 'Чистая выручка']]


wb = load_workbook(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Запланированные.xlsx')
sheet = wb.get_sheet_by_name('1')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(zaplan, index=False, header=True):
    sheet.append(r)
    
wb.save(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Запланированные.xlsx')

#-----------------------------------------------------------------------

wb = load_workbook(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Планирование на день РОП.xlsx')

sheet = wb.get_sheet_by_name('1')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(df, index=False, header=True):
    sheet.append(r)


wb.save(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Планирование на день РОП.xlsx')

print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Готово!")
print("")