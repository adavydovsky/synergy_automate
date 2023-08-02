import os

import pandas as pd
import numpy as np
import openpyxl as ox
from win32com import client
import re

from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

from datetime import timedelta
from datetime import datetime

import warnings
warnings.filterwarnings("ignore")

import string

import pyodbc

#now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Обновление отчёта по центру интелекта (ДОП), пожалуйста, подождите...")
#-------------------------------------------------------

def date(x):
    d = datetime(1900,1,1)
    y = timedelta(days=x-2)
    x = d+y
    x = x.strftime("%Y%m%d")
    return x


def product(x, y):
    x['ОПЛАТА ПО ФАКТУ'] = pd.to_datetime(x['ОПЛАТА ПО ФАКТУ'])
    x['Год_оплаты_факт'] = x['ОПЛАТА ПО ФАКТУ'].dt.year
    x['Месяц'] = x['ОПЛАТА ПО ФАКТУ'].dt.month
    x = x.query('Год_оплаты_факт == 2023')
    x['ID сделки'] = x['ID сделки'].drop_duplicates()
    x = x.loc[x['Продукт_2'] == y]
    x = x[['ID сделки', 'ОТДЕЛ', 'Продукт', 'ID лида', 'Клиент', 'Месяц']]
    x['Продукт'] = x['Продукт'].replace('курсы', 1)
    x = x.rename(columns={'Продукт': y})
#    x['ID сделки'] = x['ID сделки'].fillna(0)
#    x['ID сделки'] = x['ID сделки'].astype(int)
    return x

#-------------------------------------------------------

data = pd.read_excel(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\База ЧВ_бд18 0,25.xlsx', sheet_name='База ЧВ')

for row in data['ОПЛАТА ПО ФАКТУ']:
    data.loc[data['ОПЛАТА ПО ФАКТУ'] == row, ['ОПЛАТА ПО ФАКТУ']] = date(row)


sch = data
rp = data
two_in_one = data

#-------------------------------------------------------

conn = pyodbc.connect(r'Driver={SQL Server};Server=MSK1-BIDB01;Database=DWH;Trusted_Connection=yes;')
cursor = conn.cursor()

ls = f'''

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

declare @date date = getdate()-1
declare @month int = month(@date)
declare @startdate date = DATEFROMPARTS(2022,01,01);

SELECT DISTINCT
     --даты
     D.CODE as "id deal crm"
     
      ,D.ID_CFO_VISUAL  as "Визуал"
     
       


FROM [DWH].[dbo].DIC_DEAL D
    left join [DWH].[dbo].[DIC_DEAL_VISUAL] V ON V.[ID_DEAL]=D.ID_DEAL
    LEFT JOIN [DWH].[dbo].[DIC_EMPLOYEES] ed ON ed.ID_EMPLOYEES = V.[ID_EMPLOYEES_VISUAL]

    LEFT JOIN [DWH].[dbo].[DIC_EMPLOYEES] E ON D.ID_EMPLOYEES = E.ID_EMPLOYEES
    LEFT JOIN [DWH].[dbo].[ASS_EMPLOYEE_AND_ORGSTRUCTURE] ASS_OS ON E.ID_EMPLOYEES = ASS_OS.ID_EMPLOYEES
    LEFT JOIN [DWH].[dbo].[v_DIC_ORGSTRUCTURE] org on org.ID_ORGSTRUCTURE = ASS_OS.ID_ORGSTRUCTURE
    

    
    -- статус сделки
    left join [DWH].[dbo].[DIC_STATUS_DEAL] sd on sd.ID_STATUS_DEAL = d.ID_STATUS_DEAL


  
WHERE
  D.ID_DEAL <> - 1

    and D.[DATE_CREATE]  between @startdate and @date -- фильтр по дате завершения 
and org.full_NAME not like '%МАП%'
and org.full_NAME like '%Коммерческий департамент (КМВ)\%'
and D.ID_CFO_VISUAL > 0

  
'''


Employees = f'''

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

SELECT [EMPLOYEES]
      ,[LOG]
      ,[SP]

      ,CASE

       WHEN KC like '%4%' THEN 'КЦ 4'
       WHEN KC like '%3%' THEN 'КЦ 3'
       WHEN KC like '%центр 3%' THEN 'КЦ 3'
       WHEN KC like '%продаж 18%' THEN 'ОП 18'
             
       ELSE KC
  END
  as "KC"

         ,CASE

       WHEN [KC] like '%4%' THEN SUBSTRING([ОП],CHARINDEX(' ',[ОП])+1,LEN([ОП])) + ' ВР'
       WHEN ([ОП] in ('ОП 9')) and ([GP] in ('.1')) THEN [ОП] + [GP]
       WHEN ([ОП] in ('ОП 2')) and ([GP] in ('.1')) THEN 'ОП 8'
       WHEN ([ОП] in ('ОП 10')) and ([GP] in ('.2')) THEN 'ОП 10.2'
       WHEN ([ОП] in ('ОП 5')) and ([GP] in ('.1')) THEN [ОП] + [GP]
       WHEN ([ОП] in ('3 ЯР')) and ([GP] in ('.1')) THEN '3.1 ЯР'
       WHEN ([KC] in ('Отдел прямых продаж'))  THEN 'ОПП'
           WHEN [KC] like '%продаж 18%'  THEN 'ОП 18'  
       ELSE [ОП]
  END
  as "ОП"
      
               ,[PHONE]
      ,[STATUS]
      ,[ID_EMPL]
      ,[ID_ORG]

  FROM [DWH].[dbo].[KHTS_EMPL]
  where 
  --KC like '%4%' 
  [SP] in ('КД')


'''

opp = f'''
SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

declare @startdate date = '2023.01.01'
declare @date date = getdate()-1;


WITH T as
(
SELECT distinct
 
 D.[CODE] as "id_deal"
 ,D_I.[ID_INVOICE] as "inv"
,ASS_D_I.[ID_INVOICE] as "id_invoice"
 ,D_I.[DATE_PAY] as "date_pay"
 ,year(D_I.[DATE_PAY]) as "date_year"
 ,month(D_I.[DATE_PAY]) as "date_month"
 ,EM.SP as "SP"
 ,EM.KC as "KC"
 ,EM.[EMPLOYEES] as "emp"
 ,isnull(EM.ОП ,'нет') as "OP"
        
,"Form"=
 CASE 
    WHEN pin.name like '%ДОТ%'  THEN 'Дист'
     WHEN pin.name like '%Классич%'  THEN 'ПКО'
     WHEN pin.name like '%On-line индивидуальная%'  THEN 'Дист'
     WHEN pin.name like '%вечер%'  THEN 'ПВО'
     WHEN pin.name like '%выходн%'  THEN 'ПВД'
     WHEN pin.name like '%заочн%'  THEN 'ПЗО'
     WHEN pin.name like '%перепод%'  THEN 'очка'
    ELSE pin.name
  END
 
   ,CASE

    
    WHEN pin.name like '%ДПО%' THEN 'ДПО'
    WHEN pin.name like '%ДО /%' THEN 'ДПО'WHEN (pin.name like '%ВО%') or (pin.name like '%ВПО%') THEN 'ВО'
    
    WHEN pin.name like 'СПО /%' THEN 'СПО'
    
    ELSE pin.name
  END
  as "Level_1"

  ,CASE

    WHEN pin.name like '%БАК%' THEN 'БАК'
    WHEN pin.name like '%АСП%' THEN 'АСП'
    WHEN pin.name like '%ДПО%' THEN 'ДПО'
    WHEN pin.name like '%СПО /%' THEN 'СПО'
    WHEN pin.name like '%МАГ%' THEN 'МАГ'
    WHEN pin.name like '%ДО /%' THEN 'ДПО'
    
    ELSE pin.name
  END
  as "Level_2"


  ,pin.name as "name"
  ,cast(PIN.PRICE as INT) as PRICE
  
   ,
 cast(
 CASE 
    WHEN D_I.PAYED_NUM is null THEN 1
    WHEN D_I.PAYED_NUM =0 THEN 1
    
    ELSE D_I.PAYED_NUM
  END
  as decimal(33,0)) as "Semestr"

 ,CAST(iif(D_I.[AMOUNT] = 0,1,D_I.[AMOUNT]) as decimal(33,0)) as AMOUNT

  ,
 cast(
 CASE 
    WHEN D_I.PAYED_NUM = 1 THEN D_I.[AMOUNT] 
    WHEN D_I.PAYED_NUM = 2 THEN D_I.[AMOUNT]*0.5 
    WHEN D_I.PAYED_NUM > 2 THEN D_I.[AMOUNT]*0.25 
    
    ELSE D_I.[AMOUNT]
  END
  as decimal(33,0)) as "TAMOUNT"
  
 ,
 CASE 
    WHEN R.[FEATURES_1] is null THEN iif(src.name_source like '%/%',SUBSTRING(src.name_source,0,PATINDEX('%/%',src.name_source)),src.name_source)
    WHEN R.[FEATURES_1] is null THEN 'органика' 
    WHEN S.Tag1 is null THEN R.[FEATURES_1]
    
    ELSE S.Tag1
  END
   as "source_tag"
  ,
 CASE 
    WHEN R.[FEATURES_5] is null THEN iif(src.name_source like '%/%',SUBSTRING(src.name_source,0,PATINDEX('%/%',src.name_source)),src.name_source)
    
    ELSE R.FEATURES_5
  END
   as "land"

   ,d.[CONTRACT_NUMBER] as "cont"
 --  ,UF_CRM_5602E8ED144DB
   ,UF_CRM_1435304800
      
  FROM [DWH].[dbo].[ASS_DEAL_INVOICE] ASS_D_I
  --инвойсы
  LEFT JOIN [DWH].[dbo].DIC_INVOICE D_I WITH(NOLOCK) ON ASS_D_I.ID_INVOICE = D_I.ID_INVOICE
  LEFT JOIN [DWH].[dbo].DIC_STATUS_INVOICE S_I WITH(NOLOCK) ON S_I.ID_STATUS_INVOICE = D_I.ID_STATUS_INVOICE
  left join [DWH].[dbo].[DIC_INVOICE_AKADA] IA WITH(NOLOCK) ON IA.ID_INVOICE = ASS_D_I.ID_INVOICE

  LEFT JOIN [DWH].[dbo].[DIC_INVOICE_BASKET] PIN WITH(NOLOCK) ON PIN.ID_INVOICE = D_I.ID_INVOICE
  --(select*
  --from [DWH].[dbo].[DIC_INVOICE_BASKET]
  --where [SIGN_DELETED] is null)
  --PIN ON PIN.ID_INVOICE = D_I.ID_INVOICE

  --сделки
  LEFT JOIN [DWH].[dbo].DIC_DEAL D WITH(NOLOCK) ON ASS_D_I.ID_DEAL = D.ID_DEAL
  LEFT JOIN [stage].[CRM_b_uts_crm_deal] UTS_D WITH(NOLOCK) ON UTS_D.VALUE_ID = D.[CODE]

  --оргструктура
  LEFT JOIN [dbo].[KHTS_EMPL]EM WITH(NOLOCK) on EM.ID_EMPL = D_I.ID_EMPLOYEES
    
  --каталог продуктов сделки
  LEFT JOIN 
  (select*
  from [DWH].[dbo].ASS_DEAL_CATALOG_PRODUCT
  where [SIGN_DELETED] is null)
  ASS_P ON D.ID_DEAL = ASS_P.ID_DEAL
  
  
 -- LEFT JOIN [DWH].[dbo].ASS_DEAL_CATALOG_PRODUCT ASS_P ON D.ID_DEAL = ASS_P.ID_DEAL
    LEFT JOIN [DWH].[dbo].[DIC_CATALOG_PRODUCT] cp WITH(NOLOCK) on cp.ID_CATALOG_PRODUCT = ASS_P.ID_CATALOG_PRODUCT
  --подсоединение заявок
  LEFT JOIN [DWH].[dbo].ASS_REQUEST_DEAL ASS_RD WITH(NOLOCK) ON ASS_RD.ID_DEAL = D.ID_DEAL
  LEFT JOIN [DWH].[dbo].[DIC_REQUEST] R WITH(NOLOCK) ON ASS_RD.ID_REQUEST = R.ID_REQUEST
  left join [DWH].[dbo].[DIC_REQUEST_UTM] RU WITH(NOLOCK) on RU.[ID_REQUEST] = R.[ID_REQUEST]

  LEFT JOIN ASS_REQUEST_SOURCE ARS WITH(NOLOCK) ON R.ID_REQUEST = ARS.ID_REQUEST
  LEFT JOIN DIC_SOURCE S WITH(NOLOCK) ON ARS.ID_SOURCE = S.ID_SOURCE

  -- источник
            left join [DWH].[stage].[CRM_b_uts_crm_lead] ul WITH(NOLOCK) on ul.value_id = R.CODE
            left join [DWH].[stage].[CRM_b_crm_lead] l WITH(NOLOCK) on l.ID = R.code
            left join (SELECT [NAME] as name_source, [STATUS_ID]
                           FROM [DWH].[stage].[CRM_b_crm_status]
                          where [ENTITY_ID] = 'SOURCE') src on src.[STATUS_ID] = l.[SOURCE_ID]
            left join [DWH].[dbo].[DIC_REQUEST_FORM] RF on RF.ID_REQUEST = R.ID_REQUEST

     -- продукт бюджета
            LEFT JOIN [DWH].[dbo].ASS_REQUEST_PRODUCT_BUDGET RPB With(nolock) ON R.ID_REQUEST=RPB.ID_REQUEST and R.DATE_CREATE= RPB.R_DATE
            LEFT JOIN [DWH].[dbo].DIC_PRODUCT_BUDGET PB With(nolock) ON RPB.ID_PRODUCT_BUDGET=PB.ID_PRODUCT_BUDGET

 WHERE
 cast(D_I.[DATE_PAY] as date) between @startdate and @date
and isnull(pin.[SIGN_DELETED],0) <> 1
 and S_I.[CODE] in ('P','L')
 --and R.[FEATURES_4] like '%самарский%'
 --and R.[FEATURES_1] like '%edunetwork%'
and em.[sp] like '%КД%'
--and pb.name like '%Дубай%'
and pin.name like '%Сверхпамять%' OR pin.name like '%чтения%'
AND kc = 'Отдел прямых продаж'
--AND EM.[EMPLOYEES] like '%Хамзина%'

--and D.[CODE] in ('4556618',

--and D_I.[ID_INVOICE] = 274853333

)

select distinct

t.date_pay
,t.inv
,t.sp
,t.kc
,t.op
,t.emp
,t.level_1
,t.level_2
,t.Form
--,sum(t.AMOUNT) as amount
--,max(t.Price) as Price
--,max(t.Semestr) as Sem
,sum(t.[TAMOUNT]) as TAmount
, t.[id_deal]
,iif(t.source_tag in ('Веб-сайт'),'органика',t.source_tag) as source_tag
,t.land
,t.name
,t.date_year
,t.date_month
,t.cont

from t
group by
t.date_pay 
,t.inv
,t.[id_deal]
,t.date_year
,t.date_month
,t.sp
,t.[kc]
,t.emp
,t.op
,t.name
,t.level_1
,t.level_2
,t.Form
,t.source_tag
,t.land
,t.cont


 order by t.date_month
 
 '''

ls = pd.read_sql_query(ls, conn)
Employees = pd.read_sql_query(Employees, conn)
opp = pd.read_sql_query(opp, conn)

#-------------------------------------------------------

ls = ls.query('Визуал == 64627099')
ls['Визуал'] = ls['Визуал'].replace(64627099, 1)
ls.columns = ['ID сделки', 'Визуал']
ls['ID сделки'] = ls['ID сделки'].astype(int)

sch = product(sch, "СЧ")
rp = product(rp, "РП")
two_in_one = product(two_in_one, "2в1")
two_in_one['2в1'] = two_in_one['2в1'].replace(1, 2)

#display(sch.info())
#display(rp.info())
#display(two_in_one.info())

#-------------------------------------------------------

data['ОПЛАТА ПО ФАКТУ'] = pd.to_datetime(data['ОПЛАТА ПО ФАКТУ'])
data['Год_оплаты_факт'] = data['ОПЛАТА ПО ФАКТУ'].dt.year
data['Месяц'] = data['ОПЛАТА ПО ФАКТУ'].dt.month
data = data.query('Год_оплаты_факт == 2023')
data['ID сделки'] = data['ID сделки'].drop_duplicates()
data = data.query('Продукт_2 != "РП" & Продукт_2 != "сч" & Продукт_2 != "СЧ"')
data = data.query('Чистая >= 20000')
data['ID сделки'] = data['ID сделки'].fillna(0)
data['ID сделки'] = data['ID сделки'].astype(int)

data = data.merge(ls, how='left', on='ID сделки')
data = data.query('Визуал != 1')

rp = rp.merge(ls, how='left', on='ID сделки')
rp = rp.query('Визуал != 1')

sch = sch.merge(ls, how='left', on='ID сделки')
sch = sch.query('Визуал != 1')

two_in_one = two_in_one.merge(ls, how='left', on='ID сделки')
two_in_one = two_in_one.query('Визуал != 1')

data = data[["ID сделки", "КЦ", "ОТДЕЛ", "ОПЕРАТОР", "ОПЛАТА ПО ФАКТУ", 
             "ДАТА ПОДАЧИ", "Чистая", "Продукт", "Срок", "Форма", "Клиент", 
             "Примечание", "Год_оплаты_факт", "Продукт_2", "Месяц", "Визуал"]]


opp['kc'] = opp['kc'].replace("Отдел прямых продаж", "ОПП")

for i in opp['Form']:

    if "2 в 1 " in i:
        opp.loc[opp['Form'] == i, ['t']] = "2в1"
        
    elif "чтени" in i:
        opp.loc[opp['Form'] == i, ['t']] = "СЧ"    
        
    elif "память" in i:
        opp.loc[opp['Form'] == i, ['t']] = "РП"
    
    else:
        opp.loc[opp['Form'] == i, ['t']] = ""

opp = opp[['date_pay', 't']]

#-------------------------------------------------------

wb = load_workbook(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\ДОП.xlsx')

sheet = wb.get_sheet_by_name('Sheet_1')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(data, index=False, header=True):
    sheet.append(r)
    
sheet = wb.get_sheet_by_name('2в1')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(two_in_one, index=False, header=True):
    sheet.append(r)

sheet = wb.get_sheet_by_name('РП')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(rp, index=False, header=True):
    sheet.append(r)

sheet = wb.get_sheet_by_name('СЧ')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(sch, index=False, header=True):
    sheet.append(r)

sheet = wb.get_sheet_by_name('ОПП')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(opp, index=False, header=True):
    sheet.append(r)


wb.save(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\ДОП.xlsx')

#-------------------------------------------------------

xlApp = client.Dispatch("Excel.Application")
books = xlApp.Workbooks.Open(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\ДОП.xlsx')
ws = books.Worksheets[0]
ws.Visible = 1
ws.Range("B1:O50").ExportAsFixedFormat(0, r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Аналитика КМВ\PDF\ДОП.pdf')
books.Close(True)
#os.startfile(r'C:\Users\ADavydovskiy\Desktop\Аналитика\По конвертации в 1.pdf', 'print')


print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Готово!")
print("")