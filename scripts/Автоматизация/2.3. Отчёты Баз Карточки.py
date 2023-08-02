import pandas as pd
import numpy as np
import openpyxl as ox
import os
from stat import S_IREAD, S_IRGRP, S_IROTH, S_IWUSR

from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

from datetime import datetime, timedelta

import warnings
warnings.filterwarnings("ignore")

import string

import pyodbc

#now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Обновление отчётов баз карточек vpo, пожалуйста, подождите...")

conn = pyodbc.connect(r'Driver={SQL Server};Server=MSK1-BIDB01;Database=DWH;Trusted_Connection=yes;')
cursor = conn.cursor()

leads = f'''

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

declare @date date = getdate()-1
declare @month int = month(@date)
declare @edate date = DATEFROMPARTS(2022,12,31)
declare @startdate date = DATEFROMPARTS(2022,04,01);

SELECT 
month(L.[Дата лида]) as "Месяц"
       ,L.[Дата лида]
      --,L.[DATETIME_CREATE]
      --,L.[ID_REQUEST]
      ,L.[id_lead]
     -- ,cast(L.[Сумма оплат] as int) as "Выручка"
      ,cast(L.[Сумма оплат (успешные сделки)]as int) as "Выручка"
      ,cast(L.[Расход с НДС] as int) as "Расход"
      ,L.[CAMPAIGN_TYPE]
      ,L.[Источник (utm_source)]
      ,L.[SOURCE_LV1]
      ,L.[SOURCE_LV4]
      ,L.[SITE]
      --,L.[Term]
      ,L.[Ленд]
      ,L.[Форма отправки]
      ,L.[Lead from (ручная)]
    --  ,L.[Название формы]
      ,L.[Продукт бюджета_LV1]
      ,L.[Продукт бюджета_LV2]
      ,L.[Регион]
      ,L.[Статус лида]
      ,L.[Статус лида ур.1]
      ,L.[Статус лида ур.2]
      ,L.[is_new]
      ,L.[Version]
      ,L.[Ответственный]
      ,L.[Департамент ответсвенного_LV2]
      ,L.[Департамент ответсвенного_LV3]
      ,L.[Департамент ответсвенного_LV4]
      ,L.[Департамент ответсвенного_FULL]
      ,L.[Департамент ответсвенного_LV2 (для отчётов)]
      ,L.[Marketer]
      ,L.[Эфф. лид]
      ,L.[deals]
      ,D.[Вероятность сделки]
      ,D.[Сумма сделки]
      ,[Дата закрытия сделки]
      ,L.[INVOICES_WITH_PAY]
      ,L.[Клиент с оплатой]
      ,L.[Месяц первой оплаты лида]
      ,L.[Дата первой оплаты лида]
      ,L.[Месяцев до первой оплаты]
      ,L.[Дней до первой оплаты]
      ,L.[Первый проданный продукт (по лиду)]
      ,L.[Кампания (utm_campaign)] as Компания
 
  FROM [Analytic].[dbo].[DM_funnel_from_lead] L

  LEFT JOIN
  -- [Analytic].[dbo].[DM_full_funnel_from_deal] D on D.[id_lead] = L.[id_lead]

  (select*
  from [Analytic].[dbo].[DM_full_funnel_from_deal]
  where [Статус сделки]  in ('Отправлено в АКАДА / Sent to AKADA',
'Назначена личная встреча / Assigned a personal meeting',
'Переговоры в процессе / Negotiations in the process',
'Оплата по квитанции / Payment on receipt',
'Договор / Agreement')
and [Дата закрытия сделки] between @date and @edate
)
  D ON D.[id_lead] = L.[id_lead]

  WHERE
  
 L.[Департамент ответсвенного_LV2] in ('Коммерческий департамент (КМВ)')
 and L.[Дата лида] between @startdate and @date

'''
empl = f'''

SELECT [EMPLOYEES] as "Ответственный"
      ,CASE

       WHEN KC like '%4%' THEN 'КЦ 4'
       WHEN KC like '%3%' THEN 'КЦ 3'
       WHEN KC like '%центр 3%' THEN 'КЦ 3'
       WHEN KC like '%продаж 18%' THEN 'ОП 18'
             
       ELSE KC
  END
  as "KC"

         ,CASE

       WHEN [KC] like '%4%' THEN SUBSTRING([ОП],CHARINDEX(' ',[ОП])+1,LEN([ОП])) + ' ВР'
       WHEN ([ОП] in ('ОП 9')) and ([GP] in ('.1')) THEN [ОП] + [GP]
       WHEN ([ОП] in ('ОП 2')) and ([GP] in ('.1')) THEN 'ОП 8'
       WHEN ([ОП] in ('ОП 10')) and ([GP] in ('.2')) THEN 'ОП 10.2'
       WHEN ([ОП] in ('ОП 5')) and ([GP] in ('.1')) THEN [ОП] + [GP]
       WHEN ([ОП] in ('3 ЯР')) and ([GP] in ('.1')) THEN '3.1 ЯР'
       WHEN ([KC] in ('Отдел прямых продаж'))  THEN 'ОПП'
           WHEN [KC] like '%продаж 18%'  THEN 'ОП 18'  
       ELSE [ОП]
  END
  as "ОП"


        ,CASE

       WHEN [KC] = 'КЦ 3' THEN 'КЦ 3'
       WHEN [KC] like '%4%' THEN 'КЦ 4'
       WHEN [ОП] = 'ОП 5' THEN 'Группа № 1'
       WHEN [ОП] = 'ОП 5.1' THEN 'Группа № 1'
       WHEN [ОП] = 'ОП 7' THEN 'Группа № 1'
       WHEN [ОП] = 'ОП 8' THEN 'Группа № 1'
       WHEN [ОП] = 'ОП 9' THEN 'Группа № 1'
       WHEN [ОП] = 'ОП 10' THEN 'Группа № 1'
       WHEN [ОП] = 'ОП 16' THEN 'Группа № 1'
       WHEN [ОП] = 'ОП 6' THEN 'Группа № 2'
       WHEN [ОП] = 'ОП 1' THEN 'Группа № 2'
       WHEN [ОП] = 'ОП 10.1' THEN 'Группа № 2'
       WHEN [ОП] = 'ОП 12' THEN 'Группа № 2'
       WHEN [ОП] = 'ОП 13' THEN 'Группа № 2'
       WHEN [ОП] = 'ОП 15' THEN 'Группа № 2'
       WHEN [ОП] = 'ОП 3' THEN 'ОП 3'
       WHEN [ОП] like '%14%' THEN 'ОП 14'
       ELSE 'др'
  END
  as "Группа"






  FROM [DWH].[dbo].[KHTS_EMPL]
  WHERE 
  [SP] in ('КД')
  AND [STATUS] = 'работает'
 

'''

empl = pd.read_sql_query(empl, conn)
leads = pd.read_sql_query(leads, conn)

data = pd.read_excel(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\База ЧВ 0,25.xlsx', 
                     sheet_name='База ЧВ')

data = data.loc[data['ID лида'] != 0]
data = data.loc[data['ОПЛАТА ПО ФАКТУ'] >= "20.10.2022"]
data = data.pivot_table(index='ID лида', values='Чистая', aggfunc='sum').reset_index()
data = data.rename(columns={'ID лида': 'id_lead'})

leads = leads.merge(data, on='id_lead', how='left')
leads = leads.merge(empl, on='Ответственный', how='left')

leads = leads[['Месяц', 'id_lead', 'Выручка', 'Чистая', 'Источник (utm_source)', 'Ленд',
               'Регион', 'Статус лида', 'Ответственный', 'Департамент ответсвенного_LV3',
               'Департамент ответсвенного_LV4', 'Вероятность сделки', 'Дата закрытия сделки', 
               'Сумма сделки', 'deals', 'INVOICES_WITH_PAY', 'Дней до первой оплаты',
               'Первый проданный продукт (по лиду)', 'Дата лида', 'Группа', 'ОП', 'Version', 'Компания']]

civil_law = leads.query('Version == "civil_law_baccalaureate"')
med = leads.query('Компания == "МЕД_КД"')
isit = leads.query('Компания == "ИСИТ_КД_vpo"')
programm = leads.query('Компания == "программирование - повтор"')
yr = leads.query('Компания == "юр фак-повтор"')\

y = S_IREAD|S_IRGRP|S_IROTH
n = S_IWUSR|S_IREAD

def chek(first):
    chekpoint = first
    l = ['Отчёт База civil_law', 'Отчёт База Med_KD', 
         'Отчёт База Исит vpo', 'Отчёт База Исит synru',
         'Отчёт База Программирование', 'Отчёт База юр фак']
    for index in l:
        filename = f'\\\\synergy.local\\Documents\\11.Коммерческий департамент\\01. Аналитика КД\\06. Общая аналитика\\Факультеты\\Отработка баз\\{index}.xlsx'
        os.chmod(filename, chekpoint)

chek(n)

wb = load_workbook(r'\\synergy.local\Documents\11.Коммерческий департамент\01. Аналитика КД\06. Общая аналитика\Факультеты\Отработка баз\Отчёт База civil_law.xlsx')

sheet = wb.get_sheet_by_name('Выгрузка')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(civil_law, index=True, header=True):
    sheet.append(r)
    
wb.save(r'\\synergy.local\Documents\11.Коммерческий департамент\01. Аналитика КД\06. Общая аналитика\Факультеты\Отработка баз\Отчёт База civil_law.xlsx')

wb = load_workbook(r'\\synergy.local\Documents\11.Коммерческий департамент\01. Аналитика КД\06. Общая аналитика\Факультеты\Отработка баз\Отчёт База Med_KD.xlsx')

sheet = wb.get_sheet_by_name('Выгрузка')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(med, index=True, header=True):
    sheet.append(r)
    
wb.save(r'\\synergy.local\Documents\11.Коммерческий департамент\01. Аналитика КД\06. Общая аналитика\Факультеты\Отработка баз\Отчёт База Med_KD.xlsx')

wb = load_workbook(r'\\synergy.local\Documents\11.Коммерческий департамент\01. Аналитика КД\06. Общая аналитика\Факультеты\Отработка баз\Отчёт База Исит synru.xlsx')

sheet = wb.get_sheet_by_name('Выгрузка')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(isit, index=True, header=True):
    sheet.append(r)
    
wb.save(r'\\synergy.local\Documents\11.Коммерческий департамент\01. Аналитика КД\06. Общая аналитика\Факультеты\Отработка баз\Отчёт База Исит synru.xlsx')

wb = load_workbook(r'\\synergy.local\Documents\11.Коммерческий департамент\01. Аналитика КД\06. Общая аналитика\Факультеты\Отработка баз\Отчёт База Исит vpo.xlsx')

sheet = wb.get_sheet_by_name('Выгрузка')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(isit, index=True, header=True):
    sheet.append(r)
    
wb.save(r'\\synergy.local\Documents\11.Коммерческий департамент\01. Аналитика КД\06. Общая аналитика\Факультеты\Отработка баз\Отчёт База Исит vpo.xlsx')

wb = load_workbook(r'\\synergy.local\Documents\11.Коммерческий департамент\01. Аналитика КД\06. Общая аналитика\Факультеты\Отработка баз\Отчёт База Программирование.xlsx')

sheet = wb.get_sheet_by_name('Выгрузка')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(programm, index=True, header=True):
    sheet.append(r)
    
wb.save(r'\\synergy.local\Documents\11.Коммерческий департамент\01. Аналитика КД\06. Общая аналитика\Факультеты\Отработка баз\Отчёт База Программирование.xlsx')

wb = load_workbook(r'\\synergy.local\Documents\11.Коммерческий департамент\01. Аналитика КД\06. Общая аналитика\Факультеты\Отработка баз\Отчёт База юр фак.xlsx')

sheet = wb.get_sheet_by_name('Выгрузка')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(yr, index=True, header=True):
    sheet.append(r)
    
wb.save(r'\\synergy.local\Documents\11.Коммерческий департамент\01. Аналитика КД\06. Общая аналитика\Факультеты\Отработка баз\Отчёт База юр фак.xlsx')

chek(y)

print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Готово!")
print("")