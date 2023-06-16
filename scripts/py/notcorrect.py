import os

import pandas as pd
import numpy as np
import openpyxl as ox
import re

from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

from datetime import datetime, timedelta

import warnings
warnings.filterwarnings("ignore")

import string

import pyodbc

print("Артём бдително отделяет корректные сделки от некорректных, пожалуйста, подождите...")

conn = pyodbc.connect(r'Driver={SQL Server};Server=MSK1-BIDB01;Database=DWH;Trusted_Connection=yes;')
cursor = conn.cursor()

Employees = f'''

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

SELECT [EMPLOYEES]
      ,[LOG]
      ,[SP]

      ,CASE

       WHEN KC like '%4%' THEN 'КЦ 4'
       WHEN KC like '%3%' THEN 'КЦ 3'
       WHEN KC like '%центр 3%' THEN 'КЦ 3'
         WHEN KC like '%продаж 18%' THEN 'ОП 18'
             
       ELSE KC
  END
  as "КЦ"

         ,CASE

       WHEN [KC] like '%4%' THEN SUBSTRING([ОП],CHARINDEX(' ',[ОП])+1,LEN([ОП])) + ' ВР'
       WHEN ([ОП] in ('ОП 9')) and ([GP] in ('.1')) THEN [ОП] + [GP]
       WHEN ([ОП] in ('ОП 2')) and ([GP] in ('.1')) THEN 'ОП 8'
       WHEN ([ОП] in ('ОП 10')) and ([GP] in ('.2')) THEN [ОП] + [GP]
       WHEN ([ОП] in ('ОП 5')) and ([GP] in ('.1')) THEN [ОП] + [GP]
       WHEN ([ОП] in ('ОП 14')) and ([GP] in ('.1')) THEN [ОП] + [GP]
       WHEN ([ОП] in ('3 ЯР')) and ([GP] in ('.1')) THEN '3.1 ЯР'
         WHEN ([KC] in ('Отдел прямых продаж'))  THEN 'ОПП'
           WHEN [KC] like '%продаж 18%'  THEN 'ОП 18'  
       ELSE [ОП]
  END
  as "ОП"
      
               ,[PHONE]
      ,[STATUS]
      ,[ID_EMPL]
      ,[ID_ORG]

  FROM [DWH].[dbo].[KHTS_EMPL]
  where 
  --KC like '%4%' 
  [SP] in ('КД')

'''

Employees = pd.read_sql_query(Employees, conn)
Employees = Employees.rename(columns={'EMPLOYEES': 'Ответственный'})
Employees = Employees[['Ответственный', 'КЦ', 'ОП']]

not_cor = pd.read_excel('C:\\Users\\ADavydovskiy\\scripts\\Выгрузки\\NotCorrectDealsByDepartmentAndDate.xlsx', sheet_name='Лист3')
over = pd.read_excel('C:\\Users\\ADavydovskiy\\scripts\\Выгрузки\\Overdue_affairs_in_deals.xlsx', sheet_name='Лист3')


not_cor=not_cor[['ID', 'Название сделки', 'Стадия сделки', 'Дата создания',
       'Дата Изменения', 'Предложение', 'Ответственный', 'Контакт', 'Компания',
       'Ленд', 'Дата отправки СМС Email', 'Товар', 'Группа продуктов',
       'Кампания utm', 'Сумма', 'ID Лида']]

not_cor = not_cor.merge(Employees, how='left', on='Ответственный')

not_cor=not_cor[['КЦ', 'ОП', 'Ответственный', 'ID', 'Название сделки', 'Стадия сделки', 'Дата создания',
       'Дата Изменения', 'Предложение', 'Ответственный', 'Контакт', 'Компания',
       'Ленд', 'Дата отправки СМС Email', 'Товар', 'Группа продуктов',
       'Кампания utm', 'Сумма', 'ID Лида']]

over = over[5:].reset_index()

new_header = over.iloc[0]
over = over[1:]
over.columns = new_header

over = over[['Менеджер', 'Всего сделок', 'Сделок без дел', 'Сделок с просроченными делами']]
over.columns = ['Ответственный', 'Всего сделок', 'Сделок без дел', 'Сделок с просроченными делами']

over = over.dropna()

over = over.loc[over['Ответственный'].str.contains('Итого')==False]

new_df = over['Ответственный'].str.split(' ',expand=True)
new_df.columns = ['0', '1', '2', '3', '4', '5', '6', '7', '8']
over['Ответственный'] = new_df['5'].map(str) + ' ' + new_df['6'].map(str) + ' ' + new_df['7'].map(str)

over['Ответственный'] = over['Ответственный'].replace("     ", "")
over = over.merge(Employees, how='left', on='Ответственный')

over = over[['КЦ', 'ОП', 'Ответственный', 'Всего сделок', 
             'Сделок без дел', 'Сделок с просроченными делами']]

all_not_cor = not_cor.pivot_table(index='ОП', values='ID', aggfunc='count')

nc_bgp = not_cor.loc[not_cor['Группа продуктов'] != 'Курсы']
nc_bgp_pivot = nc_bgp.pivot_table(index='ОП', values='ID', aggfunc='count')
all_not_cor = all_not_cor.merge(nc_bgp_pivot, how='left', on='ОП')

not_cor['Товар'] = not_cor['Товар'].fillna('0')
nc_bnt = not_cor[not_cor['Товар'] == '0']
nc_bnt_pivot = nc_bnt.pivot_table(index='ОП', values='ID', aggfunc='count')
all_not_cor = all_not_cor.merge(nc_bnt_pivot, how='left', on='ОП')

not_cor['Контакт'] = not_cor['Контакт'].fillna('0')
nc_bk = not_cor[not_cor['Контакт'] == '0']
nc_bk_pivot = nc_bk.pivot_table(index='ОП', values='ID', aggfunc='count')
all_not_cor = all_not_cor.merge(nc_bk_pivot, how='left', on='ОП')

over_pivot = over.pivot_table(index='ОП',
                              values=['Сделок с просроченными делами', 
                                       'Сделок без дел', 'Всего сделок'],
                              aggfunc='count')

all_pivot = over_pivot.merge(all_not_cor, how='left', on='ОП')
all_pivot.columns = ['Всего сделок', 'Сделок без дел', 'Сделок с просроченными делами',
                     'Всего некорректных', 'Без группы продуктов', 'Без наименования товара', 
                     'Без контакта']
all_pivot = all_pivot.fillna(0).astype(int)

wb = load_workbook('C:\\Users\\ADavydovskiy\\scripts\\Отчёты\\Некорректные.xlsx')

sheet = wb.get_sheet_by_name('Сводник_1')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(all_pivot, index=True, header=True):
    sheet.append(r)


wb.save('C:\\Users\\ADavydovskiy\\scripts\\Отчёты\\Некорректные.xlsx')