import os
from tqdm import tqdm

import pandas as pd
import numpy as np
import openpyxl as ox
import re
import time

from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

from datetime import timedelta
from datetime import datetime

import warnings
warnings.filterwarnings("ignore")

import string

import pyodbc
import win32com.client
from PIL import ImageGrab
import pywhatkit


d = datetime.today()
d = (d - timedelta(days=1)).strftime('%Y-%m-%d')
d

conn = pyodbc.connect(r'Driver={SQL Server};Server=MSK1-BIDB01;Database=DWH;Trusted_Connection=yes;')
cursor = conn.cursor()

ls = f'''

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

declare @edate date = getdate()-1
declare @month int = month(getdate()-1)
declare @sdate date = DATEFROMPARTS(2023, @month, 01);


SELECT DISTINCT
           cast(dr.[DATE_OP] as date)  as "DATE"
           ,month(dr.[DATE_OP]) as "month"
           ,ul.UF_CRM_1464341216 as "Город (IP)"
           ,Rdop.CODE_LABEL_BASE_KD as "Метка базы"
           ,iif(dsr.[NAME] not like '%/%',dsr.[NAME], SUBSTRING(dsr.[NAME],0,PATINDEX('% / %',dsr.[NAME])))   as "Статус"             
           ,concat(e.[LAST_NAME]+' ',e.[NAME]+' ',e.[SECOND_NAME]) as "Ответственный"
           ,fc.[SPENT_BY_REQ] as "ras"         
           ,dr.[CODE] as id  
           ,UTS_D.[UF_CRM_ID_MEETING]  as "Встреча id"
           ,cast(tt.[UF_PLAN_MEET_DATE] as date) as "Дата вс план"
           ,cast(tt.[UF_FACT_MEET_DATE] as date) as "Дата вс факт"
           ,tt.[UF_MEETING_STATUS] as "Статус вс"
             

FROM [DWH].[dbo].[DIC_REQUEST] dr 

            left join [DWH].[dbo].[DIC_REQUEST_UTM] RU on RU.[ID_REQUEST] = DR.[ID_REQUEST]

-- статусы

            left join  [DWH].[dbo].[DIC_STATUS_REQUEST] dsr on dsr.ID_STATUS_REQUEST = dr.ID_STATUS_REQUEST

-- источник

LEFT JOIN [DWH].[STAGE].[CRM_B_UTS_CRM_LEAD] UL WITH(NOLOCK) ON UL.VALUE_ID = dr.CODE
LEFT JOIN [DWH].[STAGE].[CRM_B_CRM_LEAD] L WITH(NOLOCK) ON L.ID = dr.CODE
INNER JOIN (SELECT [NAME], [STATUS_ID]
           FROM [DWH].[STAGE].[CRM_B_CRM_STATUS] WITH(NOLOCK)
           WHERE [ENTITY_ID] = 'SOURCE' AND (STATUS_ID='WEB' or STATUS_ID= '7')
           ) SRC ON SRC.[STATUS_ID] = L.[SOURCE_ID]

            LEFT JOIN ASS_REQUEST_SOURCE ARS WITH(NOLOCK) ON dR.ID_REQUEST = ARS.ID_REQUEST
            LEFT JOIN DIC_SOURCE S WITH(NOLOCK) ON ARS.ID_SOURCE = S.ID_SOURCE
            
-- исполнители
LEFT JOIN [DWH].[dbo].DIC_EMPLOYEES E With(nolock) ON dr.Id_EMPLOYEES = E.ID_EMPLOYEES
JOIN (SELECT * FROM  [DWH].[dbo].[v_DIC_ORG_EMPL] where [lv1_NAME] like '%Коммерческий департамент (КМВ)%') org on org.[ID_EMPLOYEES] = E.ID_EMPLOYEES

-- продукт бюджета

            LEFT JOIN [DWH].[dbo].ASS_REQUEST_PRODUCT_BUDGET RPB With(nolock) ON dR.ID_REQUEST=RPB.ID_REQUEST and dR.DATE_CREATE= RPB.R_DATE
            LEFT JOIN [DWH].[dbo].DIC_PRODUCT_BUDGET PB With(nolock) ON RPB.ID_PRODUCT_BUDGET=PB.ID_PRODUCT_BUDGET

            --расход
            LEFT JOIN [DWH].[dbo].[FCT_REQUEST] fc With(nolock) on fc.[ID_REQUEST] = dR.ID_REQUEST
            -- доп. коды лидов
            LEFT JOIN [DWH].[dbo].[DIC_REQUEST_STAT] Rdop With(nolock) on dR.ID_REQUEST = Rdop.ID_REQUEST

            -- сделки
            LEFT JOIN [DWH].[dbo].ASS_REQUEST_DEAL ASS_RD With(nolock) ON ASS_RD.ID_REQUEST = dR.ID_REQUEST
            LEFT JOIN [DWH].[dbo].[DIC_DEAL] D With(nolock) ON D.ID_DEAL = ASS_RD.ID_DEAL
            -- инвойсы
            LEFT JOIN [DWH].[dbo].ASS_DEAL_INVOICE ASS_DI With(nolock) ON ASS_DI.ID_DEAL = ASS_RD.ID_DEAL
            LEFT JOIN [DWH].[dbo].DIC_INVOICE D_I With(nolock) ON D_I.ID_INVOICE = ASS_DI.ID_INVOICE
            LEFT JOIN [DWH].[dbo].DIC_STATUS_INVOICE S_I With(nolock) ON S_I.ID_STATUS_INVOICE = D_I.ID_STATUS_INVOICE
            --встречи
             LEFT JOIN [stage].[CRM_b_uts_crm_deal] UTS_D ON UTS_D.VALUE_ID = D.code
             LEFT JOIN [stage].[CRM_b_uts_tasks_task] tt ON tt.[VALUE_ID] = UTS_D.[UF_CRM_ID_MEETING]
             -- выбор телефона и почты
            LEFT JOIN [DWH].[dbo].DIC_CLIENT_CRM dcc With(nolock) on dcc.code = dR.CODE and ID_TYPE_CLIENT_CRM = 2
            left join [DWH].[dbo].[ASS_CLIENT_CRM_TELEPHONE] act With(nolock) on act.ID_CLIENT_CRM = dcc.ID_CLIENT_CRM
            left join [DWH].[dbo].[DIC_TELEPHONE] dt With(nolock) on dt.ID_TELEPHONE = act.ID_TELEPHONE
            -- адрес по телефону для поля Регион
            left join [DWH].[dbo].[DIC_TYPE_TELEPHONE] dtt With(nolock) on dtt.ID_TYPE_TELEPHONE = act.ID_TYPE_TELEPHONE
            left join [DWH].[dbo].[v_DIC_ADDRESS_LIM] va With(nolock) on va.ID_ADDRESS = dt.ID_ADDRESS 


WHERE 1=1 

and cast(dr.[DATE_OP] as date) between @sdate and @edate
--and org.full_NAME like '%\Коммерческий департамент (КМВ)\%'
and org.[NAME] not like '%МАП%'
and va.full_NAME like '%Москва%'
AND dsr.[NAME] NOT IN ('Повторные заявки / Reapplication','') --убрать -- для эфф лидов

  
'''


Employees = f'''

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

SELECT [EMPLOYEES]
      ,[LOG]
      ,[SP]

      ,CASE

       WHEN KC like '%4%' THEN 'КЦ 4'
       WHEN KC like '%3%' THEN 'КЦ 3'
       WHEN KC like '%центр 3%' THEN 'КЦ 3'
       WHEN KC like '%продаж 18%' THEN 'ОП 18'
             
       ELSE KC
  END
  as "KC"

         ,CASE

       WHEN [KC] like '%4%' THEN SUBSTRING([ОП],CHARINDEX(' ',[ОП])+1,LEN([ОП])) + ' ВР'
       WHEN ([ОП] in ('ОП 9')) and ([GP] in ('.1')) THEN [ОП] + [GP]
       WHEN ([ОП] in ('ОП 2')) and ([GP] in ('.1')) THEN 'ОП 8'
       WHEN ([ОП] in ('ОП 10')) and ([GP] in ('.2')) THEN 'ОП 10.2'
       WHEN ([ОП] in ('ОП 5')) and ([GP] in ('.1')) THEN [ОП] + [GP]
       WHEN ([ОП] in ('3 ЯР')) and ([GP] in ('.1')) THEN '3.1 ЯР'
       WHEN ([KC] in ('Отдел прямых продаж'))  THEN 'ОПП'
           WHEN [KC] like '%продаж 18%'  THEN 'ОП 18'  
       ELSE [ОП]
  END
  as "ОП"
      
               ,[PHONE]
      ,[STATUS]
      ,[ID_EMPL]
      ,[ID_ORG]

  FROM [DWH].[dbo].[KHTS_EMPL]
  where 
  --KC like '%4%' 
  [SP] in ('КД')


'''

ls = pd.read_sql_query(ls, conn)
Employees = pd.read_sql_query(Employees, conn)

ls['DATE'] = pd.to_datetime(ls['DATE'])
ls['Дата вс план'] = pd.to_datetime(ls['Дата вс план'])
ls['Дата вс факт'] = pd.to_datetime(ls['Дата вс факт'])
ls.rename({'Метка базы': 'Метка_базы',}, axis=1, inplace=True)
ls = ls.query('Метка_базы != "ЧБ" & Метка_базы != "ЧБ "')
ls = ls.query('Статус == "Назначен ответственный" | Статус == "Перезвон" | Статус == "Перспектива" | Статус == "Сконвертирован"')


Employees.rename({'EMPLOYEES': 'Ответственный',}, axis=1, inplace=True)
Employees = Employees[['Ответственный', 'KC', 'ОП']]


ls = ls.merge(Employees, how='left', on='Ответственный')


all_date_1_ОП = ls.pivot_table(index='ОП', values='id', aggfunc='count')
all_date_2_ОП = ls.pivot_table(index='ОП', values='Встреча id', aggfunc='count')
all_date_ОП = all_date_1_ОП.merge(all_date_2_ОП, how='outer', on='ОП')
all_date_ОП.rename({'ОП': 'СП',}, axis=1, inplace=True)

all_date_1_КЦ = ls.pivot_table(index='KC', values='id', aggfunc='count')
all_date_2_КЦ = ls.pivot_table(index='KC', values='Встреча id', aggfunc='count')
all_date_КЦ = all_date_1_КЦ.merge(all_date_2_КЦ, how='outer', on='KC')
all_date_КЦ.rename({'KC': 'СП',}, axis=1, inplace=True)

all_date = pd.concat([all_date_ОП, all_date_КЦ])


ls_ = ls.query('DATE == @d')

all_date_1_ОП_d = ls_.pivot_table(index='ОП', values='id', aggfunc='count')
all_date_2_ОП_d = ls_.pivot_table(index='ОП', values='Встреча id', aggfunc='count')
all_date_ОП_d = all_date_1_ОП_d.merge(all_date_2_ОП_d, how='outer', on='ОП')
all_date_ОП_d.rename({'ОП': 'СП',}, axis=1, inplace=True)

all_date_1_КЦ_d = ls_.pivot_table(index='KC', values='id', aggfunc='count')
all_date_2_КЦ_d = ls_.pivot_table(index='KC', values='Встреча id', aggfunc='count')
all_date_КЦ_d = all_date_1_КЦ_d.merge(all_date_2_КЦ_d, how='outer', on='KC')
all_date_КЦ_d.rename({'KC': 'СП',}, axis=1, inplace=True)

all_date_d = pd.concat([all_date_ОП_d, all_date_КЦ_d])


wb = load_workbook('C:\\Users\\ADavydovskiy\\scripts\\Отчёты\\Отчёт ЛВ.xlsx')

sheet = wb.get_sheet_by_name('Sheet_1')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(all_date, index=True, header=True):
    sheet.append(r)

sheet = wb.get_sheet_by_name('Sheet_2')
sheet.delete_rows(1, sheet.max_row)    

for r in dataframe_to_rows(all_date_d, index=True, header=True):
    sheet.append(r)


wb.save('C:\\Users\\ADavydovskiy\\scripts\\Отчёты\\Отчёт ЛВ.xlsx')

print("")
print("Готово! Отправяю картинку на зелёную почту!))")

if os.path.isfile('C:\\Users\\ADavydovskiy\\scripts\\Скрины\\Отчёт ЛВ.jpg'): 
    os.remove('C:\\Users\\ADavydovskiy\\scripts\\Скрины\\Отчёт ЛВ.jpg') 
    print("Предыдущая картинка успешно удалена") 
else: 
    print("Очередной фейл!")


client = win32com.client.Dispatch("Excel.Application")
wb = client.Workbooks.Open('C:\\Users\\ADavydovskiy\\scripts\\Отчёты\\Отчёт ЛВ.xlsx')
ws = wb.Worksheets('Отчёт')

ws.Range("B1:K38").CopyPicture(Format = 2)
wb.Close()
client.Quit()

img = ImageGrab.grabclipboard()
img.save('C:\\Users\\ADavydovskiy\\scripts\\Скрины\\Отчёт ЛВ.jpg')


#pywhatkit.sendwhats_image("+79919249593", "C:\\Users\\ADavydovskiy\\scripts\\Скрины\\Отчёт ЛВ.jpg", wait_time = 30)
#time.sleep(10)