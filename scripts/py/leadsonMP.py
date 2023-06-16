import os

import pandas as pd
import numpy as np
import openpyxl as ox
import re

from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

from datetime import datetime, timedelta

import warnings
warnings.filterwarnings("ignore")

import string

import pyodbc

print("Артём старательно распределяет лиды на МП, пожалуйста, подождите...")

day_m = datetime.today()

d = day_m.day
m = day_m.month

if m == 1:
    month = 'Январь'
elif m == 2:
    month = 'Февраль'
elif m == 3:
    month = 'Март'
elif m == 4:
    month = 'Апрель'
elif m == 5:
    month = 'Май'
elif m == 6:
    month = 'Июнь'
elif m == 7:
    month = 'Июль'
elif m == 8:
    month = 'Август'
elif m == 9:
    month = 'Сентябрь'
elif m == 10:
    month = 'Октябрь'
elif m == 11:
    month = 'Ноябрь'
else:
    month = 'Декабрь'

conn = pyodbc.connect(r'Driver={SQL Server};Server=MSK1-BIDB01;Database=DWH;Trusted_Connection=yes;')
cursor = conn.cursor()

leads_eff = f'''

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

declare @date date = getdate()-1
declare @month int = month(getdate()-1)
declare @startdate date = DATEFROMPARTS(2022, @month, 12)
declare @endate date = @date;

with t as(
select distinct
dr.[DATE_CREATE]    as "DATE"
,dr.[DATE_OP] as "Дата соединения на ОП"
,year(dr.[DATE_CREATE]) as "year"

,[FEATURES_6]  as "Старый/новый"       
,dr.[FEATURES_5] as "Ленд"
,CASE

    WHEN s.tag1 like '%органика%' THEN 'органика'
    WHEN s.tag1 like '%yandex%' THEN 'Яндекс'
    WHEN s.tag1 like '%google%' THEN 'Google'
    WHEN s.tag1 like '%facebook%' THEN 'facebook'
    WHEN s.tag1 like '%edunetwork%' THEN 'edunetwork'
    WHEN s.tag1 like '%(пусто)%' THEN 'органика'
    WHEN s.tag1 like '%TikTok%' THEN 'TikTok'
    WHEN s.tag1 like '%рассылка%' THEN 'рассылка'
    WHEN dr.[FEATURES_1] like '%studika%' THEN 'studika'
    
    ELSE 'прочее'
END as "tag"


,dr.[FEATURES_1] as "source"
,fc.[SPENT_BY_REQ] as "Расход"
,iif(dsr.[NAME] not like '%/%',dsr.[NAME], SUBSTRING(dsr.[NAME],0,PATINDEX('% / %',dsr.[NAME])))   as "Статус"              
,concat(e.[LAST_NAME]+' ',e.[NAME]+' ',e.[SECOND_NAME]) as "Ответственный"
,CASE

    WHEN org.full_NAME like '%центр 1%' THEN 'КЦ 1'
    WHEN org.full_NAME like '%центр 2%' THEN 'КЦ 2'
    WHEN org.full_NAME like '%центр 3%' THEN 'КЦ 3'
    WHEN org.full_NAME like '%центр 4%' THEN 'КЦ 4'
    
    ELSE 'др'
  END as "СП"

,RU.[Marketer] as Marketer
,dr.[CODE] as id  
  

from [DWH].[dbo].[DIC_REQUEST] dr 

            left join [DWH].[dbo].[DIC_REQUEST_UTM] RU on RU.[ID_REQUEST] = DR.[ID_REQUEST]
-- статусы
            left join  [DWH].[dbo].[DIC_STATUS_REQUEST] dsr on dsr.ID_STATUS_REQUEST = dr.ID_STATUS_REQUEST
-- источник
            left join [DWH].[stage].[CRM_b_uts_crm_lead] ul on ul.value_id = dr.CODE
            left join [DWH].[stage].[CRM_b_crm_lead] l on l.ID = dr.code
            left join (SELECT [NAME] as name_source, [STATUS_ID]
            FROM [DWH].[stage].[CRM_b_crm_status]
            where [ENTITY_ID] = 'SOURCE') src on src.[STATUS_ID] = l.[SOURCE_ID]

              LEFT JOIN ASS_REQUEST_SOURCE ARS WITH(NOLOCK) ON dR.ID_REQUEST = ARS.ID_REQUEST
              LEFT JOIN DIC_SOURCE S WITH(NOLOCK) ON ARS.ID_SOURCE = S.ID_SOURCE
-- исполнители
            LEFT JOIN [DWH].[dbo].DIC_EMPLOYEES E ON dR.Id_EMPLOYEES = E.ID_EMPLOYEES
            LEFT JOIN [DWH].[dbo].[ASS_EMPLOYEE_AND_ORGSTRUCTURE] ASS_OS ON E.ID_EMPLOYEES = ASS_OS.ID_EMPLOYEES
-- оргструктура
            LEFT JOIN [DWH].[dbo].[v_DIC_ORGSTRUCTURE] org on org.ID_ORGSTRUCTURE = ISNULL(ASS_OS.ID_ORGSTRUCTURE,-1)
-- продукт бюджета
            LEFT JOIN [DWH].[dbo].ASS_REQUEST_PRODUCT_BUDGET RPB With(nolock) ON dR.ID_REQUEST=RPB.ID_REQUEST and dR.DATE_CREATE= RPB.R_DATE
            LEFT JOIN [DWH].[dbo].DIC_PRODUCT_BUDGET PB With(nolock) ON RPB.ID_PRODUCT_BUDGET=PB.ID_PRODUCT_BUDGET
--расход
            LEFT JOIN [DWH].[dbo].[FCT_REQUEST] fc With(nolock) on fc.[ID_REQUEST] = dR.ID_REQUEST

where 1=1 

and isnull(dr.SIGN_DELETED,0) =0
and year(dr.[DATE_CREATE]) in ('2021', '2022','2023')
and month(dr.[DATE_CREATE]) = @month
and day(dr.[DATE_CREATE]) <= day(getdate()-1)
and org.full_NAME like '%Коммерческий департамент\%'
and org.name_2 not like '%МАП%'
and src.name_source like '%Веб%'
aND dsr.[NAME] NOT IN ('Дубль / Double','Спам / Spam','Ошибка номера / Error number', 'Спам / Spam','Повторные заявки / Reapplication')
and dr.[FEATURES_5] not in ('estr2022' , 'ege_rf', 'leadform_CZ_747828', 'leadform_CZN_SYN_PP_713311','proftest')
)

select
t.[year]
, count (t.id) as "Кол"

from t
group by 
t.[year]

'''

leads = f'''

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

declare @date date = getdate()-1
declare @month int = month(getdate())
declare @startdate date = DATEFROMPARTS(2022, @month, 12)
declare @endate date = @date;

select distinct
              
dr.[DATE_CREATE]    as "DATE"
,dr.[DATE_OP] as "Дата соединения на ОП"
,year(dr.[DATE_OP]) as "Год"

,[FEATURES_6]  as "Старый/новый"       
,dr.[FEATURES_5] as "Ленд"

,CASE

    WHEN s.tag1 like '%органика%' THEN 'органика'
    WHEN s.tag1 like '%yandex%' THEN 'Яндекс'
    WHEN s.tag1 like '%google%' THEN 'Google'
    WHEN s.tag1 like '%facebook%' THEN 'facebook'
    WHEN s.tag1 like '%edunetwork%' THEN 'edunetwork'
    WHEN s.tag1 like '%(пусто)%' THEN 'органика'
    WHEN s.tag1 like '%TikTok%' THEN 'TikTok'
    WHEN s.tag1 like '%рассылка%' THEN 'рассылка'
    WHEN dr.[FEATURES_1] like '%studika%' THEN 'studika'
    
    ELSE 'прочее'
  END as "tag"


,dr.[FEATURES_1] as "source"
,fc.[SPENT_BY_REQ] as "Расход"
,iif(dsr.[NAME] not like '%/%',dsr.[NAME], SUBSTRING(dsr.[NAME],0,PATINDEX('% / %',dsr.[NAME])))   as "Статус"              
,concat(e.[LAST_NAME]+' ',e.[NAME]+' ',e.[SECOND_NAME]) as "Ответственный"
,CASE

    WHEN org.full_NAME like '%центр 1%' THEN 'КЦ 1'
    WHEN org.full_NAME like '%центр 2%' THEN 'КЦ 2'
    WHEN org.full_NAME like '%центр 3%' THEN 'КЦ 3'
    WHEN org.full_NAME like '%центр 4%' THEN 'КЦ 4'
    ELSE 'др'
  END as "СП"

,RU.[Marketer] as Marketer
,dr.[CODE] as id  
             

         from [DWH].[dbo].[DIC_REQUEST] dr 

            left join [DWH].[dbo].[DIC_REQUEST_UTM] RU on RU.[ID_REQUEST] = DR.[ID_REQUEST]
-- статусы
            left join  [DWH].[dbo].[DIC_STATUS_REQUEST] dsr on dsr.ID_STATUS_REQUEST = dr.ID_STATUS_REQUEST
-- источник
            left join [DWH].[stage].[CRM_b_uts_crm_lead] ul on ul.value_id = dr.CODE
            left join [DWH].[stage].[CRM_b_crm_lead] l on l.ID = dr.code
            left join (SELECT [NAME] as name_source, [STATUS_ID]
               FROM [DWH].[stage].[CRM_b_crm_status]
               where [ENTITY_ID] = 'SOURCE') src on src.[STATUS_ID] = l.[SOURCE_ID]

                  LEFT JOIN ASS_REQUEST_SOURCE ARS WITH(NOLOCK) ON dR.ID_REQUEST = ARS.ID_REQUEST
                  LEFT JOIN DIC_SOURCE S WITH(NOLOCK) ON ARS.ID_SOURCE = S.ID_SOURCE
-- исполнители

LEFT JOIN [DWH].[dbo].DIC_EMPLOYEES E ON dR.Id_EMPLOYEES = E.ID_EMPLOYEES
LEFT JOIN [DWH].[dbo].[ASS_EMPLOYEE_AND_ORGSTRUCTURE] ASS_OS ON E.ID_EMPLOYEES = ASS_OS.ID_EMPLOYEES
-- оргструктура
            LEFT JOIN [DWH].[dbo].[v_DIC_ORGSTRUCTURE] org on org.ID_ORGSTRUCTURE = ISNULL(ASS_OS.ID_ORGSTRUCTURE,-1)
-- продукт бюджета
            LEFT JOIN [DWH].[dbo].ASS_REQUEST_PRODUCT_BUDGET RPB With(nolock) ON dR.ID_REQUEST=RPB.ID_REQUEST and dR.DATE_CREATE= RPB.R_DATE
            LEFT JOIN [DWH].[dbo].DIC_PRODUCT_BUDGET PB With(nolock) ON RPB.ID_PRODUCT_BUDGET=PB.ID_PRODUCT_BUDGET
--расход
            LEFT JOIN [DWH].[dbo].[FCT_REQUEST] fc With(nolock) on fc.[ID_REQUEST] = dR.ID_REQUEST

where 1=1 

and isnull(dr.SIGN_DELETED,0) =0
and year(dr.[DATE_OP]) in ('2021', '2022','2023')
and month(dr.[DATE_OP]) = @month
and org.full_NAME like '%Коммерческий департамент\%'
and org.name_2 not like '%МАП%'
and src.name_source like '%Веб%'
and dr.[FEATURES_5] not in ('estr2022' , 'ege_rf', 'leadform_CZ_747828', 'leadform_CZN_SYN_PP_713311','proftest')

 

'''

leads_eff = pd.read_sql_query(leads_eff, conn)
leads = pd.read_sql_query(leads, conn)
leads['День'] = leads['Дата соединения на ОП'].dt.day

baza_planov = pd.read_excel('C:\\Users\\ADavydovskiy\\scripts\\Выгрузки\\База планов 2020.xlsx', sheet_name='Планы_2020')
baza_planov = baza_planov.query('Менеджер != "Вакансия"')

leads = leads.query('День <= (@d-1)')
leads_pivot = leads.pivot_table(index='СП', columns='Год', values='id', aggfunc='count')

baza_planov_m = baza_planov.query('Месяц == @month')
baza_planov_m_pivot = baza_planov_m.pivot_table(index='КЦ', columns='Год', values='Менеджер', aggfunc='count')

baza_planov_pivot = baza_planov.pivot_table(index='Месяц.1', columns='Год', values='Менеджер', aggfunc='count')



wb = load_workbook('C:\\Users\\ADavydovskiy\\scripts\\Отчёты\\Распределение лидов на МП.xlsx')

sheet = wb.get_sheet_by_name('Сводник_1')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(leads_pivot, index=True, header=True):
    sheet.append(r)
    

sheet = wb.get_sheet_by_name('Сводник_2')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(baza_planov_m_pivot, index=True, header=True):
    sheet.append(r)
    
    
sheet = wb.get_sheet_by_name('Сводник_3')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(baza_planov_pivot, index=True, header=True):
    sheet.append(r)
    
    
sheet = wb.get_sheet_by_name('Сводник_4')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(leads_eff, index=True, header=True):
    sheet.append(r)

wb.save('C:\\Users\\ADavydovskiy\\scripts\\Отчёты\\Распределение лидов на МП.xlsx')