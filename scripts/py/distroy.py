import pandas as pd
import numpy as np
import pyodbc
from datetime import datetime, timedelta

import openpyxl as ox
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from tqdm import tqdm

import warnings

warnings.simplefilter("ignore")

print('Введите имя пользователя:')
user = input()
print('Введите количество дней в месяце:')
x = input()
x = int(x)
print('Введите наименование месяца в нижнем регистре (ПРИМЕР: январь):')
month = input()
print('Введите год:')
year = input()

data_кц_1 = pd.read_excel("C:\\Users\\{}\\Desktop\\График работы КД.xlsx".format(user),
                          sheet_name='кц_1_{}_{}'.format(month, year))
data_кц_2 = pd.read_excel("C:\\Users\\{}\\Desktop\\График работы КД.xlsx".format(user),
                          sheet_name='кц_2_{}_{}'.format(month, year))
data_кц_3 = pd.read_excel("C:\\Users\\{}\\Desktop\\График работы КД.xlsx".format(user),
                          sheet_name='кц_3_{}_{}'.format(month, year))
data_кц_4 = pd.read_excel("C:\\Users\\{}\\Desktop\\График работы КД.xlsx".format(user),
                          sheet_name='кц_4_{}_{}'.format(month, year))

name_sheet = (data_кц_1, data_кц_2, data_кц_3, data_кц_4)

data = []
sheet_num = (1, 2, 3, 4)
name_headers = ('name_1', 'name_2', 'name_3', 'name_4', 'name_5')

# -------------------------------------------------------------

conn = pyodbc.connect(r'Driver={SQL Server};Server=MSK1-BIDB01;Database=DWH;Trusted_Connection=yes;')
cursor = conn.cursor()

employees = f'''

SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED

SELECT [EMPLOYEES]
      ,[LOG]
      ,[SP]

      ,CASE

       WHEN KC like '%4%' THEN 'КЦ 4'
       WHEN KC like '%3%' THEN 'КЦ 3'
       WHEN KC like '%центр 3%' THEN 'КЦ 3'
         WHEN KC like '%продаж 18%' THEN 'ОП 18'
             
       ELSE KC
  END
  as "KC"

         ,CASE

       WHEN [KC] like '%4%' THEN SUBSTRING([ОП],CHARINDEX(' ',[ОП])+1,LEN([ОП])) + ' ВР'
       WHEN ([ОП] in ('ОП 9')) and ([GP] in ('.1')) THEN [ОП] + [GP]
       WHEN ([ОП] in ('ОП 2')) and ([GP] in ('.1')) THEN 'ОП 8'
       WHEN ([ОП] in ('ОП 10')) and ([GP] in ('.2')) THEN [ОП] + [GP]
       WHEN ([ОП] in ('ОП 5')) and ([GP] in ('.1')) THEN [ОП] + [GP]
       WHEN ([ОП] in ('ОП 14')) and ([GP] in ('.1')) THEN [ОП] + [GP]
       WHEN ([ОП] in ('3 ЯР')) and ([GP] in ('.1')) THEN '3.1 ЯР'
         WHEN ([KC] in ('Отдел прямых продаж'))  THEN 'ОПП'
           WHEN [KC] like '%продаж 18%'  THEN 'ОП 18'  
       ELSE [ОП]
  END
  as "ОП"
      
               ,[PHONE]
      ,[STATUS]
      ,[ID_EMPL]
      ,[ID_ORG]

  FROM [DWH].[dbo].[KHTS_EMPL]
  where 
  --KC like '%4%' 
   [SP] in ('КД')


'''

empl = pd.read_sql_query(employees, conn)

# -------------------------------------------------------------


header = list(range(1, x + 1))

h = list(map(str, header))

mass_header = list(range(1, x + 1))

headers = ["МП"]
headers.extend(mass_header)

td_col = ['МП']
td_col.extend(mass_header)

# -------------------------------------------------------------

number = ('1:', '2:', '3:', '4:', '5:', '6:', '7:', '8:', '9:',
          '1-', '2-', '3-', '4-', '5-', '6-', '7-', '8-', '9-',
          '10', '11', '12', '13', '14', '15', '16', '17', '18',
          '18', '19', '20', '21', '21', '22', '23', '24')

numbers = ('01', '02', '03', '04', '05', '06',
           '07', '08', '09', '10', '11', '12',
           '13', '14', '15', '16', '17', '18',
           '19', '20', '21', '22', '23', '00')

# -------------------------------------------------------------

for index in name_sheet:
    index.columns = headers
    data.append(index)

data = pd.concat(data)


for i in tqdm(headers):
    for row in data[i]:
        if str(row)[:2] in number:
            data.loc[data[i] == row, [i]] = str(row)[:2]
pass


for index in tqdm(headers):
    data[index] = data[index].str.lower()
pass


names = {0: 'name_1', 1: 'name_2', 2: 'name_3', 3: 'name_4', 4: 'name_5'}

data_names = data['МП'].str.split(' ', expand=True).rename(columns=names).fillna('')
data_names = data_names[['name_1', 'name_2', 'name_3', 'name_4']]

data[['name_1', 'name_2', 'name_3', 'name_4']] = data_names[['name_1', 'name_2', 'name_3', 'name_4']]

for index in tqdm(headers):
    data[index] = data[index].str.replace(' ', '')
    data[index] = data[index].str.replace(':', '')
    data[index] = data[index].str.replace('-', '')
pass

for index in tqdm(header):
    data.loc[data[index] == '07001600', [index]] = '7'
    data.loc[data[index] == '08001700', [index]] = '8'
    data.loc[data[index] == '09002100', [index]] = '9'
    data.loc[data[index] == '09001800', [index]] = '9'
    data.loc[data[index] == '09002000', [index]] = '9'
pass

for index in tqdm(headers[1:]):
    data.loc[~data[index].isin(h), [index]] = np.NaN
pass


data['МП'] = data['name_1'] + " " + data['name_2']  # + " " + data['name_3'] + " " + data['name_4']

data = data[td_col]

data = data.fillna('пропуск')
data = data.query('МП != "пропуск"')

# data = data[headers]

for index in tqdm(headers):
    data.loc[data[index] == 'пропуск', [index]] = np.NaN
pass


td = pd.read_excel('C:\\Users\\{}\\Desktop\\Трудовая_дисциплина.xlsx'.format(user), sheet_name='График работы')

td.columns = td.iloc[0]
td = td[1:]

col = []
col_2 = td.columns

s = '(п)'

for index in tqdm(col_2):
    if s in index:
        col.append(index)
pass

col = col[:x]

td['МП'] = td[col_2[:1]]

td[mass_header] = td[col]

td = td[td_col]

td_headers = td.columns

for i in tqdm(td_headers):
    for row in td[i]:
        if str(row)[:2] in numbers:
            td.loc[td[i] == row, [i]] = str(row)[:2] + '.' + str(row)[3:5]
pass


for index in tqdm(td_headers[1:]):
    for row in td[index]:
        try:
            float(td.loc[td[index] == row, [index]].iloc[0, 0])
        except:
            td.loc[td[index] == row, [index]] = np.NaN
pass

for index in tqdm(td_headers):
    td[index] = td[index].str.lower()
pass

td_names = td['МП'].str.split(' ', expand=True).rename(columns=names).fillna('')
td_names = td_names[['name_1', 'name_2', 'name_3', 'name_4']]

td[['name_1', 'name_2', 'name_3', 'name_4']] = td_names[['name_1', 'name_2', 'name_3', 'name_4']]

td['МП'] = td['name_1'] + " " + td['name_2']  # + " " + td['name_3'] + " " + td['name_4']

td = td[td_col]

td = td.query('МП != " "')


def search_partial_text(src, dst):
    dst_buf = dst
    result = 0
    for char in src:
        if char in dst_buf:
            dst_buf = dst_buf.replace(char, '', 1)
            result += 1
    r1 = int(result / len(src) * 100)
    r2 = int(result / len(dst) * 100)
    return r1 if r1 < r2 else r2


замены = []



for i in tqdm(data['МП']):

    s1 = i
    t9_2 = []
    t9_k = []
    one = data.loc[data['МП'] == i, ['МП']].iloc[0, 0]

    for index in td['МП']:
        s2 = index
        two = td.loc[td['МП'] == index, ['МП']].iloc[0, 0]
        search_partial = search_partial_text(s1, s2)

        if search_partial >= 90:
            t9_2.append(two)
            t9_k.append(search_partial)

    t9 = pd.DataFrame({'2': t9_2,
                       'k': t9_k})

    if len(t9) > 1:

        tr = t9.sort_values(by='k', ascending=False).iloc[0, 1]
        fl = t9.sort_values(by='k', ascending=False).iloc[1, 1]
        mp = t9.sort_values(by='k', ascending=False).iloc[0, 0]

        if tr != fl:
            data.loc[data['МП'] == s1, ['МП']] = mp
            if s1 != mp:
                замены.append({s1: mp})

    elif len(t9) == 1:

        mp = t9.iloc[0, 0]
        data.loc[data['МП'] == s1, ['МП']] = mp
        if s1 != mp:
            замены.append({s1: mp})
pass

print('')
print("Произведённые замены фамилий:")
print('')
print(замены)

all = data.merge(td, how='left', on='МП')

mp = []
delay = []

for row in tqdm(all['МП']):
    mp.append(row)
    delay.append(0)
    for index in td_headers[1:]:
        if float(all.loc[all['МП'] == row, ['{}_x'.format(index)]].iloc[0, 0]) >= 0:
            x = float(all.loc[all['МП'] == row, ['{}_x'.format(index)]].iloc[0, 0])
        else:
            x = 0
        if float(all.loc[all['МП'] == row, ['{}_y'.format(index)]].iloc[0, 0]) >= 0:
            y = float(all.loc[all['МП'] == row, ['{}_y'.format(index)]].iloc[0, 0])
        else:
            y = 0
        if y > x:
            delay[-1] += 1
pass

опездалы = pd.DataFrame(
    {'МП': mp,
     'Количество нарушений за месяц': delay
     })

опездалы = опездалы.sort_values(by='Количество нарушений за месяц', ascending=False)
опездалы = опездалы.drop_duplicates(subset=['МП'], keep='first')

# -------------------------------------------------------------

empl = empl[['EMPLOYEES', 'KC', 'ОП']]
empl.columns = ['МП', 'КЦ', 'ОП']

empl['МП'] = empl['МП'].replace('ё', "е")

empl['МП'] = empl['МП'].str.lower()

empl_names = empl['МП'].str.split(' ', expand=True).rename(columns=names).fillna('')
empl_names = empl_names[['name_1', 'name_2', 'name_3', 'name_4']]

empl[['name_1', 'name_2', 'name_3', 'name_4']] = empl_names[['name_1', 'name_2', 'name_3', 'name_4']]

empl['МП'] = empl['name_1'] + " " + empl['name_2']

empl = empl[['МП', 'КЦ', 'ОП']]

опездалы = опездалы.merge(empl, how='left', on='МП')

# -------------------------------------------------------------

print('')
print("ПРЕДПРОСМОТР... Файл имортируется в Excel :)")
print('')
print(опездалы.head(50))
print('')

wb = load_workbook('C:\\Users\\{}\\Desktop\\Опоздания.xlsx'.format(user))

sheet = wb.get_sheet_by_name('Sheet_1')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(опездалы, index=False, header=True):
    sheet.append(r)

sheet = wb.get_sheet_by_name('Отчёт')

sheet['A1'] = ("Отчётный период: " + month + " " + year)
sheet['A2'] = ("Дата формирования: " + datetime.today().strftime('%d.%m.%Y'))


wb.save('C:\\Users\\{}\\Desktop\\Опоздания.xlsx'.format(user))

print("")
print("Готово!")