import os

import pandas as pd
import numpy as np
import openpyxl as ox

from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

from datetime import datetime, timedelta

import warnings
warnings.filterwarnings("ignore")

import string

import pyodbc

print("Артём с особой старательностью протягивает лиды, пожалуйста, подождите...")

data = pd.read_excel('C:\\Users\\ADavydovskiy\\scripts\\Выгрузки\\БАЗА КД ОБЩАЯ 2018.xlsx', sheet_name='2018')


conn = pyodbc.connect(r'Driver={SQL Server};Server=MSK1-BIDB01;Database=DWH;Trusted_Connection=yes;')
cursor = conn.cursor()


data = data[['ID сделки', 'ленд']]


data = data.tail(1000)
data = data[data['ленд'] != np.nan]


dt = data['ID сделки'].astype('str')

dt = [f", '{el}'" for el in data['ID сделки']]


dt = list(dt)

s = ''.join(dt)

data = data['ID сделки']


request_string_leads = f'''

SELECT DISTINCT
	 --даты
	  CONVERT(varchar(10), COALESCE(D.DATE_CREATE, ASS_AR.Data, R.DATE_CREATE),104)  as "Дата создания" -- DATE_CREATE,
	  --,CONVERT(varchar(10), D.DATE_MODIFY, 104) as "Дата изменения"
	 -- ,CONVERT(varchar(10), D.CLOSE_DATE, 104)  as "Дата завершения"
	  --,R.[DATE_OP]
	  --маркетинговые метки
	 -- ,d.[UTM_CAMPAIGN]  as "Кампания"
	 ,src.name_source as "Источник 1"
	  ,R.FEATURES_1 "Источник"
      ,R.FEATURES_5  "Ленд"
	 -- ,R.[DATE_CREATE]
	 -- --клиент
	 -- ,D.name  as "Название сделки" --NAME_DEAL
	 -- ,concat(cc.[FAMILY]+' ',cc.[NAME]+' ',cc.[SURNAME]) as "Контакт"
	 -- ,cc.BIRTHDATE "Дата рождения"
	 -- ,isnull(PhHome.PHONE,'') "Телефон дом."
	 -- ,isnull(PhMob.PHONE,'') "Телефон моб."
	 -- ,isnull(PhJob.PHONE,'') "Телефон раб."
	 -- ,isnull(PhOst.PHONE,'') "Телефон проч."
	 -- ,coalesce(cc.EMAIL_HOME,cc.EMAIL_WORK) as "e-mail"
	 -- --продукт
	 -- ,cp.[NAME] "Продукт сделки"
	 -- ,dgp.NAME as "Группа продуктов"
	 -- ,PB.[Name] as "Продукт бюджета"
	 -- ,CAST(ASS_P.[QUANTITY] AS INT) AS "Количество продукта"
 -- ,CAST(D.AMOUNT / COUNT(*) OVER (PARTITION BY D.ID_DEAL)  as decimal(33,0)) "Сумма"
	 -- ,CAST(D.AMOUNT as decimal(33,0)) as "СуммаСделки"	      
	 -- ,IIF(isnull(D.SIGN_MAT_CAP,0)=1, 'Да', '') "Материнский капитал в сделке"
	 -- --вероятность+прослушка
	-- ,D.[PROBABILITY] "Вероятность"
	 -- ,replace(replace(isnull(utsD.[AUDITION_ROP],'0'),'0','нет'),'1','да') as "Прослушано РОП"
	 -- --статус
	--,sd.NAME as "Стадия сделки"
	 -- --ответственный	 
	 -- ,concat(e.[LAST_NAME]+' ',e.[NAME]+' ',e.[SECOND_NAME]) as "Ответственный"
	--  ,org.full_NAME "Структурное подразделение"
	 -- ,D.ID_CFO_VISUAL  as "Визуал"
	 -- --инвойсы
	 --  --,CONVERT(varchar(10), D_I.[DATE_CREATE],104) as "Дата создания инвойса"
	 --  ,CONVERT(varchar(10), D_I.[DATE_PAY],104) as "Дата оплаты"
  --     ,CAST(D_I.[AMOUNT] as decimal(33,0)) as "Сумма инвойса"
	 ----  ,S_I.[NAME] as "Статус инвойса"
	 --  --дела
		--,CONVERT(varchar(10), CAST(CA.[CREATED] as date),104) as "Дата создания дела"
		--,CONVERT(varchar(10), CAST(CA.[START_TIME] as date),104) as "Дата дела"
	 --   ,CAST(CA.[START_TIME] as time(0)) as "Время дела"
		--,CA.[COMPLETED] as "Статус дела"
		--,utsD.[CODE_LABEL_BASE_KD]
	 --  --id
	  ,D.CODE as "id deal crm"
	  ,R.[CODE] as "id lead crm"
	  --,R.[ID_REQUEST] as "id lead base"
	  --,RPB.[ID_PRODUCT_BUDGET] as "id продукт бюджета"	  
	  --,cc.ID_CLIENT_CRM as "id контакта"
	  --,ASS_D_I.[ID_INVOICE] as "id инвойса base"
	  --,D_I.[CODE] as "id инвойса crm"
	  


FROM [DWH].[dbo].DIC_DEAL D
		  LEFT JOIN [DWH].[dbo].[DIC_EMPLOYEES] E ON D.ID_EMPLOYEES = E.ID_EMPLOYEES
		  LEFT JOIN [DWH].[dbo].[ASS_EMPLOYEE_AND_ORGSTRUCTURE] ASS_OS ON E.ID_EMPLOYEES = ASS_OS.ID_EMPLOYEES
		  LEFT JOIN [DWH].[dbo].[v_DIC_ORGSTRUCTURE] org on org.ID_ORGSTRUCTURE = ASS_OS.ID_ORGSTRUCTURE
		 
		  --подсоединение заявок
		  LEFT JOIN [DWH].[dbo].ASS_REQUEST_DEAL ASS_RD ON ASS_RD.ID_DEAL = D.ID_DEAL
		  LEFT JOIN [DWH].[dbo].[DIC_REQUEST] R ON ASS_RD.ID_REQUEST = R.ID_REQUEST

		   -- источник
			left join [DWH].[stage].[CRM_b_uts_crm_lead] ul on ul.value_id = R.CODE
			left join [DWH].[stage].[CRM_b_crm_lead] l on l.ID = R.CODE
			left join (SELECT [NAME] as name_source, [STATUS_ID]
						   FROM [DWH].[stage].[CRM_b_crm_status]
						  where [ENTITY_ID] = 'SOURCE') src on src.[STATUS_ID] = l.[SOURCE_ID]
			left join [DWH].[dbo].[DIC_REQUEST_FORM] RF on RF.ID_REQUEST = R.ID_REQUEST
		    --объявления
		  LEFT JOIN [DWH].[dbo].ASS_AD_REQUEST ASS_AR ON ASS_AR.[ID_TYPE_ASS_AD_REQUEST] = 0   --Связь с учетом всех коректировок
										 AND ASS_AR.ID_REQUEST = R.ID_REQUEST   
		  LEFT JOIN [DWH].[dbo].DIC_AD AD ON AD.ID_AD = ASS_AR.ID_AD
		  LEFT JOIN [DWH].[dbo].DIC_AD_GROUP AG ON AD.ID_AD_GROUP = AG.ID_AD_GROUP
		  LEFT JOIN [DWH].[dbo].DIC_AD_CAMPAIGN AC ON AG.ID_AD_CAMPAIGN = AC.ID_AD_CAMPAIGN
   		  --подсоединение инвойсов к сделке
		  LEFT JOIN [DWH].[dbo].ASS_DEAL_INVOICE ASS_D_I ON ASS_D_I.ID_DEAL = D.ID_DEAL
		  LEFT JOIN [DWH].[dbo].DIC_INVOICE D_I ON ASS_D_I.ID_INVOICE = D_I.ID_INVOICE
		  LEFT JOIN [DWH].[dbo].DIC_STATUS_INVOICE S_I ON S_I.ID_STATUS_INVOICE = D_I.ID_STATUS_INVOICE
		  --каталог продуктов сделки
		  LEFT JOIN [DWH].[dbo].ASS_DEAL_CATALOG_PRODUCT ASS_P ON D.ID_DEAL = ASS_P.ID_DEAL
		  LEFT JOIN [DWH].[dbo].[DIC_CATALOG_PRODUCT] cp on cp.ID_CATALOG_PRODUCT = ASS_P.ID_CATALOG_PRODUCT
		  -- статус сделки
		  left join [DWH].[dbo].[DIC_STATUS_DEAL] sd on sd.ID_STATUS_DEAL = d.ID_STATUS_DEAL
		  --contact
		  left join [DWH].[dbo].[DIC_CLIENT_CRM] cc on cc.[ID_CLIENT_CRM] = d.[ID_CLIENT_CRM]
		  left join [DWH].[dbo].[DIC_CLIENT_CRM_PHONE] PhHome ON PhHome.[ID_CLIENT_CRM]=d.[ID_CLIENT_CRM] AND PhHome.[ID_TYPE_CLIENT_CRM_PHONE]=1	--	Домашний
		  left join [DWH].[dbo].[DIC_CLIENT_CRM_PHONE] PhMob ON PhMob.[ID_CLIENT_CRM]=d.[ID_CLIENT_CRM] AND PhMob.[ID_TYPE_CLIENT_CRM_PHONE]=2		--Мобильный
		  left join [DWH].[dbo].[DIC_CLIENT_CRM_PHONE] PhJob ON PhJob.[ID_CLIENT_CRM]=d.[ID_CLIENT_CRM] AND PhJob.[ID_TYPE_CLIENT_CRM_PHONE]=5		--Рабочий
		  left join [DWH].[dbo].[DIC_CLIENT_CRM_PHONE] PhOst ON PhOst.[ID_CLIENT_CRM]=d.[ID_CLIENT_CRM] AND PhOst.[ID_TYPE_CLIENT_CRM_PHONE]=3		--Прочий
		  -- group product deal
		  left join [DWH].[dbo].[DIC_GROUP_PRODUCT_DEAL] dgp on dgp.[ID_GROUP_PRODUCT_DEAL] = d.[GROUP_PRODUCT_ID]
		  -- расширение свойств сделок
		  left join [DWH].[dbo].[DIC_DEAL_STAT] utsD ON D.[ID_DEAL]=utsD.[ID_DEAL]
		  -- продукт бюджета
			LEFT JOIN [DWH].[dbo].ASS_REQUEST_PRODUCT_BUDGET RPB With(nolock) ON R.ID_REQUEST=RPB.ID_REQUEST and R.DATE_CREATE= RPB.R_DATE
			LEFT JOIN [DWH].[dbo].DIC_PRODUCT_BUDGET PB With(nolock) ON RPB.ID_PRODUCT_BUDGET=PB.ID_PRODUCT_BUDGET
			--дела
			LEFT JOIN [DWH].[stage].[CRM_b_crm_act] CA ON CA.[OWNER_ID] = D.[CODE]
		  
  
WHERE
  D.ID_DEAL <> - 1
	--and  cast(COALESCE(D.DATE_CREATE, ASS_AR.Data, R.DATE_CREATE)  as date) between '2020-06-01' and '2020-06-04' -- фильтр по дате создания
	--and  D.DATE_CREATE  between '2020-05-01' and '2020-06-08' -- фильтр по дате завершения 
	--and  cc.BIRTHDATE  between '1990-01-01' and '1999-12-31' -- фильтр по  дате рождения
	--КАК ВАРИАНТ  and 	 YEAR(cc.BIRTHDATE)  between 1990 AND 1999 -- фильтр по  ГОДУ даты рождения	
--	and R.[DATE_CREATE]  between '2020-05-01' and '2020-06-08' -- фильтр по дате завершения 
--and org.full_NAME like '%Коммерческий департамент\%'
--and dgp.NAME like '%ШБ%'
--and sd.NAME like '%eУспешная%'
and D.CODE in ('5101506'{s})

'''


leads = pd.read_sql_query(request_string_leads, conn)


leads = leads[['id deal crm', 'Дата создания', 'Источник 1', 'Источник', 'Ленд', 'id lead crm']]
leads[['id deal crm', 'id lead crm']] = leads[['id deal crm', 'id lead crm']].astype(int)


leads['Источник'] = leads['Источник'].fillna('0')
leads_2 = leads.query('Источник == "0"')
leads = leads.query('Источник != "0"')


leads_2['Источник'] = leads_2['Источник 1']
leads_2['Источник'] = leads_2['Источник'].map({'Аккаунтинг/Accounting': 'свой контакт', 
                                               'Входящий звонок / Incoming call': 'входящий звонок', 
                                               'Веб-сайт / Website':'органик', 
                                               'Импорт / Import':'импорт'})

leads = leads.append(leads_2, ignore_index=True)


leads['Ленд'] = leads['Ленд'].fillna('0')
leads_2 = leads.query('Ленд == "0"')
leads = leads.query('Ленд != "0"')


leads_2['Ленд'] = leads_2['Источник 1']
leads_2['Ленд'] = leads_2['Ленд'].map({'Аккаунтинг/Accounting': 'свой контакт', 
                                        'Входящий звонок / Incoming call': 'входящий звонок',  
                                        'Импорт / Import':'импорт'})

leads = leads.append(leads_2, ignore_index=True)


data = data.reset_index()


wb = load_workbook('C:\\Users\\ADavydovskiy\\scripts\\Отчёты\\leads.xlsx')

sheet = wb.get_sheet_by_name('Sheet_1')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(data, index=False, header=True):
    sheet.append(r)

sheet = wb.get_sheet_by_name('Sheet_2')
sheet.delete_rows(1, sheet.max_row)    

for r in dataframe_to_rows(leads, index=False, header=True):
    sheet.append(r)


wb.save('C:\\Users\\ADavydovskiy\\scripts\\Отчёты\\leads.xlsx')