import pandas as pd
import numpy as np
import time
from tqdm import tqdm
import openpyxl as ox
import re
import os
import pywhatkit

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains

import warnings
warnings.filterwarnings("ignore")

print("Артём снова выгружает планирование, пожалуйста, подождите...")


driver = webdriver.Chrome(executable_path = 'C:\\Users\\ADavydovskiy\\scripts\\YandexDriver\\yandexdriver.exe')
x = driver.get('https://corp.synergy.ru/crm/reports.synergy/plan_oplat/')
elem = Select(driver.find_element(By.ID, 'period'))
elem.select_by_value('cday')

rec = driver.find_element(By.ID, 'rec').click()

prob = driver.find_element(By.XPATH, 
                           '/html/body/table/tbody/tr[2]/td/table/tbody/tr[1]/td[2]/table/tbody/tr[2]/td/div/div[2]/div/div[2]/form/div[6]/input[2]'
                          ).click()


stage = Select(driver.find_element(By.ID, 'stage'))
stage.select_by_value('В работе')

elem = driver.find_element(By.XPATH, '//*[@id="dept_chosen"]/a/span').click()

element = driver.find_element(By.XPATH, '//*[@id="dept_chosen"]/div/div/input')
ActionChains(driver) \
    .send_keys_to_element(element, "(КМВ)", Keys.ENTER)\
    .perform()


    
find = driver.find_element(By.XPATH, 
                           '//*[@id="workarea-content"]/div/div[2]/form/div[14]/input'
                          ).click()
time.sleep(20)

tab = driver.find_element(By.XPATH, '//*[@id="mytable"]')    
tab_html=tab.get_attribute('outerHTML')
tab_dfs=pd.read_html(tab_html)
df = tab_dfs[0]
y = df['Сумма за день'][1]
df['Сумма за день'] = df['Сумма за день'].replace(y, np.nan)
df['Сумма за день'] = df['Сумма за день'][-2:].fillna(y)

driver.close()

#-----------------------------------------------------------------------

wb = load_workbook(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Планирование на день РОП.xlsx')

sheet = wb.get_sheet_by_name('1')
sheet.delete_rows(1, sheet.max_row)

for r in dataframe_to_rows(df, index=False, header=True):
    sheet.append(r)


wb.save(r'\\synergy.local\Documents\19.Группа мониторинга и сопровождения сделок\01.Отчеты\Планирование на день РОП.xlsx')


print("")
print("Готово!")
time.sleep(5)